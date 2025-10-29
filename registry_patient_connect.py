# -*- coding: utf-8 -*-
"""
(ปรับปรุงจาก registry_patient_connect.py — แก้ strike-through logic & ปรับสไตล์ตาราง)
"""
import os, sys, json, argparse, csv, base64, secrets, hashlib, unicodedata, re
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Set
from datetime import datetime, timedelta, time as dtime, date
from concurrent.futures import ThreadPoolExecutor

import requests
from requests.adapters import HTTPAdapter, Retry

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import QSettings, QUrl, QLocale
from PySide6.QtGui import QIcon, QPixmap, QPainter, QLinearGradient
from PySide6.QtWebSockets import QWebSocket
from PySide6.QtWidgets import QDialog

from icd10_catalog import (
    add_custom_entry,
    diagnosis_suggestions,
    get_custom_entries,
    get_diagnoses,
    get_operations,
)

try:
    from rapidfuzz import fuzz, process  # type: ignore

    _HAS_RAPIDFUZZ = True
except Exception:  # pragma: no cover - optional dependency
    _HAS_RAPIDFUZZ = False

# ---------------------- Modern theme ----------------------
def apply_modern_theme(widget: QtWidgets.QWidget):
    widget.setStyleSheet("""
    QWidget{font-family:'Segoe UI','Inter','Noto Sans',system-ui;font-size:11pt;color:#0f172a;background:#f4f6fb;}
    QLineEdit, QDateEdit, QTimeEdit, QComboBox, QPlainTextEdit{
        padding:8px 12px;border-radius:12px;border:1px solid #e6eaf2;background:#fff;
    }
    QLineEdit:hover, QDateEdit:hover, QTimeEdit:hover, QComboBox:hover, QPlainTextEdit:hover{border-color:#cfd8e6;}
    QLineEdit:focus, QDateEdit:focus, QTimeEdit:focus, QComboBox:focus, QPlainTextEdit:focus{border:1px solid #7aa2ff;background:#fff;}
    QPushButton{padding:9px 16px;border-radius:14px;border:1px solid #e7ecf4;background:#fff;font-weight:800;}
    QPushButton:hover{background:#f7faff;border-color:#d9e4fb;}
    QPushButton:pressed{background:#eef4ff;}
    QPushButton[variant="primary"]{background:#2563eb;border-color:#2563eb;color:#fff;}
    QPushButton[variant="primary"]:hover{background:#1d4ed8;}
    QPushButton[variant="danger"]{background:#ef4444;border-color:#ef4444;color:#fff;}
    QPushButton[variant="danger"]:hover{background:#dc2626;}
    QPushButton[variant="ghost"]{background:transparent;border-color:#e7ecf4;color:#0f172a;}
    QTableWidget, QTreeWidget{background:#fff;border:1px solid #e6e6ef;border-radius:12px;gridline-color:#e6e6ef;selection-background-color:#e0f2fe;}
    QTreeWidget::item{height:36px;}
    QTreeWidget::item:hover{ background: rgba(2,132,199,0.06); }
    QHeaderView::section{background:#f1f5f9;border-bottom:1px solid #e6eaf2;padding:10px 12px;font-weight:900;color:#0f172a;}
    QLabel[role='t']{ font-weight:900; font-size:16pt; letter-spacing:.2px; }
    QLabel[role='s']{ color:#64748b; font-size:10pt; }
    QLabel[role='h']{
        font-weight:900; font-size:12.5pt; color:#0f2167; letter-spacing:.2px;
    }
    QFrame#SectionHeader{
        background:#eef4ff; border:1px solid #dbeafe; border-radius:12px; padding:6px 10px;
    }
    QLabel[hint="1"]{color:#64748b;}
    QLabel[warn="1"]{color:#b91c1c;font-weight:700;}
    TabWidget::pane{border:0;}
    QTabBar::tab{
        padding:10px 16px;border-radius:12px;margin:4px;background:#e9eef8;font-weight:700;color:#0f172a;
    }
    QTabBar::tab:hover{ background:#eef3ff; }
    QTabBar::tab:selected{ background:#2563eb;color:#fff; }
    """)


def add_shadow(widget: QtWidgets.QWidget, blur=28, x=0, y=8, color="#24000000"):
    eff = QtWidgets.QGraphicsDropShadowEffect(widget)
    eff.setBlurRadius(blur); eff.setOffset(x, y); eff.setColor(QtGui.QColor(color))
    widget.setGraphicsEffect(eff)

class NoWheelComboBox(QtWidgets.QComboBox):
    """คอมโบที่ไม่ยอมให้เมาส์สกรอลล์เปลี่ยนค่า (กันเผลอเลื่อน)"""
    def wheelEvent(self, e: QtGui.QWheelEvent) -> None:
        e.ignore()  # ให้ scroll ที่ parent แทน
        return


def make_search_combo(options: list[str]) -> QtWidgets.QComboBox:
    cb = NoWheelComboBox()
    cb.setEditable(True)
    cb.addItems([""] + options)
    cb.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
    comp = QtWidgets.QCompleter(options)
    comp.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
    comp.setFilterMode(QtCore.Qt.MatchContains)
    cb.setCompleter(comp)
    cb.setMinimumWidth(180)
    return cb


def section_header(text: str) -> QtWidgets.QFrame:
    wrap = QtWidgets.QFrame()
    wrap.setObjectName("SectionHeader")
    h = QtWidgets.QHBoxLayout(wrap)
    h.setContentsMargins(10, 6, 10, 6)
    lab = QtWidgets.QLabel(text)
    lab.setProperty("role", "h")
    h.addWidget(lab)
    h.addStretch(1)
    return wrap

class Card(QtWidgets.QFrame):
    def __init__(self, title="", subtitle=""):
        super().__init__()
        self.setObjectName("Card")
        self.setStyleSheet("""
            QFrame#Card { background:#ffffff; border:1px solid #e6eaf2; border-radius:22px; }
            QLabel[role='t'] { font-weight:900; font-size:16pt; letter-spacing:0.2px; }
            QLabel[role='s'] { color:#64748b; font-size:10pt; }
        """)
        v = QtWidgets.QVBoxLayout(self)
        v.setContentsMargins(20,20,20,20); v.setSpacing(12)
        self.title_lbl = QtWidgets.QLabel(title); self.title_lbl.setProperty("role","t"); v.addWidget(self.title_lbl)
        if subtitle:
            s = QtWidgets.QLabel(subtitle); s.setProperty("role","s"); v.addWidget(s)
        self.body = QtWidgets.QWidget()
        self.grid = QtWidgets.QGridLayout(self.body)
        self.grid.setHorizontalSpacing(14); self.grid.setVerticalSpacing(12)
        v.addWidget(self.body); add_shadow(self)

class InfoBanner(QtWidgets.QFrame):
    def __init__(self, title: str = "", subtitle: str = "", variant: str = "blue", icon: str = "📁"):
        super().__init__()
        self.setObjectName("InfoBanner")
        self._variants = {
            "blue": {"bg": "#eaf6ff", "bd": "#cfe4ff", "accent": "#3b82f6"},
            "violet": {"bg": "#f4efff", "bd": "#e0d4ff", "accent": "#7c3aed"},
            "green": {"bg": "#eafaf3", "bd": "#cfeedd", "accent": "#10b981"},
        }
        pal = self._variants.get(variant, self._variants["blue"])

        self.setStyleSheet(
            f"""
        QFrame#InfoBanner {{
            background:{pal['bg']};
            border:1px solid {pal['bd']};
            border-radius:14px;
        }}
        QLabel[role='title']{{ font-weight:900; font-size:14pt; color:#0f172a; letter-spacing:.2px; }}
        QLabel[role='sub']  {{ color:#64748b; font-size:10pt; }}
        """
        )
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(12)

        accent = QtWidgets.QFrame()
        accent.setFixedWidth(8)
        accent.setStyleSheet(f"QFrame{{background:{pal['accent']}; border-radius:8px;}}")
        lay.addWidget(accent, 0)

        inner = QtWidgets.QVBoxLayout()
        inner.setSpacing(4)
        top = QtWidgets.QHBoxLayout()
        top.setSpacing(8)

        self.icon_lbl = QtWidgets.QLabel(icon)
        self.icon_lbl.setStyleSheet("font-size:16pt;")
        self.title_lbl = QtWidgets.QLabel(title)
        self.title_lbl.setProperty("role", "title")
        top.addWidget(self.icon_lbl, 0)
        top.addWidget(self.title_lbl, 1)
        top.addStretch(1)

        self.sub_lbl = QtWidgets.QLabel(subtitle)
        self.sub_lbl.setProperty("role", "sub")

        inner.addLayout(top)
        inner.addWidget(self.sub_lbl)
        lay.addLayout(inner, 1)

        add_shadow(self, blur=30, x=0, y=6, color="#2a000000")

    def set_title(self, text: str):
        self.title_lbl.setText(text or "")

    def set_subtitle(self, text: str):
        self.sub_lbl.setText(text or "")

    def set_icon(self, text: str):
        self.icon_lbl.setText(text or "📁")

# ---------------------- Diagnosis search helpers ----------------------
_COMBINE_RE = re.compile(r"[\u0300-\u036f\u0E31\u0E34-\u0E3A\u0E47-\u0E4E]")
_NON_ALNUM_RE = re.compile(r"[^0-9A-Za-z\u0E00-\u0E7F]+")


def normalize_text(text: str) -> str:
    if not text:
        return ""
    val = unicodedata.normalize("NFKD", text).lower()
    val = _COMBINE_RE.sub("", val)
    val = _NON_ALNUM_RE.sub(" ", val).strip()
    return " ".join(val.split())


class FastSearchIndex:
    def __init__(self, items: List[str], prefix_len: int = 3):
        self.items: List[str] = items or []
        self.norms: List[str] = [normalize_text(x) for x in self.items]
        self.prefix_len = max(1, int(prefix_len or 1))
        self.prefix_map: Dict[str, List[int]] = {}
        for idx, (raw, norm) in enumerate(zip(self.items, self.norms)):
            keys = set()
            left = (raw.split(" - ", 1)[0] if raw else "")
            for token in (raw, left, norm):
                token_norm = normalize_text(token)
                if not token_norm:
                    continue
                tokens = token_norm.split()
                if not tokens:
                    continue
                first = tokens[0]
                prefixes = {first[: self.prefix_len]}
                if len(first) >= 2:
                    prefixes.add(first[:2])
                for tk in tokens[:3]:
                    prefixes.add(tk[: self.prefix_len])
                keys.update({p for p in prefixes if p})
            for key in keys:
                self.prefix_map.setdefault(key, []).append(idx)

    def search(self, query: str, limit: int = 100) -> List[str]:
        if not self.items:
            return []
        limit = max(1, int(limit or 1))
        q = normalize_text(query)
        if not q:
            return self.items[:limit]

        q_parts = q.split()
        first_key = q_parts[0][: self.prefix_len]
        cand_idx = list(dict.fromkeys(self.prefix_map.get(first_key, [])))
        results: List[Tuple[int, float]] = []

        for idx in cand_idx:
            norm = self.norms[idx]
            if all(part in norm for part in q_parts):
                results.append((idx, 1.0))
                if len(results) >= limit:
                    break

        if len(results) < max(15, limit // 3):
            extra: List[Tuple[int, float]] = []
            cap = min(len(self.items), 5000)
            seen = {idx for idx, _ in results}
            for idx in range(cap):
                if idx in seen:
                    continue
                norm = self.norms[idx]
                if all(part in norm for part in q_parts):
                    extra.append((idx, 0.9))
                    if len(results) + len(extra) >= limit:
                        break
            for idx, score in extra:
                results.append((idx, score))

        if _HAS_RAPIDFUZZ and len(results) < limit:
            seen = {idx for idx, _ in results}
            corpus = {idx: norm for idx, norm in enumerate(self.norms) if idx not in seen}
            try:
                top_matches = process.extract(  # type: ignore[misc]
                    q,
                    corpus,
                    scorer=fuzz.partial_ratio,
                    score_cutoff=65,
                    limit=max(50, limit),
                )
            except Exception:
                top_matches = []
            for _, score, idx in top_matches:
                results.append((idx, score / 100.0))
                if len(results) >= limit:
                    break

        seen_idx = set()
        ordered: List[Tuple[int, float]] = []
        for idx, score in sorted(results, key=lambda item: (-item[1], len(self.items[item[0]]))):
            if idx in seen_idx:
                continue
            seen_idx.add(idx)
            ordered.append((idx, score))
            if len(ordered) >= limit:
                break

        return [self.items[idx] for idx, _ in ordered]

# ---------------------- Config ----------------------
DEFAULT_HOST = os.getenv("SURGIBOT_CLIENT_HOST", "127.0.0.1")
DEFAULT_PORT = int(os.getenv("SURGIBOT_CLIENT_PORT", "8088"))
DEFAULT_TOKEN = os.getenv("SURGIBOT_SECRET", "8HDYAANLgTyjbBK4JPGx1ooZbVC86_OMJ9uEXBm3EZTidUVyzhGiReaksGA0ites")

API_HEALTH = "/api/health"; API_LIST="/api/list"; API_LIST_FULL="/api/list_full"; API_WS="/api/ws"

STATUS_COLORS = {
    "รอผ่าตัด": "#fde047", "กำลังผ่าตัด": "#ef4444", "กำลังพักฟื้น": "#22c55e",
    "กำลังส่งกลับตึก": "#a855f7", "เลื่อนการผ่าตัด": "#64748b",
}
PULSE_STATUS = {"กำลังผ่าตัด","กำลังพักฟื้น","กำลังส่งกลับตึก"}
DEFAULT_OR_ROOMS = ["OR1","OR2","OR3","OR4","OR5","OR6","OR8"]

# --- สถานะจาก monitor ที่ใช้จับเวลา / auto-complete ---
STATUS_OP_START = "กำลังผ่าตัด"
STATUS_OP_END = "กำลังพักฟื้น"
STATUS_RETURNING = "กำลังส่งกลับตึก"

WARD_LIST = [
    "— กรุณาเลือก —",
    "หอผู้ป่วยอภิบาลสงฆ์",
    "หอผู้ป่วยพิเศษศัลยกรรม ชั้น 4",
    "หอผู้ป่วยศัลยกรรมกระดูกและข้อ",
    "หอผู้ป่วยศัลยกรรมหญิง",
    "หอผู้ป่วยศัลยกรรมชาย",
    "หอผู้ป่วยพิเศษอายุรกรรม ชั้น 5",
    "หอผู้ป่วยพิเศษอายุรกรรม ชั้น 4",
    "หอผู้ป่วยICU-MED",
    "หอผู้ป่วย ICU รวม",
    "หอผู้ป่วยอายุรกรรมหญิง",
    "หอผู้ป่วยอายุรกรรมชาย",
    "หอผู้ป่วยพิเศษรวมน้ำใจ",
    "หอผู้ป่วยหนักกุมารเวช",
    "หอผู้ป่วยหู ตา คอ จมูก",
    "หอผู้ป่วยกุมารเวช",
    "หอผู้ป่วยพิเศษสูติ-นรีเวช ชั้น 5",
    "หอผู้ป่วยพิเศษสูติ-นรีเวช ชั้น 4",
    "หอผู้ป่วยศัลยกรรมประสาทและสมอง",
    "หอผู้ป่วยสูติ-นรีเวช",
    "ห้องคลอด",
    "ห้องผ่าตัด",
    "แผนกอุบัติเหตุและฉุกเฉิน",
]

WARD_PLACEHOLDER = WARD_LIST[0]

SCRUB_NURSES = [
    "อรุณี", "ศิวดาติ์", "กัญญณัช", "ชัญญาภัค", "สุนทรี", "พิศมัย", "เทวัญ", "กันต์พงษ์",
    "ปนัฏฐา", "สุจิตรา", "ชัยยงค์", "สุภาวัลย์", "จันทจร", "วรรณิภา", "ณัฐพงษ์", "ตะวัน",
    "ปวีณา", "นิฤมล", "ปริญญา", "สยุมพร", "สุรสิทธ์", "บุศรินทร์", "ศิริกัญญา", "นราวัตน์",
    "บัณฑิตา", "วรรณวิสา", "ชลดา", "วรีสา",
]


ORG_NAME="ORNBH"; APP_SHARED="SurgiBotShared"; OR_KEY="schedule/or_rooms"; ENTRIES_KEY="schedule/entries"; SEQ_KEY="schedule/seq"
APP_SETTINGS="RegistryPatientConnect"
PDPA_ACK_KEY="pdpa/ack"
SECRET_SALT_KEY="sec/hn_salt"
FERNET_KEY="sec/fernet_key"  # เผื่อจะต่อยอดเข้ารหัสข้อความในอนาคต

DEPT_DOCTORS = {
    "Surgery | ศัลยกรรมทั่วไป": ["นพ.สุริยา คุณาชน","นพ.ธนวัฒน์ พันธุ์พรหม","พญ.สุภาภรณ์ พิณพาทย์","พญ.รัฐพร ตั้งเพียร","พญ.พิชัย สุวัฒนพูนลาภ"],
    "Orthopedics | ศัลยกรรมกระดูกและข้อ": ["นพ.ชัชพล องค์โฆษิต","นพ.ณัฐพงศ์ ศรีโพนทอง","นพ.อำนาจ อนันต์วัฒนกุล","นพ.อภิชาติ ลักษณะ","นพ.กฤษฎา อิ้งอำพร","นพ.วิษณุ ผูกพันธ์"],
    "Urology | ศัลยกรรมระบบทางเดินปัสสาวะ": ["พญ.สายฝน บรรณจิตร์"],
    "ENT | ศัลยกรรม โสต ศอ นาสิก": ["พญ.พิรุณยา แสนวันดี","พญ.สุทธิพร หมวดไธสง","นพ.วรวิช พลเวียงธรรม"],
    "Obstetrics-Gynecology | สูติ-นรีเวช": ["นพ.สุรจิตต์ นิมิตรวงษ์สกุล","พญ.ขวัญตา ทุนประเทือง","พญ.วัชราภรณ์ อนวัชชกุล","พญ.รุ่งฤดี โขมพัตร","พญ.ฐิติมน ชัยชนะทรัพย์"],
    "Ophthalmology | จักษุ": ["นพ.สราวุธ สารีย์","พญ.ดวิษา อังศรีประเสริฐ","พญ.สาวิตรี ถนอมวงศ์ไทย","พญ.สีกชมพู ตั้งสัตยาธิษฐาน","พญ.นันท์นภัส ชีวะเกรียงไกร"],
    "Maxillofacial | ศัลยกรรมขากรรไกร": ["นพ.ฉลองรัฐ เดชา","พญ.อรุณนภา คิสารัง"],
}

DEPT_KEY_MAP = {
    "Surgery": "Surgery",
    "ศัลยกรรมทั่วไป": "Surgery",
    "Surgery | ศัลยกรรมทั่วไป": "Surgery",
    "Orthopedics": "Orthopedics",
    "ศัลยกรรมกระดูกและข้อ": "Orthopedics",
    "Orthopedics | ศัลยกรรมกระดูกและข้อ": "Orthopedics",
    "Urology": "Urology",
    "ศัลยกรรมระบบทางเดินปัสสาวะ": "Urology",
    "ระบบทางเดินปัสสาวะ": "Urology",
    "Urology | ศัลยกรรมระบบทางเดินปัสสาวะ": "Urology",
    "ENT": "ENT",
    "ศัลยกรรม โสต ศอ นาสิก": "ENT",
    "หู คอ จมูก": "ENT",
    "ENT | ศัลยกรรม โสต ศอ นาสิก": "ENT",
    "Obstetrics-Gynecology": "OBGYN",
    "สูติ-นรีเวช": "OBGYN",
    "Obstetrics-Gynecology | สูติ-นรีเวช": "OBGYN",
    "Ophthalmology": "Ophthalmology",
    "จักษุ": "Ophthalmology",
    "Ophthalmology | จักษุ": "Ophthalmology",
    "Maxillofacial": "Maxillofacial",
    "ศัลยกรรมขากรรไกร": "Maxillofacial",
    "ศัลยกรรมช่องปากและใบหน้า": "Maxillofacial",
    "Maxillofacial | ศัลยกรรมขากรรไกร": "Maxillofacial",
}


def _dept_to_specialty_key(label: str) -> str:
    text = (label or "").strip()
    if not text:
        return ""
    parts = [p.strip() for p in text.split("|")]
    for part in parts + [text]:
        if part in DEPT_KEY_MAP:
            return DEPT_KEY_MAP[part]
        lowered = part.lower()
        for name, key in DEPT_KEY_MAP.items():
            if lowered == name.lower():
                return key
    lowered_all = text.lower()
    if "กระดูก" in lowered_all or "ortho" in lowered_all:
        return "Orthopedics"
    if "ปัสสาวะ" in lowered_all or "uro" in lowered_all:
        return "Urology"
    if "สูติ" in lowered_all or "ob" in lowered_all:
        return "OBGYN"
    if "โสต" in lowered_all or "ent" in lowered_all or "คอ" in lowered_all:
        return "ENT"
    if "จักษุ" in lowered_all or "oph" in lowered_all:
        return "Ophthalmology"
    if "ขากรรไกร" in lowered_all or "ช่องปาก" in lowered_all or "max" in lowered_all:
        return "Maxillofacial"
    return ""

class Toast(QtWidgets.QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            QFrame{background:#111827; color:#fff; border-radius:12px; padding:10px 14px;}
            QLabel{color:#fff;}
        """)
        add_shadow(self, blur=30, x=0, y=8, color="#40000000")
        self.lab = QtWidgets.QLabel("", self)
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(14,10,14,10)
        lay.addWidget(self.lab)
        self._anim: Optional[QtCore.QPropertyAnimation] = None
        self.hide()

    def show_toast(self, text: str, msec: int = 2200):
        self.lab.setText(text)
        if self.parent():
            p = self.parent()
            geo = p.geometry()
            self.adjustSize()
            w, h = self.width(), self.height()
            self.setGeometry(geo.width() - w - 24, 24, w, h)

        if self._anim is not None:
            try:
                self._anim.stop()
            except Exception:
                pass
            self._anim.deleteLater()
            self._anim = None

        self.setWindowOpacity(0.0)
        self.show()
        self.raise_()

        anim = QtCore.QPropertyAnimation(self, b"windowOpacity", self)
        anim.setDuration(msec)
        anim.setStartValue(0.0)
        anim.setKeyValueAt(0.1, 1.0)
        anim.setKeyValueAt(0.9, 1.0)
        anim.setEndValue(0.0)

        anim.finished.connect(self.hide)
        self._anim = anim
        anim.start()


class SweetAlert:
    @staticmethod
    def info(parent: QtWidgets.QWidget, title: str, text: str) -> None:
        QtWidgets.QMessageBox.information(parent, title, text)

    @staticmethod
    def success(
        parent: QtWidgets.QWidget,
        title: str,
        text: str,
        auto_close_msec: Optional[int] = None,
    ) -> None:
        box = QtWidgets.QMessageBox(parent)
        box.setIcon(QtWidgets.QMessageBox.Information)
        box.setWindowTitle(title)
        box.setText(f"✅ {text}")
        box.setStandardButtons(QtWidgets.QMessageBox.Ok)
        box.setDefaultButton(QtWidgets.QMessageBox.Ok)
        if auto_close_msec and auto_close_msec > 0:
            QtCore.QTimer.singleShot(auto_close_msec, box.accept)
        box.exec()

    @staticmethod
    def warning(parent: QtWidgets.QWidget, title: str, text: str) -> None:
        QtWidgets.QMessageBox.warning(parent, title, text)

    @staticmethod
    def confirm(parent: QtWidgets.QWidget, title: str, text: str) -> bool:
        box = QtWidgets.QMessageBox(parent)
        box.setIcon(QtWidgets.QMessageBox.Question)
        box.setWindowTitle(title)
        box.setText(text)
        box.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        box.setDefaultButton(QtWidgets.QMessageBox.No)
        return box.exec() == QtWidgets.QMessageBox.Yes

    @staticmethod
    def loading(parent: QtWidgets.QWidget, text: str = "กำลังดำเนินการ...") -> QtWidgets.QProgressDialog:
        dlg = QtWidgets.QProgressDialog(text, None, 0, 0, parent)
        dlg.setWindowModality(QtCore.Qt.ApplicationModal)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setWindowTitle("โปรดรอสักครู่")
        return dlg

class StatusChipWidget(QtWidgets.QWidget):
    def __init__(self, text:str, color:str, pulse:bool=False, parent=None):
        super().__init__(parent)
        self._text=text; self._color=color; self._pulse=pulse
        if pulse:
            self.eff = QtWidgets.QGraphicsOpacityEffect(self); self.setGraphicsEffect(self.eff)
            self.anim = QtCore.QPropertyAnimation(self.eff, b"opacity", self)
            self.anim.setDuration(1200); self.anim.setStartValue(0.5); self.anim.setEndValue(1.0)
            self.anim.setEasingCurve(QtCore.QEasingCurve.InOutQuad); self.anim.setLoopCount(-1); self.anim.start()
    def minimumSizeHint(self):
        fm = QtGui.QFontMetrics(self.font())
        w = fm.horizontalAdvance(self._text) + 22 + 16
        h = fm.height() + 10
        return QtCore.QSize(w, h)
    def paintEvent(self, e):
        p=QtGui.QPainter(self); p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        rect = self.rect().adjusted(2,2,-2,-2)
        bg = QtGui.QColor(self._color); bg.setAlpha(205)
        p.setPen(QtCore.Qt.NoPen); p.setBrush(bg)
        p.drawRoundedRect(rect, 10, 10)
        p.setPen(QtGui.QColor("#ffffff"))
        p.drawText(rect.adjusted(12,0,-8,0), QtCore.Qt.AlignVCenter|QtCore.Qt.AlignLeft, self._text)

class ScheduleEntry(QtCore.QObject):
    def __init__(
        self,
        or_room="",
        dt=None,
        time_str="",
        hn="",
        name="",
        age=0,
        dept="",
        doctor="",
        diags=None,
        ops=None,
        ward="",
        case_size="",
        queue=0,
        period="in",
        urgency="Elective",
        assist1="",
        assist2="",
        scrub="",
        circulate="",
        time_start="",
        time_end="",
        case_uid: str = "",
        version: int = 1,
        state: str = "scheduled",
        returning_started_at: str = "",
        returned_to_ward_at: str = "",
        postop_completed: bool = False,
    ):
        super().__init__()
        self.or_room = or_room
        self.date = dt or datetime.now().date()
        self.time = time_str
        self.hn = (hn or "").strip()
        self.name = (name or "").strip()
        self.age = int(age) if str(age).isdigit() else 0
        self.dept = dept
        self.doctor = doctor
        self.diags = diags or []
        self.ops = ops or []
        self.ward = (ward or "").strip()
        self.case_size = (case_size or "").strip()  # Minor/Major
        self.queue = int(queue) if str(queue).isdigit() else 0
        self.period = period  # "in" | "off"
        self.urgency = (urgency or "Elective")
        self.assist1 = assist1
        self.assist2 = assist2
        self.scrub = scrub
        self.circulate = circulate
        self.time_start = time_start
        self.time_end = time_end
        self.case_uid = case_uid or self._gen_case_uid()
        self.version = int(version or 1)
        self.state = state or "scheduled"
        self.returning_started_at = returning_started_at or ""
        self.returned_to_ward_at = returned_to_ward_at or ""
        self.postop_completed = bool(postop_completed)

    def _gen_case_uid(self) -> str:
        base = f"{self.or_room}|{self.hn}|{self.time}|{self.date}"
        return hashlib.sha1(base.encode("utf-8", "ignore")).hexdigest()

    def to_dict(self):
        return {
            "or": self.or_room,
            "date": str(self.date),
            "time": self.time,
            "hn": self.hn,
            "name": self.name,
            "age": self.age,
            "dept": self.dept,
            "doctor": self.doctor,
            "diags": self.diags,
            "ops": self.ops,
            "ward": self.ward,
            "case_size": self.case_size,
            "queue": self.queue,
            "period": self.period,
            "urgency": self.urgency,
            "assist1": self.assist1,
            "assist2": self.assist2,
            "scrub": self.scrub,
            "circulate": self.circulate,
            "time_start": self.time_start,
            "time_end": self.time_end,
            "case_uid": self.case_uid,
            "version": self.version,
            "state": self.state,
            "returning_started_at": self.returning_started_at,
            "returned_to_ward_at": self.returned_to_ward_at,
            "postop_completed": self.postop_completed,
        }

    @staticmethod
    def from_dict(d:dict):
        try:
            fromiso = datetime.fromisoformat(d.get("date")).date()
        except Exception:
            fromiso = datetime.now().date()
        return ScheduleEntry(
            d.get("or",""),
            fromiso,
            d.get("time",""),
            d.get("hn",""),
            d.get("name",""),
            d.get("age",0),
            d.get("dept",""),
            d.get("doctor",""),
            d.get("diags",[]) or [],
            d.get("ops",[]) or [],
            d.get("ward",""),
            d.get("case_size",""),
            d.get("queue",0),
            d.get("period","in"),
            d.get("urgency","Elective"),
            d.get("assist1",""),
            d.get("assist2",""),
            d.get("scrub",""),
            d.get("circulate",""),
            d.get("time_start",""),
            d.get("time_end",""),
            d.get("case_uid",""),
            d.get("version", 1),
            d.get("state","scheduled"),
            d.get("returning_started_at",""),
            d.get("returned_to_ward_at",""),
            bool(d.get("postop_completed", False)),
        )

    def uid(self)->str:
        return f"{self.or_room}|{self.hn}|{self.time}|{self.date}"

class SharedScheduleModel:
    def __init__(self):
        self.s = QSettings(ORG_NAME, APP_SHARED)
        self.entries = self._load(); self.or_rooms = self._load_or()
        if not self.s.contains(SEQ_KEY): self.s.setValue(SEQ_KEY, 0)
    def _load(self)->List[ScheduleEntry]:
        raw=self.s.value(ENTRIES_KEY, []); out=[]
        if isinstance(raw,list):
            for d in raw:
                if isinstance(d,dict): out.append(ScheduleEntry.from_dict(d))
        return out
    def _save(self):
        self.s.setValue(ENTRIES_KEY, [e.to_dict() for e in self.entries])
        self.s.setValue(SEQ_KEY, int(self.s.value(SEQ_KEY,0))+1); self.s.sync()
    def _load_or(self)->List[str]:
        lst=self.s.value(OR_KEY)
        if not isinstance(lst,list) or not lst: lst=DEFAULT_OR_ROOMS[:]; self.s.setValue(OR_KEY, lst)
        return [str(x) for x in lst]
    def set_or_rooms(self, rooms:List[str]):
        norm=[]
        for r in rooms:
            r=r.strip().upper()
            if r and r.startswith("OR") and r!="OR7" and r not in norm: norm.append(r)
        if not norm: norm=DEFAULT_OR_ROOMS[:]
        self.or_rooms=norm; self.s.setValue(OR_KEY, norm); self.s.setValue(SEQ_KEY, int(self.s.value(SEQ_KEY,0))+1); self.s.sync()
    def add(self, e:ScheduleEntry): self.entries.append(e); self._save()
    def update(self, idx:int, e:ScheduleEntry):
        if 0<=idx<len(self.entries): self.entries[idx]=e; self._save()
    def delete(self, idx:int):
        if 0<=idx<len(self.entries): self.entries.pop(idx); self._save()
    def seq(self)->int: return int(self.s.value(SEQ_KEY, 0))
    def all(self) -> List[ScheduleEntry]:
        return list(self.entries)
    def clear(self) -> int:
        removed = len(self.entries)
        if removed:
            self.entries.clear()
            self._save()
        return removed
    def remove_by_date(self, day: date) -> int:
        before = len(self.entries)
        self.entries = [e for e in self.entries if getattr(e, "date", None) != day]
        removed = before - len(self.entries)
        if removed:
            self._save()
        return removed
    def replace_all(self, items: List[ScheduleEntry]) -> None:
        self.entries = list(items)
        self._save()

class LocalDBLogger:
    def __init__(self, elective_path="schedule_elective.db", emergency_path="schedule_emergency.db"):
        import sqlite3
        self.sqlite3 = sqlite3
        self.conn_e = sqlite3.connect(elective_path)
        self.conn_x = sqlite3.connect(emergency_path)
        self._init(self.conn_e)
        self._init(self.conn_x)

    def _init(self, conn):
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS schedule(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                urgency TEXT,
                period TEXT,
                or_room TEXT,
                date TEXT,
                time TEXT,
                hn TEXT,
                name TEXT,
                age INTEGER,
                dept TEXT,
                doctor TEXT,
                diagnosis TEXT,
                operation TEXT,
                ward TEXT,
                queue INTEGER,
                time_start TEXT,
                time_end TEXT,
                case_size TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS surgery_events(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                case_uid TEXT,
                event TEXT,
                at TEXT,
                details TEXT
            )
            """
        )
        conn.commit()

    def append_entry(self, e: 'ScheduleEntry'):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = (
            ts,
            e.urgency,
            e.period,
            e.or_room,
            str(e.date),
            e.time,
            e.hn,
            e.name,
            int(e.age or 0),
            e.dept,
            e.doctor,
            " with ".join(e.diags),
            " with ".join(e.ops),
            e.ward,
            int(e.queue or 0),
            e.time_start or "",
            e.time_end or "",
            e.case_size or "",
        )

        conn = self.conn_x if str(e.urgency).lower() == "emergency" else self.conn_e
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO schedule(
                timestamp, urgency, period, or_room, date, time,
                hn, name, age, dept, doctor,
                diagnosis, operation, ward, queue,
                time_start, time_end, case_size
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            row,
        )
        conn.commit()

    def log_event(self, case_uid: str, event: str, details: Optional[dict] = None, emergency: bool = False):
        conn = self.conn_x if emergency else self.conn_e
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO surgery_events(case_uid,event,at,details) VALUES(?,?,?,?)",
            (case_uid, event, _now_iso(), json.dumps(details or {})),
        )
        conn.commit()

# ---------------------- Security helpers (salt & hash) ----------------------
def _app_settings() -> QSettings:
    # ใช้ settings ชุดเดียวกับตัวแอป เพื่อเก็บ salt/ack
    return QSettings(ORG_NAME, APP_SETTINGS)

def _get_or_create_secret(key: str, nbytes: int = 32) -> str:
    s = _app_settings()
    if not s.contains(key):
        # ใช้ urlsafe token เพื่อ copy/backup ได้ง่าย
        tok = secrets.token_urlsafe(nbytes)
        s.setValue(key, tok); s.sync()
    return str(s.value(key))

def hn_hash(hn: str) -> str:
    """De-identified hash ของ HN: SHA-256(HN + salt)"""
    salt = _get_or_create_secret(SECRET_SALT_KEY, 32)
    x = (str(hn) + salt).encode("utf-8", "ignore")
    return hashlib.sha256(x).hexdigest()

# (พื้นที่ต่อยอด: ถ้าต้องการเข้ารหัสชื่อ/หมายเลข)
# from cryptography.fernet import Fernet
# def _fernet() -> Fernet:
#     key = _get_or_create_secret(FERNET_KEY, 32)
#     # Fernet key ต้องเป็น base64 32 bytes → แปลงให้เป็น 32 bytes แล้ว b64
#     k = hashlib.sha256(key.encode()).digest()
#     return Fernet(base64.urlsafe_b64encode(k))
# def enc(txt:str)->str: return _fernet().encrypt(txt.encode()).decode()
# def dec(tok:str)->str: return _fernet().decrypt(tok.encode()).decode()

# ---------------------- PDPA / Consent ----------------------
class PDPANoticeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("แจ้งเตือน PDPA / ข้อกำกับการใช้ข้อมูล")
        self.setModal(True)
        lay = QtWidgets.QVBoxLayout(self)
        text = QtWidgets.QTextEdit(self)
        text.setReadOnly(True)
        text.setMinimumHeight(220)
        text.setStyleSheet("QTextEdit{background:#fff;border:1px solid #e6eaf2;border-radius:12px;padding:10px;}")
        text.setText(
            "วัตถุประสงค์การใช้ข้อมูล:\n"
            "- ใช้เพื่อการลงทะเบียน/บริหารจัดการคิวผ่าตัด และสื่อสารการทำงานในห้องผ่าตัด\n"
            "- ใช้สถิติภาพรวมแบบไม่ระบุตัวตน (de-identified) เพื่อปรับปรุงคุณภาพบริการ (QI)\n\n"
            "การคุ้มครองข้อมูลส่วนบุคคล (PDPA):\n"
            "- เก็บเท่าที่จำเป็น (data minimization)\n"
            "- มีปุ่ม Export แบบไม่ระบุตัวตน (แฮช HN) สำหรับงานวิเคราะห์\n"
            "- ห้ามส่งออก/ถ่ายโอนข้อมูลที่ระบุตัวบุคคลโดยไม่ได้รับอนุญาต\n"
            "- การบันทึก Log จะไม่เก็บข้อมูลที่ระบุตัวบุคคลโดยไม่จำเป็น\n\n"
            "การดำเนินการต่อถือว่าท่านเข้าใจและยอมรับตามข้างต้น"
        )
        chk = QtWidgets.QCheckBox("ฉันอ่านและยอมรับการใช้ข้อมูลตาม PDPA แล้ว")
        btn = QtWidgets.QPushButton("ตกลง"); btn.setProperty("variant","primary"); btn.setEnabled(False)
        chk.toggled.connect(lambda b: btn.setEnabled(b))
        btn.clicked.connect(self.accept)
        lay.addWidget(text); lay.addWidget(chk); lay.addWidget(btn)

def _fmt_td(td: timedelta) -> str:
    total = int(abs(td.total_seconds())); h = total // 3600; m = (total % 3600) // 60; s = total % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def _parse_iso(ts: str):
    if not ts: return None
    try: return datetime.fromisoformat(ts.replace("Z",""))
    except Exception: return None

def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

def _is_postop_complete_entry(e: "ScheduleEntry") -> bool:
    if not (e.time_start and e.time_end):
        return False
    try:
        hh1, mm1 = e.time_start.split(":")
        hh2, mm2 = e.time_end.split(":")
        t1 = int(hh1) * 60 + int(mm1)
        t2 = int(hh2) * 60 + int(mm2)
        if t2 < t1:
            return False
    except Exception:
        return False
    if not (e.scrub or e.circulate or e.assist1 or e.assist2):
        return False
    if not (e.ops or e.diags):
        return False
    return True

def _app_icon() -> QIcon:
    pm=QtGui.QPixmap(64,64); pm.fill(QtCore.Qt.transparent)
    pa=QPainter(pm); pa.setRenderHint(QtGui.QPainter.Antialiasing,True)
    gr=QLinearGradient(0,0,64,64)
    gr.setColorAt(0,"#d9ecff"); gr.setColorAt(.55,"#e1f5ff"); gr.setColorAt(1,"#e6fff5")
    pa.setBrush(gr); pa.setPen(QtCore.Qt.NoPen); pa.drawEllipse(6,6,52,52)
    pa.setBrush(QtGui.QColor("#0f172a"))
    pa.drawEllipse(26,20,12,12)
    pa.end()
    return QIcon(pm)


def _load_app_icon() -> QIcon:
    p = Path("assets/app.ico")
    if p.exists():
        ico = QIcon(str(p))
        if not ico.isNull():
            return ico
    return _app_icon()

def _now_period(dt_val: datetime) -> str:
    start = dtime(8,30); end = dtime(16,30)
    return "in" if (start <= dt_val.time() < end) else "off"

def _period_label(code: str) -> str: return "ในเวลาราชการ" if code=="in" else "นอกเวลาราชการ"

class ClientHTTP:
    def __init__(self, host=DEFAULT_HOST, port=DEFAULT_PORT, token=DEFAULT_TOKEN, timeout=1.2):
        self.base, self.token, self.timeout = f"http://{host}:{port}", token, timeout
        self.s = requests.Session()
        self.s.mount("http://", HTTPAdapter(max_retries=Retry(
            total=3, connect=2, read=2, backoff_factor=0.35,
            status_forcelist=(429,500,502,503,504),
            allowed_methods=frozenset(["GET","POST"])
        )))
    def health(self):
        r=self.s.get(self.base+API_HEALTH, timeout=self.timeout, headers={"Accept":"application/json"})
        r.raise_for_status(); return r.json()
    def list_items(self):
        try:
            r=self.s.get(f"{self.base}{API_LIST_FULL}?token={self.token}",timeout=self.timeout,headers={"Accept":"application/json"})
            if r.status_code==200: return self._wrap(r.json())
        except Exception: pass
        try:
            r=self.s.get(self.base+API_LIST,timeout=self.timeout,headers={"Accept":"application/json"})
            if r.status_code==200: return self._wrap(r.json())
        except Exception: pass
        return {"items":[]}
    @staticmethod
    def _wrap(d):
        if isinstance(d,list): return {"items":d}
        if isinstance(d,dict):
            for k in ("items","data","table","rows","list"):
                if k in d and isinstance(d[k],list): return {"items":d[k]}
            for v in d.values():
                if isinstance(v,list): return {"items":v}
            return d
        return {"items":[]}

def extract_rows(payload):
    if isinstance(payload,list): src=payload
    elif isinstance(payload,dict):
        for k in ("items","data","table","rows","list"):
            if k in payload and isinstance(payload[k],list): src=payload[k]; break
        else:
            src=next((v for v in payload.values() if isinstance(v, list)), [])
    else: src=[]
    rows=[]
    for i,it in enumerate(src,1):
        if not isinstance(it,dict): continue
        hn=str(it.get("hn_full") or it.get("hn") or "").strip()
        pid=str(it.get("patient_id") or it.get("pid") or it.get("queue_id") or "")
        if not pid:
            orr=str(it.get("or") or it.get("or_room") or ""); q=str(it.get("queue") or it.get("q") or "")
            if orr and q: pid=f"{orr}-{q}"
        status=str(it.get("status") or "")
        ts=(it.get("timestamp") or it.get("ts") or it.get("updated_at") or it.get("created_at") or it.get("time"))
        eta=it.get("eta_minutes", it.get("eta", it.get("eta_min")))
        if isinstance(eta,str) and eta.isdigit(): eta=int(eta)
        elif not isinstance(eta,int): eta=None
        rows.append({"id": hn if hn else i, "hn_full": hn or None, "patient_id": pid, "status": status, "timestamp": ts, "eta_minutes": eta})
    return rows

class QueueSelectWidget(QtWidgets.QWidget):
    changed = QtCore.Signal(int)
    def __init__(self, value:int=0, parent=None):
        super().__init__(parent)
        h = QtWidgets.QHBoxLayout(self); h.setContentsMargins(2, 0, 2, 0); h.setSpacing(6)
        self.combo = QtWidgets.QComboBox()
        self.combo.addItem("ตามเวลา", 0)
        for i in range(1, 10): self.combo.addItem(str(i), i)
        self.combo.setStyleSheet("""
            QComboBox{
                padding:6px 10px;border-radius:10px;border:1px solid #e6eaf2;background:#ffffff;
                min-width: 110px; font-weight:700;
            }
            QComboBox:hover{border-color:#cfd8e6;}
            QComboBox:focus{border:1px solid #7aa2ff;background:#ffffff;}
        """)
        idx = 0
        for i in range(self.combo.count()):
            if int(self.combo.itemData(i)) == int(value): idx = i; break
        self.combo.setCurrentIndex(idx)
        self.combo.currentIndexChanged.connect(self._emit_changed)
        h.addWidget(self.combo, 1)
        self.setMinimumWidth(120)
        self.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
    def _emit_changed(self, _i): self.changed.emit(int(self.combo.currentData() or 0))
    def value(self)->int: return int(self.combo.currentData() or 0)


# --------------------------- Fixed Excel import helpers ---------------------------
FIXED_MAPPING_TH = {
    "time": "สั่งผ่าตัดเวลา",
    "hn": "HN",
    "name": "ชื่อ",
    "age": "อายุ",
    "diags": "ICD Name",
    "ops": "ชื่อการผ่าตัด",
    "doctor": "แพทย์ผู้สั่ง",
    "ward": "Ward",
}

_re_hhmm = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*$")
_re_hhmmss = re.compile(r"^\s*(\d{1,2}):(\d{2}):(\d{2})\s*$")
_re_dt_hhmm = re.compile(r".*?(\d{1,2}):(\d{2})(?::\d{2})?\s*$")
_year_re = re.compile(r"(\d+)\s*ปี")


def _excel_time_to_hhmm(value: float) -> str:
    try:
        fraction = float(value)
    except Exception:
        return ""
    if not (0 <= fraction < 1.1):
        return ""
    total_seconds = int(round(fraction * 24 * 3600))
    hours = (total_seconds // 3600) % 24
    minutes = (total_seconds // 60) % 60
    return f"{hours:02d}:{minutes:02d}"


def parse_time_hhmm_or_tf(raw_value) -> str:
    """แปลงค่าจาก Excel/ข้อความให้เป็น HH:MM หรือ 'TF' หากไม่ทราบเวลา"""
    if raw_value is None:
        return "TF"

    if isinstance(raw_value, (int, float)):
        return _excel_time_to_hhmm(raw_value) or "TF"

    if isinstance(raw_value, datetime):
        return f"{raw_value.hour:02d}:{raw_value.minute:02d}"
    if isinstance(raw_value, dtime):
        return f"{raw_value.hour:02d}:{raw_value.minute:02d}"

    text = str(raw_value).strip()
    if not text:
        return "TF"

    match = _re_hhmmss.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    match = _re_hhmm.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    match = _re_dt_hhmm.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    return "TF"


def parse_age_years(txt: str) -> int:
    """ดึงเฉพาะจำนวน 'ปี' เช่น '23 ปี 4 เดือน 30 วัน' -> 23; '23' -> 23; อื่นๆ -> 0"""
    if not txt:
        return 0
    s = str(txt)
    m = _year_re.search(s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return 0
    try:
        return int(float(s))
    except Exception:
        return 0


def normalize_doctor(txt: str) -> str:
    """คงคำนำหน้าและจัดการ alias ชื่อแพทย์ให้สะอาด"""
    return normalize_doctor_name(txt)


def map_to_known_ward(src: str, known_wards: List[str]) -> str:
    """
    จับคู่ชื่อวอร์ดให้ตรงกับรายการที่แอปมีอยู่แล้ว
    กลไก: เทียบแบบ case-insensitive + ตัดช่องว่างเกิน + หา 'คีย์เวิร์ดหลัก'
    คุณสามารถปรับ synonyms ได้ตามชื่อในระบบจริง
    """

    s = " ".join((src or "").lower().split())
    synonyms = {
        "หูคอจมูก": ["หู คอ จมูก", "ent", "โสตศอนาสิก"],
    }

    for w in known_wards:
        if s == " ".join(w.lower().split()):
            return w

    for canonical, words in synonyms.items():
        for kw in words:
            if kw.replace(" ", "") in s.replace(" ", ""):
                for w in known_wards:
                    if canonical.replace(" ", "") in w.replace(" ", "").lower():
                        return w
                return src

    for w in known_wards:
        lowered = w.lower()
        if any(token and token in s for token in lowered.split()):
            return w

    return src


# ================== CONFIG: ตารางแพทย์ประจำ OR ==================
# ฟอร์แมต:
# WEEKLY_DOCTOR_OR_PLAN[weekday]["ORx"] = [
#   {"doctor": "ชื่อแพทย์", "when": "ALLDAY|AM|PM", "weeks": [1,2,3,4]},
#   {"doctor": ["ดร.ก", "ดร.ข"], ...}
# ]
# หมายเหตุ: Monday=0 ... Sunday=6
WEEKLY_DOCTOR_OR_PLAN: Dict[int, Dict[str, List[Dict[str, object]]]] = {
    0: {
        "OR1": [
            {"doctor": "นพ.สุริยา คุณาชน", "when": "ALLDAY", "weeks": [1]},
            {"doctor": "พญ.รัฐพร ตั้งเพียร", "when": "ALLDAY", "weeks": [2]},
            {"doctor": "พญ.พิชัย สุวัฒนพูนลาภ", "when": "ALLDAY", "weeks": [3]},
            {"doctor": "นพ.ธนวัฒน์ พันธุ์พรหม", "when": "ALLDAY", "weeks": [4]},
        ],
        "OR2": [],
        "OR3": [{"doctor": "พญ.พิรุณยา แสนวันดี", "when": "ALLDAY", "weeks": [1, 2, 3, 4]}],
        "OR5": [{"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]}],
        "OR6": [{"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]}],
        "OR8": [{"doctor": "EYE_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]}],
    },
    1: {
        "OR1": [
            {"doctor": "พญ.สายฝน บรรณจิตร์", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR2": [
            {"doctor": "นพ.ชัชพล องค์โฆษิต", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR3": [
            {"doctor": "พญ.สุภาภรณ์ พิณพาทย์", "when": "AM", "weeks": [1, 2, 3, 4]},
            {"doctor": "ทพญ.อรุณนภา คิสารัง", "when": "PM", "weeks": [1, 2, 3, 4]},
        ],
        "OR5": [
            {"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR6": [
            {"doctor": "นพ.พิชัย สุวัฒนพูนลาภ", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR8": [
            {"doctor": "EYE_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
    },
    2: {
        "OR1": [
            {"doctor": "พญ.สายฝน บรรณจิตร์", "when": "AM", "weeks": [1, 2, 3, 4]},
            {"doctor": "นพ.ชัชพล องค์โฆษิต", "when": "PM", "weeks": [1, 3]},
            {"doctor": "นพ.ณัฐพงศ์ ศรีโพนทอง", "when": "PM", "weeks": [2, 4]},
            {"doctor": "นพ.วิษณุ ผูกพันธ์", "when": "PM", "weeks": [2, 4]},
            {"doctor": "นพ.กฤษฎา อิ้งอำพร", "when": "PM", "weeks": [2, 4]},
        ],
        "OR2": [
            {"doctor": "นพ.วิษณุ ผูกพันธ์", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR3": [
            {"doctor": "CLOSED", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR5": [
            {"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR6": [
            {"doctor": "พญ.รัฐพร ตั้งเพียร", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR8": [
            {"doctor": "EYE_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
    },
    3: {
        "OR1": [],
        "OR2": [
            {"doctor": "นพ.อำนาจ อนันต์วัฒนกุล", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR3": [
            {"doctor": "นพ.วรวิช พลเวียงธรรม", "when": "AM", "weeks": [1, 2, 3, 4]},
            {"doctor": "ทพ.ฉลองรัฐ เดชา", "when": "PM", "weeks": [1, 2, 3, 4]},
        ],
        "OR5": [
            {"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR6": [
            {"doctor": "นพ.ธนวัฒน์ พันธุ์พรหม", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR8": [
            {"doctor": "EYE_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
    },
    4: {
        "OR1": [
            {"doctor": "พญ.สุภาภรณ์ พิณพาทย์", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR2": [
            {"doctor": "นพ.กฤษฎา อิ้งอำพร", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR3": [
            {"doctor": "พญ.สุทธิพร หมวดไธสง", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR5": [
            {"doctor": "OBGYN_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR6": [
            {"doctor": "CLOSED", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
        "OR8": [
            {"doctor": "EYE_ANY", "when": "ALLDAY", "weeks": [1, 2, 3, 4]},
        ],
    },
}

GROUPS: Dict[str, List[str]] = {
    # Surgery | ศัลยกรรมทั่วไป
    "SUR_ANY": [
        "นพ.สุริยา คุณาชน",
        "นพ.ธนวัฒน์ พันธุ์พรหม",
        "พญ.สุภาภรณ์ พิณพาทย์",
        "พญ.รัฐพร ตั้งเพียร",
        "พญ.พิชัย สุวัฒนพูนลาภ",
    ],

    # Orthopedics | ศัลยกรรมกระดูกและข้อ
    "ORTHO_ANY": [
        "นพ.ชัชพล องค์โฆษิต",
        "นพ.ณัฐพงศ์ ศรีโพนทอง",
        "นพ.อำนาจ อนันต์วัฒนกุล",
        "นพ.อภิชาติ ลักษณะ",
        "นพ.กฤษฎา อิ้งอำพร",
        "นพ.วิษณุ ผูกพันธ์",
    ],

    # Urology | ระบบทางเดินปัสสาวะ
    "URO_ANY": [
        "พญ.สายฝน บรรณจิตร์",
    ],

    # ENT | โสต ศอ นาสิก
    "ENT_ANY": [
        "พญ.พิรุณยา แสนวันดี",
        "พญ.สุทธิพร หมวดไธสง",
        "นพ.วรวิช พลเวียงธรรม",
    ],

    # Obstetrics-Gynecology | สูติ-นรีเวช
    "OBGYN_ANY": [
        "นพ.สุรจิตต์ นิมิตรวงษ์สกุล",
        "พญ.ขวัญตา ทุนประเทือง",
        "พญ.วัชราภรณ์ อนวัชชกุล",
        "พญ.รุ่งฤดี โขมพัตร",
        "พญ.ฐิติมน ชัยชนะทรัพย์",
    ],

    # Ophthalmology | จักษุ
    "EYE_ANY": [
        "นพ.สราวุธ สารีย์",
        "พญ.ดวิษา อังศรีประเสริฐ",
        "พญ.สาวิตรี ถนอมวงศ์ไทย",
        "พญ.สีกชมพู ตั้งสัตยาธิษฐาน",
        "พญ.นันท์นภัส ชีวะเกรียงไกร",
    ],

    # Maxillofacial | ศัลยกรรมขากรรไกร
    "MAXILO_ANY": [
        "ทพ.ฉลองรัฐ เดชา",
        "ทพญ.อรุณนภา คิสารัง",
    ],
}

DOCTOR_ALIASES: Dict[str, str] = {
    # ยกตัวอย่างสะกด/วรรคต่างกัน/พิมพ์ผิดที่เจอบ่อย
    "นพ.สุริยะ คุณาชน": "นพ.สุริยา คุณาชน",
    "นพ.สุริยา": "นพ.สุริยา คุณาชน",
    "นพ.ธนวัฒน์": "นพ.ธนวัฒน์ พันธุ์พรหม",
    "พญ.รัฐพร": "พญ.รัฐพร ตั้งเพียร",
    "พญ.พิชัย": "พญ.พิชัย สุวัฒนพูนลาภ",
    "พญ.พิริยา": "พญ.พิรุณยา แสนวันดี",
    "พญ.พิรุณยา": "พญ.พิรุณยา แสนวันดี",
    "พญ.สายฝน": "พญ.สายฝน บรรณจิตร์",
    "นพ.ชัชพล": "นพ.ชัชพล องค์โฆษิต",
    "นพ.ณัฐพงศ์": "นพ.ณัฐพงศ์ ศรีโพนทอง",
    "นพ.วิษณุ": "นพ.วิษณุ ผูกพันธ์",
    "นพ.กฤษฎา": "นพ.กฤษฎา อิ้งอำพร",
    "พญ.สุภาภรณ์": "พญ.สุภาภรณ์ พิณพาทย์",
    "พญ.สุทธิพร": "พญ.สุทธิพร หมวดไธสง",
    "พญ.สุภาภรณ์ พิณพาท": "พญ.สุภาภรณ์ พิณพาทย์",
    "พญ.พิชัย สุวัฒนพูนลาภ": "พญ.พิชัย สุวัฒนพูนลาภ",
    "นพ.วิษณุ ผูกพัน": "นพ.วิษณุ ผูกพันธ์",
    "ทพญ.อรุณนภา คิสาลัง": "ทพญ.อรุณนภา คิสารัง",
    "พญ.สีกชมพู ตั้งสัตยาธ": "พญ.สีกชมพู ตั้งสัตยาธิษฐาน",
}

DOCTOR_ALIASES.update(
    {
        "ทพญ.อรุณนภา": "ทพญ.อรุณนภา คิสารัง",
        "ทพ.ฉลองรัฐ": "ทพ.ฉลองรัฐ เดชา",
        "นพ.วรวิช": "นพ.วรวิช พลเวียงธรรม",
        "พญ.สุทธิพร": "พญ.สุทธิพร หมวดไธสง",
    }
)

TOKEN_DISPLAY_NAMES: Dict[str, str] = {
    "SUR_ANY": "ทีมศัลยกรรมทั่วไป",
    "ORTHO_ANY": "ทีมศัลยกรรมกระดูก",
    "URO_ANY": "ทีมระบบทางเดินปัสสาวะ",
    "ENT_ANY": "ทีมโสต ศอ นาสิก",
    "OBGYN_ANY": "ทีมสูติ-นรีเวช",
    "EYE_ANY": "ทีมจักษุ",
    "MAXILO_ANY": "ทีมศัลยกรรมขากรรไกร",
    "CLOSE": "ปิดห้อง",
    "CLOSED": "ปิดห้อง",
}

CLOSED_TOKENS = {"CLOSE", "CLOSED"}
# ================================================================


def normalize_doctor_name(name: str) -> str:
    s = " ".join(str(name or "").split())
    return DOCTOR_ALIASES.get(s, s)


GROUP_MEMBER_LOOKUP: Dict[str, Set[str]] = {
    token: {normalize_doctor_name(member) for member in members}
    for token, members in GROUPS.items()
}


def week_of_month(d: date) -> int:
    first = d.replace(day=1)
    return ((d.day + first.weekday() - 1) // 7) + 1


def time_to_period(hhmm_or_tf: str) -> str:
    if hhmm_or_tf == "TF":
        return "ANY"
    try:
        hour = int(hhmm_or_tf.split(":", 1)[0])
    except Exception:
        return "ANY"
    return "AM" if hour < 12 else "PM"


def doctor_in_group(doc: str, token: str) -> bool:
    normalized = normalize_doctor_name(doc)
    return normalized in GROUP_MEMBER_LOOKUP.get(token, set())


def match_doctor(token_or_name: str, doctor_name: str) -> bool:
    token = token_or_name
    if not token:
        return False
    if token in CLOSED_TOKENS:
        return False
    if token in GROUPS or token in {"SUR_ANY", "ORTHO_ANY", "ENT_ANY", "EYE_ANY", "MAXILO_ANY", "OBGYN_ANY"}:
        return doctor_in_group(doctor_name, token)
    return normalize_doctor_name(token) == normalize_doctor_name(doctor_name)


def doctor_service_token(doctor_name: str) -> str:
    normalized = normalize_doctor_name(doctor_name)
    for token, members in GROUP_MEMBER_LOOKUP.items():
        if normalized in members:
            return token
    return ""


def _rule_tokens(rule: Dict[str, object]) -> List[str]:
    doctor_token = rule.get("doctor")
    if isinstance(doctor_token, list):
        return [str(tok or "") for tok in doctor_token]
    return [str(doctor_token or "")]


def _rule_matches_service(rule: Dict[str, object], service_token: str) -> bool:
    if not service_token:
        return False
    members = GROUP_MEMBER_LOOKUP.get(service_token, set())
    if not members and service_token in GROUPS:
        members = {normalize_doctor_name(name) for name in GROUPS.get(service_token, [])}

    for token in _rule_tokens(rule):
        if not token or token in CLOSED_TOKENS:
            continue
        if token == service_token:
            return True
        if token in GROUP_MEMBER_LOOKUP and token == service_token:
            return True
        if normalize_doctor_name(token) in members:
            return True
    return False


def _describe_doctor_token(token: str) -> str:
    if not token:
        return ""
    label = TOKEN_DISPLAY_NAMES.get(token)
    if label:
        return label
    if token in GROUPS:
        # ไม่มีข้อความเฉพาะ ให้ใช้ชื่อกลุ่มแรกเป็นตัวแทน
        members = GROUPS.get(token, [])
        if members:
            return f"{TOKEN_DISPLAY_NAMES.get(token, '') or members[0]}"
    return normalize_doctor_name(token)


def describe_or_plan_label(case_date: date, or_room: str) -> str:
    if not or_room or or_room == "-":
        return ""

    weekday = case_date.weekday()
    plan = WEEKLY_DOCTOR_OR_PLAN.get(weekday, {})
    rules = plan.get(or_room, []) or []
    if not rules:
        return ""

    current_week = week_of_month(case_date)

    def label_for_rule(rule: Dict[str, object]) -> str:
        doctor_token = rule.get("doctor")
        if isinstance(doctor_token, list):
            tokens = [str(tok) for tok in doctor_token]
        else:
            tokens = [str(doctor_token or "")]
        parts: List[str] = []
        for tok in tokens:
            desc = _describe_doctor_token(tok)
            if desc:
                parts.append(desc)
        label = ", ".join(parts)
        when = (str(rule.get("when") or "ALLDAY").upper())
        if when == "AM":
            return f"เช้า: {label}" if label else "เช้า"
        if when == "PM":
            return f"บ่าย: {label}" if label else "บ่าย"
        if label:
            return label
        return ""

    filtered = [rule for rule in rules if current_week in rule.get("weeks", [1, 2, 3, 4, 5])]
    if not filtered:
        filtered = rules

    labels = [label_for_rule(rule) for rule in filtered]
    labels = [lbl for lbl in labels if lbl]
    return " • ".join(labels)


def pick_or_by_doctor(case_date: date, time_str: str, doctor_name: str) -> str:
    if not doctor_name:
        return ""

    weekday = case_date.weekday()
    plan = WEEKLY_DOCTOR_OR_PLAN.get(weekday, {})
    if not plan:
        return ""

    current_week = week_of_month(case_date)
    period = time_to_period(time_str)

    def iter_rules():
        for or_room, rules in plan.items():
            for rule in rules or []:
                yield or_room, rule

    for or_room, rule in iter_rules():
        doctor_token = rule.get("doctor")
        if not doctor_token:
            continue
        weeks = rule.get("weeks", [1, 2, 3, 4, 5])
        when = (rule.get("when") or "ALLDAY").upper()
        doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
        if current_week in weeks and (when == "ALLDAY" or period == "ANY" or period == when):
            if any(match_doctor(doc, doctor_name) for doc in doctors):
                return or_room

    if period == "ANY":
        for or_room, rule in iter_rules():
            doctor_token = rule.get("doctor")
            if not doctor_token:
                continue
            weeks = rule.get("weeks", [1, 2, 3, 4, 5])
            if current_week not in weeks:
                continue
            doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
            if any(tok in CLOSED_TOKENS for tok in doctors):
                continue
            if any(match_doctor(doc, doctor_name) for doc in doctors):
                return or_room

    for or_room, rule in iter_rules():
        doctor_token = rule.get("doctor")
        if not doctor_token:
            continue
        doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
        if any(tok in CLOSED_TOKENS for tok in doctors):
            continue
        # ตรงชื่อแพทย์แบบเฉพาะเจาะจงเท่านั้น (ไม่จับคู่ token ระดับแผนก)
        explicit_tokens = [tok for tok in doctors if tok not in GROUPS]
        if not explicit_tokens:
            continue
        if any(match_doctor(doc, doctor_name) for doc in explicit_tokens):
            return or_room

    service_token = doctor_service_token(doctor_name)
    if service_token:
        for or_room, rule in iter_rules():
            weeks = rule.get("weeks", [1, 2, 3, 4, 5])
            when = (rule.get("when") or "ALLDAY").upper()
            if current_week in weeks and (when == "ALLDAY" or period == "ANY" or period == when):
                if _rule_matches_service(rule, service_token):
                    return or_room

        for or_room, rule in iter_rules():
            if _rule_matches_service(rule, service_token):
                return or_room

    return ""


class Main(QtWidgets.QWidget):
    def __init__(self, host, port, token):
        super().__init__()
        self.cli = ClientHTTP(host, port, token)
        self.sched = SharedScheduleModel()
        self.db_logger = LocalDBLogger()
        self.ws=None; self.rows_cache=[]
        self.seq_seen = self.sched.seq()
        icon = _load_app_icon()
        self.setWindowIcon(icon)
        self.tray = QtWidgets.QSystemTrayIcon(icon, self); self.tray.show()

        self._last_status_by_hn: dict[str, str] = {}

        # form edit mode
        self._edit_idx: Optional[int] = None
        self._last_focus_uid: Optional[str] = None  # ใช้ไฮไลต์หลังบันทึก

        self.toast = Toast(self)
        self._current_specialty_key = ""
        self._last_snapshot: Optional[List[Dict[str, object]]] = None

        self._diag_base_catalog: List[str] = []
        self._diag_catalog_full: List[str] = []
        self._op_catalog_full: List[str] = []
        self._dx_index: Optional[FastSearchIndex] = None
        self._op_index: Optional[FastSearchIndex] = None
        self._search_executor = ThreadPoolExecutor(max_workers=1)
        self._search_timer = QtCore.QTimer(self)
        self._search_timer.setSingleShot(True)
        try:
            self._dx_search_limit = max(1, int(os.getenv("DX_SEARCH_LIMIT", "100")))
        except ValueError:
            self._dx_search_limit = 100
        try:
            debounce_ms = max(0, int(os.getenv("DX_SEARCH_DEBOUNCE_MS", "150")))
        except ValueError:
            debounce_ms = 150
        self._search_timer.setInterval(debounce_ms)
        self._search_timer.timeout.connect(self._on_diag_search_timeout)
        self._latest_diag_query = ""
        try:
            self._op_search_limit = max(1, int(os.getenv("OP_SEARCH_LIMIT", "100")))
        except ValueError:
            self._op_search_limit = 100
        try:
            op_debounce = max(0, int(os.getenv("OP_SEARCH_DEBOUNCE_MS", "150")))
        except ValueError:
            op_debounce = 150
        self._op_search_timer = QtCore.QTimer(self)
        self._op_search_timer.setSingleShot(True)
        self._op_search_timer.setInterval(op_debounce)
        self._op_search_timer.timeout.connect(self._on_op_search_timeout)
        self._latest_op_query = ""

        self.setWindowTitle("Registry Patient Connect — ORNBH")
        self.resize(1360, 900)
        apply_modern_theme(self)
        self._build_ui(); self._load_settings(); self._pdpa_gate(); self._start_timers()

    # ---------- UI ----------
    def _build_ui(self):
        outer = QtWidgets.QVBoxLayout(self); outer.setSpacing(12); outer.setContentsMargins(14,14,14,14)
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setStyleSheet("QTabWidget::pane{border:0;} QTabBar::tab{padding:10px 16px;border-radius:12px;margin:4px;background:#e9eef8;} QTabBar::tab:selected{background:#2563eb;color:#fff;}")
        outer.addWidget(self.tabs)

        # TAB 1 — ลงทะเบียน (ห่อด้วย ScrollArea เพื่อป้องกันคอนโทรลหด)
        tab1_inner = QtWidgets.QWidget()
        t1 = QtWidgets.QVBoxLayout(tab1_inner); t1.setSpacing(12); t1.setContentsMargins(0,0,0,0)
        t1_banner = InfoBanner(
            title="ลงทะเบียนผู้ป่วย (Schedule — Private)",
            subtitle="ข้อมูลเก็บในเครื่อง และแชร์ให้โปรแกรมหลักแบบเรียลไทม์",
            variant="blue",
            icon="📝",
        )
        t1.addWidget(t1_banner)
        form = Card("ลงทะเบียนผู้ป่วย (Schedule — Private)", "ข้อมูลเก็บในเครื่อง และแชร์ให้โปรแกรมหลักแบบเรียลไทม์")
        form.title_lbl.hide()
        g=form.grid; r=0
        g.setColumnStretch(0, 0); g.setColumnStretch(1, 2); g.setColumnStretch(2, 0); g.setColumnStretch(3, 1)
        g.setColumnStretch(4, 0); g.setColumnStretch(5, 2)
        g.addWidget(QtWidgets.QLabel("OR"), r,0)
        self.cb_or=QtWidgets.QComboBox(); self._refresh_or_cb(self.cb_or); self.cb_or.setMinimumWidth(140)
        g.addWidget(self.cb_or, r,1)
        self.btn_manage_or=QtWidgets.QPushButton("จัดการ OR"); self.btn_manage_or.setProperty("variant","ghost")
        g.addWidget(self.btn_manage_or, r,2,1,2)
        r+=1
        g.addWidget(QtWidgets.QLabel("ชื่อ-สกุล"), r,0)
        self.ent_name=QtWidgets.QLineEdit()
        g.addWidget(self.ent_name, r,1,1,3)
        g.addWidget(QtWidgets.QLabel("อายุ"), r,4)
        self.ent_age=QtWidgets.QLineEdit(); self.ent_age.setValidator(QtGui.QIntValidator(0,150,self))
        g.addWidget(self.ent_age,r,5)
        r+=1
        g.addWidget(QtWidgets.QLabel("HN"), r,0)
        self.ent_hn=QtWidgets.QLineEdit(); self.ent_hn.setMaxLength(9); self.ent_hn.setValidator(QtGui.QIntValidator(0,999999999,self))
        g.addWidget(self.ent_hn,r,1)
        g.addWidget(QtWidgets.QLabel("Ward"), r,2)
        self.cb_ward = QtWidgets.QComboBox(); self.cb_ward.setEditable(True)
        self.cb_ward.addItems(WARD_LIST)
        self.cb_ward.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        ward_options = [w for w in WARD_LIST if w and w != WARD_PLACEHOLDER]
        comp = QtWidgets.QCompleter(ward_options)
        comp.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        comp.setFilterMode(QtCore.Qt.MatchContains)
        self.cb_ward.setCompleter(comp)
        self.cb_ward.setCurrentIndex(0)
        self.cb_ward.setEditText(WARD_PLACEHOLDER)
        g.addWidget(self.cb_ward, r,3)
        g.addWidget(QtWidgets.QLabel("ขนาดเคส"), r,4)
        self.cb_case = NoWheelComboBox(); self.cb_case.addItems(["","Minor","Major"])
        self.cb_case.setMinimumWidth(120)
        g.addWidget(self.cb_case, r,5)
        r+=1
        g.addWidget(QtWidgets.QLabel("ความเร่งด่วน"), r,0)
        self.cb_urgency = NoWheelComboBox(); self.cb_urgency.addItems(["Elective","Emergency"])
        self.cb_urgency.setMinimumWidth(180)
        g.addWidget(self.cb_urgency, r,1)
        self.lbl_period_info = QtWidgets.QLabel("")
        self.lbl_period_info.setProperty("hint", "1")
        g.addWidget(self.lbl_period_info, r,2,1,4)
        r+=1
        g.addWidget(QtWidgets.QLabel("วันที่"), r,0)
        self.date=QtWidgets.QDateEdit(QtCore.QDate.currentDate()); self.date.setCalendarPopup(True); self.date.setDisplayFormat("dd/MM/yyyy"); self.date.setLocale(QLocale("en_US"))
        g.addWidget(self.date,r,1)
        g.addWidget(QtWidgets.QLabel("เวลา"), r,2)
        self.time=QtWidgets.QTimeEdit(QtCore.QTime.currentTime()); self.time.setDisplayFormat("HH:mm"); self.time.setLocale(QLocale("en_US"))
        g.addWidget(self.time,r,3)
        g.addWidget(QtWidgets.QLabel("แผนก"), r,4)
        self.cb_dept=QtWidgets.QComboBox(); self.cb_dept.addItems(["— เลือกแผนก —"] + list(DEPT_DOCTORS.keys()))
        g.addWidget(self.cb_dept,r,5)
        r+=1
        self.lbl_warn = QtWidgets.QLabel(""); self.lbl_warn.setProperty("warn","1")
        g.addWidget(self.lbl_warn, r,0,1,6)
        r+=1
        self.row_doctor_label = QtWidgets.QLabel("แพทย์ผู้ผ่าตัด"); g.addWidget(self.row_doctor_label, r,0)
        self.cb_doctor=QtWidgets.QComboBox(); g.addWidget(self.cb_doctor,r,1,1,5)
        r+=1

        g.addWidget(section_header("Diagnosis"), r,0,1,6)
        r+=1
        self.diag_adder = SearchSelectAdder("ค้นหา ICD-10 / ICD-10-TM...", suggestions=[])
        self.diag_adder.requestPersist.connect(self._on_diagnosis_persist_requested)
        if self.diag_adder.search_line:
            self.diag_adder.search_line.textChanged.connect(self._on_diag_query_changed)
        g.addWidget(self.diag_adder, r,0,1,6)
        r+=1

        g.addWidget(section_header("Operation"), r,0,1,6)
        r+=1
        self.op_adder = SearchSelectAdder("ค้นหา/เลือก Operation...", suggestions=[])
        self.op_adder.itemsChanged.connect(self._on_operations_changed)
        self.op_adder.requestPersist.connect(self._on_operation_persist_requested)
        if self.op_adder.search_line:
            self.op_adder.search_line.textChanged.connect(self._on_op_query_changed)
        g.addWidget(self.op_adder, r,0,1,6)
        r+=1

        g.addWidget(section_header("Scrub Nurse / ทีมพยาบาล"), r,0,1,6)
        r+=1
        row_n = QtWidgets.QHBoxLayout(); row_n.setSpacing(8)

        def _hint(txt: str) -> QtWidgets.QLabel:
            lab = QtWidgets.QLabel(txt)
            lab.setProperty("hint", "1")
            return lab

        self.cb_assist1 = make_search_combo(SCRUB_NURSES)
        self.cb_assist2 = make_search_combo(SCRUB_NURSES)
        self.cb_scrub = make_search_combo(SCRUB_NURSES)
        self.cb_circulate = make_search_combo(SCRUB_NURSES)

        row_n.addWidget(_hint("Assist 1")); row_n.addWidget(self.cb_assist1, 1)
        row_n.addWidget(_hint("Assist 2")); row_n.addWidget(self.cb_assist2, 1)
        row_n.addWidget(_hint("Scrub")); row_n.addWidget(self.cb_scrub, 1)
        row_n.addWidget(_hint("Circulate")); row_n.addWidget(self.cb_circulate, 1)
        g.addLayout(row_n, r,0,1,6)
        r+=1

        g.addWidget(section_header("เวลาเริ่ม–จบผ่าตัด (ใส่หรือไม่ใส่ก็ได้)"), r,0,1,6)
        r+=1
        row_t = QtWidgets.QHBoxLayout(); row_t.setSpacing(10)
        self.ck_time_start = QtWidgets.QCheckBox("ระบุเวลาเริ่ม")
        self.time_start = QtWidgets.QTimeEdit(QtCore.QTime.currentTime())
        self.time_start.setDisplayFormat("HH:mm")
        self.time_start.setEnabled(False)
        self.ck_time_end = QtWidgets.QCheckBox("ระบุเวลาจบ")
        self.time_end = QtWidgets.QTimeEdit(QtCore.QTime.currentTime())
        self.time_end.setDisplayFormat("HH:mm")
        self.time_end.setEnabled(False)

        self.ck_time_start.toggled.connect(lambda ch: self.time_start.setEnabled(ch))
        self.ck_time_end.toggled.connect(lambda ch: self.time_end.setEnabled(ch))

        row_t.addWidget(self.ck_time_start)
        row_t.addWidget(self.time_start)
        row_t.addSpacing(16)
        row_t.addWidget(self.ck_time_end)
        row_t.addWidget(self.time_end)
        row_t.addStretch(1)
        g.addLayout(row_t, r,0,1,6)
        r+=1

        self.btn_add=QtWidgets.QPushButton("➕ เพิ่ม"); self.btn_add.setProperty("variant","primary")
        self.btn_cancel_edit=QtWidgets.QPushButton("ยกเลิกแก้ไข"); self.btn_cancel_edit.setProperty("variant","ghost"); self.btn_cancel_edit.hide()
        self.btn_clear=QtWidgets.QPushButton("🧹 เคลียร์"); self.btn_clear.setProperty("variant","ghost")
        rowb=QtWidgets.QHBoxLayout(); rowb.setSpacing(10); rowb.addWidget(self.btn_add); rowb.addWidget(self.btn_cancel_edit); rowb.addWidget(self.btn_clear); rowb.addStretch(1)
        g.addLayout(rowb, r,0,1,6)
        r+=1

        t1.addWidget(form); t1.addStretch(1)
        tab1_scroll = QtWidgets.QScrollArea()
        tab1_scroll.setWidgetResizable(True)
        tab1_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab1_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        tab1_scroll.setWidget(tab1_inner)
        self.tabs.addTab(tab1_scroll, "ลงทะเบียนผู้ป่วย")

        # TAB 2 — Result Schedule
        tab2 = QtWidgets.QWidget(); t2 = QtWidgets.QVBoxLayout(tab2); t2.setSpacing(12)
        self.result_banner = InfoBanner("", "ห้องผ่าตัดโรงพยาบาลหนองบัวลำภู")
        t2.addWidget(self.result_banner)
        self.card_result = Card("ตารางการผ่าตัด ประจำวัน", "ห้องผ่าตัดโรงพยาบาลหนองบัวลำภู")
        self.card_result.title_lbl.hide()
        gr2 = self.card_result.grid
        self.tree2 = QtWidgets.QTreeWidget()
        # เพิ่มคอลัมน์ให้ครอบคลุมข้อมูลจากแท็บ 1 และเปิดสกรอลล์แนวนอน
        self.tree2.setColumnCount(19)
        self.tree2.setHeaderLabels([
            "ช่วงเวลา","OR/เวลา","HN","ชื่อ-สกุล","อายุ","Diagnosis","Operation","แพทย์",
            "Ward","ขนาดเคส","แผนก","Assist1","Assist2","Scrub","Circulate","เริ่ม","จบ","คิว","ประเภทเคส"
        ])
        # ไม่พับบรรทัดและไม่ตัดข้อความเป็น "..." เพื่อให้อ่านได้เต็มโดยเลื่อนแนวนอน
        self.tree2.setWordWrap(False)
        self.tree2.setTextElideMode(QtCore.Qt.ElideNone)
        # ต้องให้หัว OR ที่เป็น widget ปรับความสูงตามเนื้อหาได้ จึงไม่ใช้ uniform row height
        self.tree2.setUniformRowHeights(False)
        self.tree2.setAlternatingRowColors(True)
        # อนุญาตให้หัวกลุ่ม (เช่น OR1, OR2) พับเก็บ/ขยายได้ จึงเปิด child indicator
        self.tree2.setRootIsDecorated(True)
        self.tree2.setIndentation(12)
        self.tree2.setMouseTracking(True)
        # เปิดสกรอลล์บาร์แนวนอนเสมอเมื่อคอลัมน์กว้าง
        self.tree2.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.tree2.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.tree2.setStyleSheet("""
            /* ตัวตาราง */
            QTreeWidget{
                background:#ffffff;
                border:1px solid #dfe6f0;
                border-radius:12px;
                gridline-color:#e8edf5;
            }

            /* Header โปร่งใสเพื่อให้มุมบนโค้งจาก section แรก/สุดท้ายทำงาน */
            QHeaderView{
                background:transparent;
                border:none;
                margin:0;
                padding:0;
            }

            /* หัวคอลัมน์: โทนขาวฟ้าอ่อน ขอบชัด ตัวหนา */
            QHeaderView::section{
                background:#f6f9ff;
                color:#0f172a;
                font-weight:900;
                letter-spacing:.2px;
                padding:12px 14px;
                border-top:1px solid #dfe6f0;
                border-bottom:1px solid #dfe6f0;
                border-right:1px solid #dfe6f0;
            }

            /* มุมบนซ้าย/ขวาโค้ง */
            QHeaderView::section:first{
                border-top-left-radius:12px;
                border-left:1px solid #dfe6f0;
            }
            QHeaderView::section:last{
                border-top-right-radius:12px;
                border-right:1px solid #dfe6f0;
            }

            /* Hover/Pressed ลดเงาเล็กน้อย */
            QHeaderView::section:hover{
                background:#eef4ff;
            }
            QHeaderView::section:pressed{
                background:#e7efff;
            }

            /* ไอเท็มในตาราง */
            QTreeWidget::item{
                height:36px;
            }
            QTreeWidget::item:alternate{
                background:#fbfdff;
            }
            QTreeWidget::item:selected{
                background:rgba(37,99,235,0.12);
                border-radius:8px;
            }
            QTreeWidget::item:hover{
                background:rgba(2,132,199,0.06);
            }
        """)
        hdr=self.tree2.header(); hdr.setStretchLastSection(False)
        hdr.setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        hdr.setFixedHeight(42)
        # ให้คอลัมน์ยืดบางส่วน และเลื่อนแนวนอนได้เมื่อกว้างเกิน
        for i in (0,1,2,3,4,7,8,9,10,11,12,13,14,15,16,17,18):
            hdr.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)   # Diagnosis
        hdr.setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)   # Operation
        self.tree2.setColumnWidth(17, 160)
        self.tree2.setColumnWidth(18, 140)
        self.tree2.setColumnHidden(0, True)
        hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.tree2.setColumnWidth(0, 0)
        self.tree2.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tree2.customContextMenuRequested.connect(self._result_ctx_menu)
        gr2.addWidget(self.tree2,0,0,1,1)

        import_bar = QtWidgets.QHBoxLayout()
        import_bar.setContentsMargins(0, 6, 0, 0)
        import_bar.setSpacing(10)
        self.btn_import_excel = QtWidgets.QPushButton("📥 นำเข้าจาก Excel")
        self.btn_import_excel.setProperty("variant", "ghost")
        import_bar.addWidget(self.btn_import_excel, 0)
        self.btn_clear_board = QtWidgets.QPushButton("🧹 ล้างกระดาน")
        self.btn_clear_board.setProperty("variant", "destructive")
        import_bar.addWidget(self.btn_clear_board, 0)
        self.btn_undo_clear = QtWidgets.QPushButton("↩️ ย้อนกลับการล้าง")
        self.btn_undo_clear.setProperty("variant", "ghost")
        self.btn_undo_clear.setEnabled(False)
        import_bar.addWidget(self.btn_undo_clear, 0)
        import_bar.addStretch(1)
        gr2.addLayout(import_bar,1,0,1,1)
        gr2.setRowStretch(0, 1)
        gr2.setRowStretch(1, 0)
        t2.addWidget(self.card_result, 1)
        self.tabs.addTab(tab2, "Result Schedule")

        self._clear_shortcut = QtGui.QShortcut(QtGui.QKeySequence("Ctrl+Del"), self)
        self._clear_shortcut.activated.connect(self._on_clear_board_clicked)

        # TAB 3 — Monitor
        tab3 = QtWidgets.QWidget(); t3 = QtWidgets.QVBoxLayout(tab3); t3.setSpacing(12); t3.setContentsMargins(0,0,0,0)
        t3_banner = InfoBanner(
            title="Result (Monitor) — จากเซิร์ฟเวอร์",
            subtitle="",
            variant="violet",
            icon="🗓️",
        )
        t3.addWidget(t3_banner)
        server_bar = QtWidgets.QFrame(); server_bar.setStyleSheet("QFrame{background:#fff;border:1px solid #e6eaf2;border-radius:14px;padding:8px;}"); add_shadow(server_bar)
        hb = QtWidgets.QHBoxLayout(server_bar); hb.setContentsMargins(8,8,8,8)
        self.ent_host = QtWidgets.QLineEdit("127.0.0.1"); self.ent_host.setMaximumWidth(180); self.ent_host.setEchoMode(QtWidgets.QLineEdit.Password)
        self.ent_port = QtWidgets.QLineEdit(str(DEFAULT_PORT)); self.ent_port.setMaximumWidth(90)
        self.ent_token = QtWidgets.QLineEdit(DEFAULT_TOKEN); self.ent_token.setEchoMode(QtWidgets.QLineEdit.Password)
        self.btn_health = QtWidgets.QPushButton("Health"); self.btn_health.setProperty("variant","ghost"); self.btn_health.clicked.connect(self._on_health)
        for w,lbl in [(self.ent_host,"Host"),(self.ent_port,"Port"),(self.ent_token,"Token")]:
            box=QtWidgets.QHBoxLayout(); lab=QtWidgets.QLabel(lbl); lab.setProperty("hint","1"); box.addWidget(lab); box.addWidget(w); hb.addLayout(box)
        hb.addWidget(self.btn_health); hb.addStretch(1)
        self.status_chip = QtWidgets.QLabel("● Offline")
        self.status_chip.setStyleSheet("color:#ef4444;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")
        hb.addWidget(self.status_chip)
        t3.addWidget(server_bar)

        mon = Card("Monitor Realtime (จากเซิร์ฟเวอร์)", "สถานะสด (ดับเบิลคลิกเพื่อส่ง HN ไปลงทะเบียน/แก้ไข)")
        gm=mon.grid
        self.table = QtWidgets.QTableWidget(0,4); self.table.setHorizontalHeaderLabels(["ID","Patient ID","สถานะ","เวลา"])
        hdr2=self.table.horizontalHeader(); hdr2.setStretchLastSection(True)
        hdr2.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        hdr2.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(34)
        gm.addWidget(self.table,0,0,1,3)
        self.btn_refresh=QtWidgets.QPushButton("รีเฟรช"); self.btn_refresh.setProperty("variant","ghost")
        self.btn_export=QtWidgets.QPushButton("Export CSV"); self.btn_export.setProperty("variant","ghost")
        self.btn_export_deid=QtWidgets.QPushButton("Export De-Identified (CSV)"); self.btn_export_deid.setProperty("variant","ghost")
        gm.addWidget(self.btn_refresh,1,0)
        gm.addWidget(self.btn_export,1,1)
        gm.addWidget(self.btn_export_deid,1,2)
        gm.setColumnStretch(0,0); gm.setColumnStretch(1,0); gm.setColumnStretch(2,1)
        t3.addWidget(mon,1)
        self.tabs.addTab(tab3, "Monitor Realtime")

        # signals
        self.btn_refresh.clicked.connect(lambda: self._refresh(True))
        self.btn_export.clicked.connect(self._export_csv)
        self.btn_export_deid.clicked.connect(self._export_deid_csv)
        self.btn_import_excel.clicked.connect(self._on_import_excel)
        self.btn_clear_board.clicked.connect(self._on_clear_board_clicked)
        self.btn_undo_clear.clicked.connect(self._on_undo_clear_clicked)
        self.btn_manage_or.clicked.connect(self._manage_or)
        self.cb_dept.currentTextChanged.connect(self._on_dept_changed)
        self.btn_add.clicked.connect(self._on_add_or_update)
        self.btn_cancel_edit.clicked.connect(self._cancel_edit_mode)
        self.btn_clear.clicked.connect(self._clear_form)
        self.table.itemDoubleClicked.connect(self._on_monitor_double_click)
        self.tree2.itemDoubleClicked.connect(self._on_result_double_click)

        # default period info (auto-calculated)
        self._update_period_info()
        self.date.dateChanged.connect(lambda *_: self._update_period_info())
        self.time.timeChanged.connect(lambda *_: self._update_period_info())

        self._set_doctor_visibility(False)
        self._on_dept_changed(self.cb_dept.currentText())
        self._render_tree2()

    # ---------- settings / timers ----------
    def _current_specialty_key_safe(self) -> str:
        return (getattr(self, "_current_specialty_key", "") or "").strip()

    def _load_settings(self):
        self.cfg = QSettings(ORG_NAME, APP_SETTINGS)
        self.tabs.setCurrentIndex(0)
    def _save_settings(self): pass
    def closeEvent(self, e):
        try:
            if self.ws: self.ws.close()
        except Exception: pass
        try:
            self._search_executor.shutdown(wait=False)
        except Exception:
            pass
        super().closeEvent(e)

    def _start_timers(self):
        self._pull = QtCore.QTimer(self); self._pull.timeout.connect(lambda: self._refresh(True)); self._pull.start(3000)
        self._seq_timer = QtCore.QTimer(self); self._seq_timer.timeout.connect(self._check_seq); self._seq_timer.start(1000)
        QtCore.QTimer.singleShot(200, lambda: self._refresh(True))
        QtCore.QTimer.singleShot(600, self._start_ws)
        self._returning_cron = QtCore.QTimer(self)
        self._returning_cron.timeout.connect(self._tick_returning_cron)
        self._returning_cron.start(30_000)

    def _tick_returning_cron(self):
        now = datetime.now()
        changed = False
        alerts: List[Tuple[str, ScheduleEntry]] = []

        for entry in self.sched.entries:
            if entry.state == "returning_to_ward" and entry.returning_started_at:
                t0 = _parse_iso(entry.returning_started_at)
                if not t0 or not entry.time_end:
                    continue
                if (now - t0) >= timedelta(minutes=3):
                    if self._is_entry_completed(entry):
                        entry.postop_completed = True
                        entry.state = "returned_to_ward"
                        entry.returned_to_ward_at = now.strftime("%Y-%m-%dT%H:%M:%S")
                        self._db_insert_case(entry)
                        alerts.append(("ok", entry))
                    else:
                        entry.postop_completed = False
                        entry.state = "postop_pending"
                        entry.returned_to_ward_at = now.strftime("%Y-%m-%dT%H:%M:%S")
                        alerts.append(("warn", entry))
                    entry.version = int(entry.version or 1) + 1
                    changed = True

        if changed:
            self.sched._save()
            self._render_tree2()
            if alerts:
                kind, entry = alerts[-1]
                if kind == "ok":
                    self._banner_returned_ok(entry)
                else:
                    self._banner_incomplete(entry)

    def _db_insert_case(self, entry: "ScheduleEntry"):
        try:
            self.db_logger.log_event(
                case_uid=entry.case_uid,
                event="returned_to_ward",
                details={
                    "hn": entry.hn,
                    "or": entry.or_room,
                    "time_start": entry.time_start,
                    "time_end": entry.time_end,
                    "assist1": entry.assist1,
                    "assist2": entry.assist2,
                    "scrub": entry.scrub,
                    "circulate": entry.circulate,
                    "diags": entry.diags,
                    "ops": entry.ops,
                },
                emergency=str(entry.urgency).lower() == "emergency",
            )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "DB error", str(exc))

    def _banner_incomplete(self, entry: "ScheduleEntry"):
        self.result_banner.set_icon("⚠️")
        self.result_banner.set_title("⚠️ ข้อมูลหลังผ่าตัดยังไม่ครบ — ยังไม่บันทึกลงฐานข้อมูล")
        self.result_banner.set_subtitle(
            f"HN {entry.hn} | OR {entry.or_room} | โปรดกรอกทีมพยาบาล/Diagnosis/Operation และตรวจเวลาเริ่ม–จบ"
        )

    def _banner_returned_ok(self, entry: "ScheduleEntry"):
        self.result_banner.set_icon("✅")
        self.result_banner.set_title("บันทึกลงฐานข้อมูลสำเร็จ (Returned)")
        self.result_banner.set_subtitle(
            f"HN {entry.hn} | OR {entry.or_room} | เวลา {entry.time_start or '-'}–{entry.time_end or '-'}"
        )

    # ---------- PDPA first-run gate ----------
    def _pdpa_gate(self):
        # เตรียม salt ทันที (ใช้สำหรับ export แบบ de-id)
        _get_or_create_secret(SECRET_SALT_KEY, 32)
        # แสดง PDPA แค่ครั้งแรก
        if not self.cfg.value(PDPA_ACK_KEY, False, type=bool):
            dlg = PDPANoticeDialog(self)
            dlg.exec()
            self.cfg.setValue(PDPA_ACK_KEY, True); self.cfg.sync()

    # ---------- helpers ----------
    def _client(self):
        try:
            return ClientHTTP(self.ent_host.text().strip() or "127.0.0.1",
                              int(self.ent_port.text().strip() or DEFAULT_PORT),
                              self.ent_token.text().strip() or DEFAULT_TOKEN)
        except Exception:
            return ClientHTTP()
    def _on_health(self):
        try: self._client().health(); self._chip(True)
        except Exception: self._chip(False)
    def _chip(self, ok:bool):
        if ok:
            self.status_chip.setText("● Online"); self.status_chip.setStyleSheet("color:#10b981;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")
        else:
            self.status_chip.setText("● Offline"); self.status_chip.setStyleSheet("color:#ef4444;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")
    def _refresh(self, prefer_server=True):
        self.btn_refresh.setEnabled(False)
        try:
            data=self._client().list_items()
            rows=extract_rows(data)
            # อัปเดต historical monitor seen ก่อน render (เก็บ HN ที่ monitor รายงานมา)
            self._scan_monitor_status_transitions(rows)
            self._rebuild_table(rows); self._chip(True)
        except Exception:
            self._chip(False); self._rebuild_table([])
        finally:
            self.btn_refresh.setEnabled(True)

    def _rebuild_table(self, rows):
        self.rows_cache=rows; self.table.setRowCount(0)
        if not rows:
            self.table.setRowCount(1); self.table.setSpan(0,0,1,4)
            empty=QtWidgets.QTableWidgetItem("ไม่มีข้อมูล (กดรีเฟรช)")
            empty.setFlags(QtCore.Qt.ItemIsEnabled); empty.setForeground(QtGui.QBrush(QtGui.QColor("#64748b")))
            self.table.setItem(0,0,empty)
            # อัปเดต Result tree ให้ปรับตัวกรองกรณี HN หายไป
            self._render_tree2()
            return
        for r in rows:
            i=self.table.rowCount(); self.table.insertRow(i)
            self.table.setItem(i,0,QtWidgets.QTableWidgetItem(str(r.get("id",""))))
            self.table.setItem(i,1,QtWidgets.QTableWidgetItem(str(r.get("patient_id",""))))
            status = str(r.get("status",""))
            col=STATUS_COLORS.get(status, "#64748b")
            chip = StatusChipWidget(status or "-", col, pulse=(status in PULSE_STATUS))
            self.table.setCellWidget(i,2, chip)
            ts=_parse_iso(r.get("timestamp")); txt=""
            if ts: txt=_fmt_td(datetime.now()-ts)
            self.table.setItem(i,3,QtWidgets.QTableWidgetItem(txt))
        # ให้ Result tree รีเฟรชเงื่อนไขแสดงผลด้วย เมื่อ monitor เปลี่ยน
        self._render_tree2()

    def _ws_url(self):
        return f"ws://{self.ent_host.text().strip() or '127.0.0.1'}:{int(self.ent_port.text().strip() or DEFAULT_PORT)}{API_WS}?token={self.ent_token.text().strip() or DEFAULT_TOKEN}"
    def _start_ws(self):
        try:
            self.ws=QWebSocket()
            self.ws.errorOccurred.connect(lambda _e:self._ws_disc())
            self.ws.connected.connect(lambda:(self._chip(True), self._pull.stop()))
            self.ws.disconnected.connect(self._ws_disc)
            self.ws.textMessageReceived.connect(self._on_ws_msg)
            self.ws.open(QUrl(self._ws_url()))
        except Exception:
            self._ws_disc()
    def _ws_disc(self):
        if self._pull.isActive()==False: self._pull.start(3000)
    def _on_ws_msg(self, msg):
        try:
            rows=extract_rows(json.loads(msg))
            self._scan_monitor_status_transitions(rows)
            self._rebuild_table(rows)
        except Exception: pass

    # ---------- schedule ----------
    def _refresh_or_cb(self, cb:QtWidgets.QComboBox):
        cb.clear(); cb.addItems(self.sched.or_rooms)
    def _manage_or(self):
        dlg=QtWidgets.QDialog(self); dlg.setWindowTitle("จัดการ OR"); v=QtWidgets.QVBoxLayout(dlg)
        lst=QtWidgets.QListWidget(); lst.addItems(self.sched.or_rooms); v.addWidget(lst)
        h=QtWidgets.QHBoxLayout(); ent=QtWidgets.QLineEdit(); ent.setPlaceholderText("เช่น OR9"); btn_add=QtWidgets.QPushButton("เพิ่ม"); btn_del=QtWidgets.QPushButton("ลบ"); h.addWidget(ent,1); h.addWidget(btn_add); h.addWidget(btn_del); v.addLayout(h)
        ok=QtWidgets.QPushButton("บันทึก"); ok.setProperty("variant","primary"); v.addWidget(ok)
        btn_add.clicked.connect(lambda: (ent.text().strip().upper() and lst.addItem(ent.text().strip().upper()), ent.setText("")))
        btn_del.clicked.connect(lambda: [lst.takeItem(lst.row(x)) for x in lst.selectedItems()])
        def save():
            rooms=[lst.item(i).text() for i in range(lst.count())]; self.sched.set_or_rooms(rooms); self._refresh_or_cb(self.cb_or); dlg.accept()
        ok.clicked.connect(save); dlg.exec()

    def _on_import_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "เลือกไฟล์นำเข้าตารางผ่าตัด",
            str(Path.home()),
            "Excel/CSV (*.xlsx *.xlsm *.xls *.csv)"
        )
        if not path:
            return

        loader = SweetAlert.loading(self, "กำลังอ่านไฟล์นำเข้า...")
        loader.setLabelText("กำลังอ่านไฟล์นำเข้า...")
        loader.show()
        QtWidgets.QApplication.processEvents()

        error_title: Optional[str] = None
        error_message: Optional[str] = None
        empty_rows = False
        ok = 0
        skipped: List[Tuple[str, str]] = []

        try:
            rows = self._load_fixed_excel_rows(path)
            if not rows:
                empty_rows = True
            else:
                loader.setLabelText("กำลังนำเข้าข้อมูลเข้าสู่ตาราง...")
                QtWidgets.QApplication.processEvents()
                ok, skipped = self._import_from_fixed_excel_rows(rows)
        except ImportError as exc:
            error_title = "นำเข้าไม่ได้"
            error_message = str(exc)
        except Exception as exc:
            error_title = "นำเข้าล้มเหลว"
            error_message = str(exc)
        finally:
            loader.close()

        if error_message:
            QtWidgets.QMessageBox.critical(self, error_title or "ผิดพลาด", error_message)
            return

        if empty_rows:
            SweetAlert.warning(self, "เตือน", "ไม่พบข้อมูลในไฟล์ที่เลือก")
            return

        if ok <= 0:
            SweetAlert.warning(self, "เตือน", "ไม่สามารถนำเข้าแถวใดได้")
            if skipped:
                QtWidgets.QMessageBox.information(
                    self,
                    "รายละเอียดแถวที่ข้าม",
                    "\n".join([f"HN {hn}: {reason}" for hn, reason in skipped[:20]])
                    + ("\n… (มีมากกว่านี้)" if len(skipped) > 20 else ""),
                )
            return

        msg = f"นำเข้าสำเร็จ {ok} แถว"
        if skipped:
            msg += f" • ข้าม {len(skipped)} แถว"
        SweetAlert.success(self, "สำเร็จ", msg, auto_close_msec=1500)

        if skipped:
            QtWidgets.QMessageBox.information(
                self,
                "รายละเอียดแถวที่ข้าม",
                "\n".join([f"HN {hn}: {reason}" for hn, reason in skipped[:20]])
                + ("\n… (มีมากกว่านี้)" if len(skipped) > 20 else ""),
            )

    def _load_fixed_excel_rows(self, path: str) -> List[dict]:
        suffix = Path(path).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as exc:  # pragma: no cover - runtime dependency
                raise ImportError("ต้องติดตั้ง openpyxl เพื่ออ่านไฟล์ Excel") from exc

            wb = load_workbook(path, data_only=True)
            sheet = wb.active
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                return []

            headers = [str(col).strip() if col is not None else "" for col in header]
            results: List[dict] = []
            for row in rows_iter:
                row_dict: Dict[str, object] = {}
                has_value = False
                for key, value in zip(headers, row):
                    if not key:
                        continue
                    row_dict[key] = value if value is not None else ""
                    if not has_value and str(value or "").strip():
                        has_value = True
                if row_dict and has_value:
                    results.append(row_dict)
            return results

        if suffix == ".csv":
            results: List[dict] = []
            with open(path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                if reader.fieldnames is None:
                    return []
                for row in reader:
                    row_dict: Dict[str, object] = {}
                    has_value = False
                    for key, value in row.items():
                        if not key:
                            continue
                        key_clean = str(key).strip()
                        row_dict[key_clean] = value if value is not None else ""
                        if not has_value and str(value or "").strip():
                            has_value = True
                    if row_dict and has_value:
                        results.append(row_dict)
            return results

        raise ValueError("รองรับเฉพาะไฟล์ Excel (.xlsx/.xlsm) หรือ CSV")

    def _import_from_fixed_excel_rows(self, rows: List[dict]) -> Tuple[int, List[Tuple[str, str]]]:
        ok = 0
        skipped: List[Tuple[str, str]] = []

        known_wards: List[str] = []
        if hasattr(self, "cb_ward") and isinstance(self.cb_ward, QtWidgets.QComboBox):
            for i in range(self.cb_ward.count()):
                text = self.cb_ward.itemText(i).strip()
                if text:
                    known_wards.append(text)
        else:
            known_wards = [w for w in WARD_LIST if w and w != WARD_PLACEHOLDER]

        qdate = self.date.date() if hasattr(self, "date") else QtCore.QDate.currentDate()
        if hasattr(qdate, "toPython"):
            base_date = qdate.toPython()
        else:  # pragma: no cover - fallback for older Qt bindings
            base_date = datetime(qdate.year(), qdate.month(), qdate.day()).date()

        default_period = self._update_period_info()

        for raw in rows:
            if not isinstance(raw, dict):
                continue

            lookup: Dict[str, object] = {}
            for key, value in raw.items():
                key_str = str(key or "").strip()
                if not key_str:
                    continue
                lookup[key_str] = value
                lookup[key_str.lower()] = value

            def get(field: str) -> str:
                header = FIXED_MAPPING_TH[field]
                val = lookup.get(header)
                if val is None:
                    val = lookup.get(header.strip())
                if val is None:
                    val = lookup.get(header.lower())
                return str(val or "").strip()

            time_raw = lookup.get(FIXED_MAPPING_TH["time"])
            time_str = parse_time_hhmm_or_tf(time_raw)
            hn = get("hn")
            name = get("name")
            doctor = get("doctor")

            if not (hn and name and doctor):
                skipped.append((hn or "-", "ต้องมี HN, ชื่อ, แพทย์ผู้สั่ง"))
                continue

            if time_str != "TF":
                try:
                    hh, mm = [int(x) for x in time_str.split(":", 1)]
                    period_code = _now_period(datetime(base_date.year, base_date.month, base_date.day, hh, mm))
                except Exception:
                    period_code = default_period
            else:
                period_code = default_period

            diag_txt = get("diags")
            op_txt = get("ops")

            or_room = pick_or_by_doctor(base_date, time_str, doctor)

            entry = ScheduleEntry(
                or_room=or_room or "-",
                dt=base_date,
                time_str=time_str,
                hn=hn,
                name=" ".join(name.split()),
                age=parse_age_years(get("age")),
                dept="",
                doctor=normalize_doctor(doctor),
                diags=[diag_txt] if diag_txt else [],
                ops=[op_txt] if op_txt else [],
                ward=map_to_known_ward(get("ward"), known_wards),
                case_size="",
                period=period_code,
                urgency="Elective",
                assist1="",
                assist2="",
                scrub="",
                circulate="",
                time_start="",
                time_end="",
            )

            self.sched.add(entry)
            try:
                self.db_logger.append_entry(entry)
            except Exception:
                pass
            ok += 1

        self._set_result_title()
        self._render_tree2()

        return ok, skipped

    def _snapshot_sched(self) -> None:
        try:
            self._last_snapshot = [entry.to_dict() for entry in self.sched.all()]
        except Exception:
            self._last_snapshot = None
        if hasattr(self, "btn_undo_clear"):
            self.btn_undo_clear.setEnabled(bool(self._last_snapshot))

    def _restore_snapshot(self) -> None:
        if not self._last_snapshot:
            SweetAlert.info(self, "ไม่มีสำเนา", "ยังไม่มีข้อมูลสำหรับย้อนกลับ")
            return

        restored: List[ScheduleEntry] = []
        for payload in self._last_snapshot:
            try:
                restored.append(ScheduleEntry.from_dict(payload))
            except Exception:
                continue

        self._last_snapshot = None
        if hasattr(self, "btn_undo_clear"):
            self.btn_undo_clear.setEnabled(False)

        self.sched.replace_all(restored)
        self._set_result_title()
        self._render_tree2()
        SweetAlert.success(self, "เรียบร้อย", "กู้คืนข้อมูลล่าสุดแล้ว", auto_close_msec=1500)

    def _on_clear_board_clicked(self) -> None:
        entries = self.sched.all()
        if not entries:
            SweetAlert.info(self, "ไม่มีข้อมูล", "ยังไม่มีรายการในตารางให้ล้าง")
            return

        box = QtWidgets.QMessageBox(self)
        box.setWindowTitle("ยืนยันการล้างกระดาน")
        box.setIcon(QtWidgets.QMessageBox.Question)
        box.setText("ต้องการล้างรายการในตารางสำหรับทดสอบใช่ไหม?")
        btn_today = box.addButton("ล้างเฉพาะวันปัจจุบัน", QtWidgets.QMessageBox.AcceptRole)
        btn_all = box.addButton("ล้างทั้งหมด", QtWidgets.QMessageBox.DestructiveRole)
        btn_cancel = box.addButton("ยกเลิก", QtWidgets.QMessageBox.RejectRole)
        box.setDefaultButton(btn_today)
        box.exec()

        clicked = box.clickedButton()
        if clicked is None or clicked is btn_cancel:
            return

        if clicked is btn_today:
            qdate = self.date.date() if hasattr(self, "date") else QtCore.QDate.currentDate()
            day = qdate.toPython()
            todays = [e for e in entries if getattr(e, "date", None) == day]
            if not todays:
                SweetAlert.info(self, "ไม่พบข้อมูล", f"ไม่มีรายการวันที่ {day.strftime('%d/%m/%Y')} ให้ล้าง")
                return

            self._snapshot_sched()
            removed = self.sched.remove_by_date(day)
            if removed <= 0:
                self._last_snapshot = None
                if hasattr(self, "btn_undo_clear"):
                    self.btn_undo_clear.setEnabled(False)
                SweetAlert.info(self, "ไม่พบข้อมูล", f"ไม่มีรายการวันที่ {day.strftime('%d/%m/%Y')} ให้ล้าง")
                return

            message = f"ลบ {removed} รายการของวันที่ {day.strftime('%d/%m/%Y')} แล้ว"
        else:
            self._snapshot_sched()
            removed = self.sched.clear()
            if removed <= 0:
                self._last_snapshot = None
                if hasattr(self, "btn_undo_clear"):
                    self.btn_undo_clear.setEnabled(False)
                SweetAlert.info(self, "ไม่มีข้อมูล", "ยังไม่มีรายการในตารางให้ล้าง")
                return

            message = f"ลบทั้งหมด {removed} รายการแล้ว"

        self._set_result_title()
        self._render_tree2()
        SweetAlert.success(self, "ล้างสำเร็จ", message, auto_close_msec=1500)

    def _on_undo_clear_clicked(self) -> None:
        self._restore_snapshot()

    def _update_period_info(self):
        qd = self.date.date()
        qtime = self.time.time()
        dt = datetime(qd.year(), qd.month(), qd.day(), qtime.hour(), qtime.minute())
        auto = _now_period(dt)
        if hasattr(self, "lbl_period_info"):
            self.lbl_period_info.setText(
                f"ระบบกำหนดช่วงเวลาอัตโนมัติ: {_period_label(auto)} (อ้างอิง {dt:%d/%m/%Y %H:%M})"
            )
        return auto

    def _on_dept_changed(self, dept_label: str):
        if dept_label and not dept_label.startswith("—"):
            self._set_doctor_visibility(True)
            self.cb_doctor.clear()
            self.cb_doctor.addItems(DEPT_DOCTORS.get(dept_label, []))
        else:
            self._set_doctor_visibility(False)
            self.cb_doctor.clear()

        specialty = _dept_to_specialty_key(dept_label or "")
        self._current_specialty_key = specialty

        if not specialty:
            self._diag_base_catalog = []
            self._diag_catalog_full = []
            self._op_catalog_full = []
            self._dx_index = None
            self._op_index = None
            if self.diag_adder.search_line:
                self.diag_adder.search_line.clear()
            if self.op_adder.search_line:
                self.op_adder.search_line.clear()
            self.diag_adder.clear()
            self.op_adder.clear()
            self.diag_adder.set_suggestions([])
            self.op_adder.set_suggestions([])
            self.diag_adder.setEnabled(False)
            self.op_adder.setEnabled(False)
            return

        loader = SweetAlert.loading(self, "กำลังเตรียมรายการสำหรับแผนกนี้...")
        QtWidgets.QApplication.processEvents()
        try:
            base_ops = get_operations(specialty) or []
            base_dx_list = get_diagnoses(specialty) or []
            user_op = get_custom_entries("operation", specialty) or []
            user_dx = get_custom_entries("diagnosis", specialty) or []
        finally:
            loader.close()

        merged_ops: List[str] = []
        for bucket in (user_op, base_ops):
            for value in bucket:
                val = (value or "").strip()
                if val and val not in merged_ops:
                    merged_ops.append(val)

        merged_dx: List[str] = []
        for bucket in (user_dx, base_dx_list):
            for value in bucket:
                val = (value or "").strip()
                if val and val not in merged_dx:
                    merged_dx.append(val)

        self._op_catalog_full = merged_ops
        self._diag_base_catalog = merged_dx
        self._diag_catalog_full = list(merged_dx)

        self._op_index = FastSearchIndex(merged_ops, prefix_len=3) if merged_ops else None
        self._dx_index = FastSearchIndex(merged_dx, prefix_len=3) if merged_dx else None

        if self.op_adder.search_line:
            self.op_adder.search_line.clear()
        if self.diag_adder.search_line:
            self.diag_adder.search_line.clear()

        self.op_adder.clear()
        self.diag_adder.clear()

        self.op_adder.setEnabled(True)
        self.diag_adder.setEnabled(True)

        self._latest_op_query = ""
        self._latest_diag_query = ""

        initial_ops = (
            self._op_index.search("", self._op_search_limit) if self._op_index else merged_ops[: self._op_search_limit]
        )
        initial_dx = (
            self._dx_index.search("", self._dx_search_limit) if self._dx_index else merged_dx[: self._dx_search_limit]
        )

        self.op_adder.set_suggestions(initial_ops)
        self.diag_adder.set_suggestions(initial_dx)

        self._refresh_diag_suggestions()

    def _on_operation_persist_requested(self, text: str):
        item = (text or "").strip()
        if not item:
            return
        specialty = self._current_specialty_key_safe()
        if not specialty:
            SweetAlert.warning(self, "เตือน", "กรุณาเลือกแผนกก่อน")
            return
        loader = SweetAlert.loading(self, "กำลังบันทึก Operation ...")
        QtWidgets.QApplication.processEvents()
        try:
            added = add_custom_entry("operation", specialty, item)
        finally:
            loader.close()
        self._on_dept_changed(self.cb_dept.currentText())
        SweetAlert.success(
            self,
            "สำเร็จ" if added else "ซ้ำ",
            "บันทึก Operation เพิ่มเข้าคลังแล้ว" if added else "มี Operation นี้อยู่แล้ว",
        )

    def _on_diagnosis_persist_requested(self, text: str):
        item = (text or "").strip()
        if not item:
            return
        specialty = self._current_specialty_key_safe()
        if not specialty:
            SweetAlert.warning(self, "เตือน", "กรุณาเลือกแผนกก่อน")
            return
        loader = SweetAlert.loading(self, "กำลังบันทึก Diagnosis ...")
        QtWidgets.QApplication.processEvents()
        try:
            added = add_custom_entry("diagnosis", specialty, item)
        finally:
            loader.close()
        self._on_dept_changed(self.cb_dept.currentText())
        SweetAlert.success(
            self,
            "สำเร็จ" if added else "ซ้ำ",
            "บันทึก Diagnosis เพิ่มเข้าคลังแล้ว" if added else "มี Diagnosis นี้อยู่แล้ว",
        )

    def _refresh_diag_suggestions(self):
        specialty = self._current_specialty_key_safe()
        if not specialty:
            self._diag_catalog_full = []
            self._dx_index = None
            self.diag_adder.set_suggestions([])
            return

        base_suggestions = diagnosis_suggestions(specialty)

        merged: List[str] = []

        def _append(values: List[str]) -> None:
            for value in values or []:
                val = (value or "").strip()
                if val and val not in merged:
                    merged.append(val)

        _append(self._diag_base_catalog)
        _append(list(base_suggestions))

        self._diag_catalog_full = merged
        self._dx_index = FastSearchIndex(self._diag_catalog_full, prefix_len=3) if self._diag_catalog_full else None
        if self.diag_adder.search_line:
            self._latest_diag_query = self.diag_adder.search_line.text()
        if self._dx_index:
            initial = self._dx_index.search(self._latest_diag_query, self._dx_search_limit)
        else:
            initial = []
        self.diag_adder.set_suggestions(initial)

    def _on_operations_changed(self, _items: List[str]):
        self._refresh_diag_suggestions()

    def _on_diag_query_changed(self, text: str):
        self._latest_diag_query = text or ""
        if not self._dx_index:
            return
        self._search_timer.stop()
        self._search_timer.start()

    def _on_diag_search_timeout(self):
        self._run_diag_search(self._latest_diag_query)

    def _run_diag_search(self, query: str):
        if not self._dx_index:
            self.diag_adder.set_suggestions([])
            return
        future = self._search_executor.submit(self._dx_index.search, query, self._dx_search_limit)

        def _apply(fut):
            try:
                results = fut.result()
            except Exception:
                results = []
            if query != self._latest_diag_query:
                return
            self.diag_adder.set_suggestions(results)

        future.add_done_callback(lambda fut: QtCore.QTimer.singleShot(0, lambda: _apply(fut)))

    def _on_op_query_changed(self, text: str):
        self._latest_op_query = text or ""
        if self._op_index:
            self._op_search_timer.stop()
            self._op_search_timer.start()
        else:
            self._run_op_search(self._latest_op_query)

    def _on_op_search_timeout(self):
        self._run_op_search(self._latest_op_query)

    def _run_op_search(self, query: str):
        if not self._op_index:
            if not self._op_catalog_full:
                self.op_adder.set_suggestions([])
                return
            normalized_query = normalize_text(query)
            if not normalized_query:
                subset = self._op_catalog_full[: self._op_search_limit]
            else:
                subset = []
                for item in self._op_catalog_full:
                    if normalized_query in normalize_text(item):
                        subset.append(item)
                    if len(subset) >= self._op_search_limit:
                        break
            self.op_adder.set_suggestions(subset)
            return

        future = self._search_executor.submit(self._op_index.search, query, self._op_search_limit)

        def _apply(fut):
            try:
                results = fut.result()
            except Exception:
                results = []
            if query != self._latest_op_query:
                return
            self.op_adder.set_suggestions(results)

        future.add_done_callback(lambda fut: QtCore.QTimer.singleShot(0, lambda: _apply(fut)))

    def _set_doctor_visibility(self, visible: bool):
        self.row_doctor_label.setVisible(visible); self.cb_doctor.setVisible(visible)

    def _collect(self):
        qd=self.date.date()
        dt = datetime(qd.year(), qd.month(), qd.day(), self.time.time().hour(), self.time.time().minute())
        auto_period = _now_period(dt)
        ward_text = self.cb_ward.currentText().strip()
        if ward_text == WARD_PLACEHOLDER:
            ward_text = ""
        return ScheduleEntry(
            or_room=self.cb_or.currentText().strip(), dt=dt.date(), time_str=self.time.time().toString("HH:mm"),
            hn=self.ent_hn.text().strip(), name=self.ent_name.text().strip(), age=self.ent_age.text().strip() or "0",
            dept=(self.cb_dept.currentText().strip() if not self.cb_dept.currentText().startswith("—") else ""),
            doctor=self.cb_doctor.currentText().strip() if self.cb_doctor.isVisible() else "",
            diags=self.diag_adder.items(), ops=self.op_adder.items(),
            ward=ward_text,
            case_size=self.cb_case.currentText().strip(),
            queue=0,
            period=auto_period,
            urgency=self.cb_urgency.currentText().strip() or "Elective",
            assist1=self.cb_assist1.currentText().strip(),
            assist2=self.cb_assist2.currentText().strip(),
            scrub=self.cb_scrub.currentText().strip(),
            circulate=self.cb_circulate.currentText().strip(),
            time_start=(self.time_start.time().toString("HH:mm") if self.ck_time_start.isChecked() else ""),
            time_end=(self.time_end.time().toString("HH:mm") if self.ck_time_end.isChecked() else ""),
        )

    def _clear_form(self):
        self.cb_or.setCurrentIndex(0)
        self.ent_name.clear(); self.ent_age.clear(); self.ent_hn.clear()
        self.cb_dept.setCurrentIndex(0); self.cb_doctor.clear(); self._set_doctor_visibility(False)
        self.diag_adder.clear(); self.op_adder.clear()
        self.cb_ward.setCurrentIndex(0); self.cb_ward.setEditText(WARD_PLACEHOLDER)
        if hasattr(self, "cb_case"):
            self.cb_case.setCurrentIndex(0)
        if hasattr(self, "cb_urgency"):
            idx = self.cb_urgency.findText("Elective")
            self.cb_urgency.setCurrentIndex(idx if idx >= 0 else 0)
        for cb in (self.cb_assist1, self.cb_assist2, self.cb_scrub, self.cb_circulate):
            cb.setCurrentIndex(0)
            cb.setEditText("")
        self.ck_time_start.setChecked(False); self.time_start.setEnabled(False); self.time_start.setTime(QtCore.QTime.currentTime())
        self.ck_time_end.setChecked(False); self.time_end.setEnabled(False); self.time_end.setTime(QtCore.QTime.currentTime())
        self.date.setDate(QtCore.QDate.currentDate())
        self.time.setTime(QtCore.QTime.currentTime())
        self._update_period_info()
        self._on_dept_changed(self.cb_dept.currentText())
        self._set_add_mode()

    # ---------- ADD / UPDATE ----------
    def _set_edit_mode(self, idx:int):
        self._edit_idx = idx
        self.btn_add.setText("💾 บันทึกการแก้ไข")
        self.btn_cancel_edit.show()
        self.toast.show_toast("เข้าสู่โหมดแก้ไข")

    def _set_add_mode(self):
        self._edit_idx = None
        self.btn_add.setText("➕ เพิ่ม")
        self.btn_cancel_edit.hide()

    def _cancel_edit_mode(self):
        self._set_add_mode()
        self.toast.show_toast("ยกเลิกโหมดแก้ไข")

    def _load_form_from_entry(self, e:ScheduleEntry):
        idx = self.cb_or.findText(e.or_room)
        if idx >= 0: self.cb_or.setCurrentIndex(idx)
        self.ent_name.setText(e.name or "")
        self.ent_age.setText(str(e.age or 0))
        self.ent_hn.setText(e.hn or "")
        if hasattr(self, "cb_urgency"):
            idx_u = self.cb_urgency.findText(e.urgency or "Elective")
            self.cb_urgency.setCurrentIndex(idx_u if idx_u >= 0 else 0)
        try:
            d = QtCore.QDate(e.date.year, e.date.month, e.date.day)
            self.date.setDate(d)
        except Exception: pass
        try:
            hh, mm = (e.time or "00:00").split(":")
            self.time.setTime(QtCore.QTime(int(hh), int(mm)))
        except Exception: pass
        self._update_period_info()
        if e.dept:
            for i in range(self.cb_dept.count()):
                if self.cb_dept.itemText(i).startswith(e.dept) or self.cb_dept.itemText(i)==e.dept:
                    self.cb_dept.setCurrentIndex(i); break
        if e.doctor and self.cb_doctor.isVisible():
            j = self.cb_doctor.findText(e.doctor)
            if j >= 0: self.cb_doctor.setCurrentIndex(j)
        self.diag_adder.clear(); [self.diag_adder.list.addItem(x) for x in (e.diags or [])]
        self.op_adder.clear();   [self.op_adder.list.addItem(x)   for x in (e.ops   or [])]
        # Ward
        j = self.cb_ward.findText(e.ward) if e.ward else -1
        if j >= 0:
            self.cb_ward.setCurrentIndex(j)
        else:
            if e.ward:
                self.cb_ward.setEditText(e.ward)
            else:
                self.cb_ward.setCurrentIndex(0)
                self.cb_ward.setEditText(WARD_PLACEHOLDER)

        # Case size
        if hasattr(self, "cb_case"):
            k = self.cb_case.findText(e.case_size) if e.case_size else -1
            if k >= 0:
                self.cb_case.setCurrentIndex(k)
            else:
                self.cb_case.setCurrentIndex(0)

        # Nurse roles
        for combo, value in (
            (self.cb_assist1, e.assist1),
            (self.cb_assist2, e.assist2),
            (self.cb_scrub, e.scrub),
            (self.cb_circulate, e.circulate),
        ):
            val = value or ""
            idx_val = combo.findText(val) if val else 0
            if val and idx_val >= 0:
                combo.setCurrentIndex(idx_val)
            else:
                combo.setCurrentIndex(0)
            combo.setEditText(val)

        # Start/End time (optional)
        if e.time_start:
            self.ck_time_start.setChecked(True)
            try:
                hh, mm = e.time_start.split(":")
                self.time_start.setTime(QtCore.QTime(int(hh), int(mm)))
            except Exception:
                pass
        else:
            self.ck_time_start.setChecked(False)
            self.time_start.setEnabled(False)
            self.time_start.setTime(QtCore.QTime.currentTime())

        if e.time_end:
            self.ck_time_end.setChecked(True)
            try:
                hh, mm = e.time_end.split(":")
                self.time_end.setTime(QtCore.QTime(int(hh), int(mm)))
            except Exception:
                pass
        else:
            self.ck_time_end.setChecked(False)
            self.time_end.setEnabled(False)
            self.time_end.setTime(QtCore.QTime.currentTime())

    def _on_add_or_update(self):
        e = self._collect()
        errs=[]
        if not e.or_room: errs.append("กรุณาเลือก OR")
        if not e.name: errs.append("กรุณากรอกชื่อ-สกุล")
        if not e.hn: errs.append("กรุณากรอก HN")
        if e.hn and not e.hn.isdigit(): errs.append("HN ต้องเป็นตัวเลขเท่านั้น")
        if errs:
            self.lbl_warn.setText(" • ".join(errs))
            try: QtWidgets.QApplication.beep()
            except Exception: pass
            return
        else:
            self.lbl_warn.setText("")

        # จำ uid ไว้เพื่อโฟกัสหลังบันทึก
        self._last_focus_uid = e.uid()

        loader = SweetAlert.loading(self, "กำลังบันทึกข้อมูล...")
        QtWidgets.QApplication.processEvents()
        try:
            if self._edit_idx is None:
                self.sched.add(e)
                try:
                    self.db_logger.append_entry(e)
                except Exception:
                    pass
                self._notify("เพิ่มรายการแล้ว", f"OR {e.or_room} • {e.time} • HN {e.hn}")
                SweetAlert.success(self, "สำเร็จ", "บันทึกรายการใหม่เรียบร้อย")
                # ไม่เพิ่มเข้า historical_monitor_seen ที่นี่ — ปล่อยให้ monitor รายงาน HN จะเป็นคนเพิ่ม
            else:
                if 0 <= self._edit_idx < len(self.sched.entries):
                    old_entry = self.sched.entries[self._edit_idx]
                    e.case_uid = old_entry.case_uid
                    e.version = int(old_entry.version or 1) + 1
                    e.state = old_entry.state
                    e.returning_started_at = old_entry.returning_started_at
                    e.returned_to_ward_at = old_entry.returned_to_ward_at
                    e.postop_completed = old_entry.postop_completed
                self.sched.update(self._edit_idx, e)
                self._notify("บันทึกการแก้ไขแล้ว", f"OR {e.or_room} • {e.time} • HN {e.hn}")
                SweetAlert.success(self, "สำเร็จ", "อัปเดตรายการเรียบร้อย")
                self._set_add_mode()
        finally:
            loader.close()

        self._set_result_title()
        self._render_tree2()

        # เด้งไปแท็บ Result และโฟกัส/ไฮไลต์ชื่อผู้ป่วย
        self.tabs.setCurrentIndex(1)
        QtCore.QTimer.singleShot(120, lambda: self._focus_uid(self._last_focus_uid))

        self._clear_form()

    # ---------- Helpers for monitor state sync ----------
    def _find_entry_by_hn_latest(self, hn: str):
        matches = [e for e in self.sched.entries if str(e.hn).strip() == str(hn).strip()]
        if not matches:
            return None

        def _key(entry: ScheduleEntry):
            try:
                hh, mm = (entry.time or "00:00").split(":")
                return (str(entry.date), int(hh) * 60 + int(mm))
            except Exception:
                return (str(entry.date), 0)

        matches.sort(key=_key, reverse=True)
        return matches[0]

    def _set_time_start_if_empty(self, entry: ScheduleEntry):
        if not entry.time_start:
            entry.time_start = datetime.now().strftime("%H:%M")
            entry.version = int(entry.version or 1) + 1

    def _set_time_end_if_empty(self, entry: ScheduleEntry):
        if not entry.time_end:
            entry.time_end = datetime.now().strftime("%H:%M")
            entry.version = int(entry.version or 1) + 1

    def _scan_monitor_status_transitions(self, rows: List[dict]):
        for row in rows:
            hn = str(row.get("hn_full") or row.get("id") or "").strip()
            if not hn:
                continue
            status = str(row.get("status") or "").strip()
            if not status:
                continue

            prev = self._last_status_by_hn.get(hn)
            if prev == status:
                continue
            self._last_status_by_hn[hn] = status

            entry = self._find_entry_by_hn_latest(hn)
            if not entry:
                continue

            changed = False
            if status == STATUS_OP_START:
                self._set_time_start_if_empty(entry)
                if entry.state in {"scheduled", "in_or", "operation_ended", "postop_pending", ""}:
                    entry.state = "operation_started"
                    changed = True
            elif status == STATUS_OP_END:
                self._set_time_end_if_empty(entry)
                if entry.state in {"operation_started", "in_or", "scheduled", ""}:
                    entry.state = "operation_ended"
                    changed = True
            elif status == STATUS_RETURNING:
                if not entry.time_end:
                    continue
                if entry.state != "returning_to_ward":
                    entry.state = "returning_to_ward"
                    entry.returning_started_at = _now_iso()
                    entry.version = int(entry.version or 1) + 1
                    changed = True

            if changed:
                self.sched._save()

    def _is_entry_completed(self, e: ScheduleEntry) -> bool:
        """ตรวจว่ารายการถูกเติมข้อมูลหลังผ่าตัดครบถ้วนพอสำหรับการปิดเคส"""
        return _is_postop_complete_entry(e)

    def _create_or_header_widget(self, or_room: str, plan_desc: str) -> QtWidgets.QWidget:
        container = QtWidgets.QFrame()
        container.setObjectName("orHeaderFrame")
        container.setStyleSheet(
            """
            QFrame#orHeaderFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #34d399, stop:1 #059669);
                border-radius: 18px;
                padding: 12px 18px;
            }
            QFrame#orHeaderFrame QLabel {
                color: #ffffff;
            }
            """
        )
        container.setMinimumHeight(56)

        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(10, 6, 10, 6)
        layout.setSpacing(2)

        title = QtWidgets.QLabel((or_room or "-").upper())
        title_font = QtGui.QFont(self.font())
        title_font.setBold(True)
        title_font.setPointSize(max(title_font.pointSize() + 2, title_font.pointSize()))
        title.setFont(title_font)

        subtitle = QtWidgets.QLabel("ห้องผ่าตัด")
        subtitle.setStyleSheet("font-size: 12px; font-weight: 600; color: rgba(255,255,255,0.85);")

        layout.addWidget(title)
        layout.addWidget(subtitle)

        if plan_desc:
            plan = QtWidgets.QLabel(plan_desc)
            plan.setWordWrap(True)
            plan.setStyleSheet("color: rgba(255,255,255,0.82); font-size: 11px;")
            layout.addWidget(plan)

        layout.addStretch(1)
        return container

    def _render_tree2(self):
        hbar = self.tree2.horizontalScrollBar()
        old_hval = hbar.value()
        self.tree2.setUpdatesEnabled(False)

        try:
            self.tree2.clear()
            self._set_result_title()

            entries_to_show: List[Tuple[int, ScheduleEntry]] = list(enumerate(self.sched.entries))
            if not entries_to_show:
                return

            groups: Dict[str, List[Tuple[int, ScheduleEntry]]] = {}
            for idx, entry in entries_to_show:
                groups.setdefault(entry.or_room or "-", []).append((idx, entry))

            order = self.sched.or_rooms

            base_date = datetime.now().date()
            try:
                if hasattr(self, "date"):
                    qdate = self.date.date()
                    if hasattr(qdate, "toPython"):
                        base_date = qdate.toPython()
                    else:
                        base_date = date(qdate.year(), qdate.month(), qdate.day())
            except Exception:
                base_date = datetime.now().date()

            def time_key(se: Tuple[int, ScheduleEntry]):
                entry = se[1]
                return entry.time or "99:99"

            status_colors = {
                "returning_to_ward": "#ede9fe",
                "postop_pending": "#fff7ed",
                "returned_to_ward": "#ecfdf5",
            }
            status_icons = {
                "returning_to_ward": "⏳",
                "postop_pending": "⚠️",
                "returned_to_ward": "✅",
            }

            for orr in sorted(groups.keys(), key=lambda x: (order.index(x) if x in order else 999, x)):
                plan_desc = describe_or_plan_label(base_date, orr)

                parent = QtWidgets.QTreeWidgetItem(["" for _ in range(self.tree2.columnCount())])
                parent.setChildIndicatorPolicy(QtWidgets.QTreeWidgetItem.ShowIndicator)
                parent.setFirstColumnSpanned(True)
                parent.setSizeHint(0, QtCore.QSize(220, 72))
                for c in range(self.tree2.columnCount()):
                    parent.setData(c, QtCore.Qt.UserRole + 99, "grp")
                self.tree2.addTopLevelItem(parent)

                header_widget = self._create_or_header_widget(orr or "-", plan_desc)
                if plan_desc:
                    header_widget.setToolTip(plan_desc)
                self.tree2.setItemWidget(parent, 0, header_widget)
                parent.setExpanded(True)

                rows_sorted = sorted(
                    groups[orr],
                    key=lambda se: (0, int(se[1].queue)) if int(se[1].queue or 0) > 0 else (1, time_key(se))
                )

                for idx, entry in rows_sorted:
                    diag_txt = " with ".join(entry.diags) if entry.diags else "-"
                    op_txt = " with ".join(entry.ops) if entry.ops else "-"
                    or_label = entry.or_room or "-"
                    time_label = entry.time if entry.time else "-"
                    if time_label == "-":
                        or_time_text = or_label
                    else:
                        or_time_text = f"{or_label} • {time_label}"
                    row = QtWidgets.QTreeWidgetItem([
                        _period_label(entry.period),
                        or_time_text,
                        entry.hn,
                        entry.name or "-",
                        str(entry.age or 0),
                        diag_txt,
                        op_txt,
                        entry.doctor or "-",
                        entry.ward or "-",
                        entry.case_size or "-",
                        entry.dept or "-",
                        entry.assist1 or "-",
                        entry.assist2 or "-",
                        entry.scrub or "-",
                        entry.circulate or "-",
                        entry.time_start or "-",
                        entry.time_end or "-",
                        "",
                        entry.urgency or "Elective",
                    ])
                    row.setData(0, QtCore.Qt.UserRole, entry.uid())
                    row.setData(0, QtCore.Qt.UserRole + 1, idx)
                    parent.addChild(row)

                    qs = QueueSelectWidget(int(entry.queue or 0))
                    uid = entry.uid()
                    qs.changed.connect(lambda new_q, u=uid: self._apply_queue_select(u, int(new_q)))
                    self.tree2.setItemWidget(row, 17, qs)

                    state = entry.state or ""
                    if state in status_colors:
                        brush = QtGui.QBrush(QtGui.QColor(status_colors[state]))
                        for col_idx in range(self.tree2.columnCount()):
                            row.setBackground(col_idx, brush)
                        icon = status_icons.get(state)
                        if icon:
                            row.setText(3, f"{icon} {row.text(3)}")
                    if state:
                        tip = [f"State: {state}"]
                        if entry.returning_started_at:
                            tip.append(f"เริ่มส่งกลับตึก: {entry.returning_started_at}")
                        if entry.returned_to_ward_at:
                            tip.append(f"กลับตึกเมื่อ: {entry.returned_to_ward_at}")
                        if entry.postop_completed:
                            tip.append("(ข้อมูลหลังผ่าตัดครบถ้วน ✓)")
                        row.setToolTip(3, "\n".join(tip))

            self.tree2.expandAll()
        finally:
            self.tree2.setUpdatesEnabled(True)
            QtCore.QTimer.singleShot(0, lambda: hbar.setValue(min(old_hval, hbar.maximum())))

    def _apply_queue_select(self, uid: str, new_q: int):
        target=None; target_idx=None
        for i, entry in enumerate(self.sched.entries):
            if entry.uid()==uid:
                target=entry; target_idx=i; break
        if not target: return
        new_q = max(0, min(9, int(new_q)))
        if new_q == target.queue: return
        target.queue = int(new_q)
        self.sched._save()
        try: QtWidgets.QApplication.beep()
        except Exception: pass
        self._notify("อัปเดตคิวสำเร็จ", f"OR {target.or_room} • HN {target.hn} → คิว {new_q or 'ตามเวลา'}")
        self._set_result_title()
        self._render_tree2()
        self._flash_row_by_uid(uid)

    def _find_item_by_uid(self, uid:str):
        root = self.tree2.invisibleRootItem()
        for i in range(root.childCount()):
            parent = root.child(i)
            for j in range(parent.childCount()):
                ch = parent.child(j)
                if ch.data(0, QtCore.Qt.UserRole)==uid:
                    return ch
        return None

    def _flash_row_by_uid(self, uid:str):
        it = self._find_item_by_uid(uid)
        if not it: return
        rect = self.tree2.visualItemRect(it)
        overlay = QtWidgets.QFrame(self.tree2.viewport())
        overlay.setGeometry(0, rect.y(), self.tree2.viewport().width(), rect.height())
        overlay.setStyleSheet("QFrame{background:rgba(250,204,21,0.35);border-radius:4px;}")
        overlay.raise_(); overlay.show()
        anim = QtCore.QPropertyAnimation(overlay, b"windowOpacity", self)
        anim.setDuration(900); anim.setStartValue(1.0); anim.setKeyValueAt(0.5, 0.0); anim.setEndValue(1.0); anim.setLoopCount(2)
        anim.finished.connect(overlay.deleteLater)
        anim.start(QtCore.QAbstractAnimation.DeleteWhenStopped)

    def _focus_uid(self, uid:str):
        if not uid: return
        it = self._find_item_by_uid(uid)
        if not it: return
        hbar = self.tree2.horizontalScrollBar()
        old_hval = hbar.value()

        rect = self.tree2.visualItemRect(it)
        if not rect.isValid():
            index = self.tree2.indexFromItem(it)
            if index.isValid():
                self.tree2.scrollTo(index, QtWidgets.QAbstractItemView.PositionAtCenter)
                rect = self.tree2.visualItemRect(it)

        if rect.isValid():
            vbar = self.tree2.verticalScrollBar()
            target = rect.y() + vbar.value() - (self.tree2.viewport().height() // 2)
            vbar.setValue(max(0, min(target, vbar.maximum())))

        self.tree2.setCurrentItem(it)
        self._flash_row_by_uid(uid)
        QtCore.QTimer.singleShot(0, lambda: hbar.setValue(min(old_hval, hbar.maximum())))

    # ---------- Result context menu / Double-click ----------
    def _result_ctx_menu(self, pos: QtCore.QPoint):
        it = self.tree2.itemAt(pos)
        if not it: return
        idx = it.data(0, QtCore.Qt.UserRole+1)
        if idx is None: return
        menu = QtWidgets.QMenu(self)
        a_edit = menu.addAction("แก้ไขรายการ")
        a_del  = menu.addAction("ลบรายการ")
        act = menu.exec(self.tree2.viewport().mapToGlobal(pos))
        if act == a_edit:
            self._on_result_double_click(it, 0)
        elif act == a_del:
            self._delete_entry_idx(int(idx))

    def _delete_entry_idx(self, idx:int):
        if 0 <= idx < len(self.sched.entries):
            entry = self.sched.entries[idx]
            if not SweetAlert.confirm(
                self,
                "ยืนยันการลบ",
                f"ต้องการลบ HN {entry.hn} ({entry.name}) หรือไม่?",
            ):
                return
            self.sched.delete(idx)
            self._render_tree2()
            SweetAlert.success(self, "สำเร็จ", "ลบรายการเรียบร้อย")

    def _on_monitor_double_click(self, item:QtWidgets.QTableWidgetItem):
        row = item.row()
        hn = self.table.item(row, 0).text().strip() if self.table.item(row,0) else ""
        if not hn:
            self.toast.show_toast("ไม่พบ HN ของแถวนี้"); return
        self._route_to_identify(hn)

    def _on_result_double_click(self, item:QtWidgets.QTreeWidgetItem, col:int):
        idx = item.data(0, QtCore.Qt.UserRole+1)
        if idx is None: return
        if 0 <= int(idx) < len(self.sched.entries):
            entry = self.sched.entries[int(idx)]
            self._load_form_from_entry(entry)
            self._set_edit_mode(int(idx))
            self.tabs.setCurrentIndex(0)  # ไปที่ฟอร์ม

    # ---------- Identify routing ----------
    def _find_entry_index_by_hn(self, hn:str)->Optional[int]:
        for i, entry in enumerate(self.sched.entries):
            if str(entry.hn).strip()==str(hn).strip():
                return i
        return None

    def _route_to_identify(self, hn:str):
        QtWidgets.QApplication.clipboard().setText(hn)
        idx = self._find_entry_index_by_hn(hn)
        if idx is not None:
            entry = self.sched.entries[idx]
            self._load_form_from_entry(entry)
            self._set_edit_mode(idx)
            self.tabs.setCurrentIndex(0)
            self.toast.show_toast(f"HN {hn}: พบในรายการ → โหมดแก้ไข")
        else:
            self._clear_form()
            self.ent_hn.setText(hn)
            self._set_add_mode()
            self.tabs.setCurrentIndex(0)
            self.toast.show_toast(f"HN {hn}: ยังไม่มี → เพิ่มใหม่")

    def apply_external_update(self, uid: str, patch: dict) -> bool:
        """รับข้อมูลจาก client ภายนอกเพื่อเติมรายละเอียดหลังผ่าตัด"""
        accepted_keys = {
            "assist1",
            "assist2",
            "scrub",
            "circulate",
            "time_start",
            "time_end",
            "ward",
            "doctor",
        }
        accepted_keys |= {
            "state",
            "returning_started_at",
            "returned_to_ward_at",
            "postop_completed",
            "version",
        }

        intent = str(patch.get("_intent") or "").strip().lower()

        for entry in self.sched.entries:
            if entry.uid() == uid:
                if intent == "mark_returning":
                    if not entry.time_end:
                        self.toast.show_toast("ยังไม่มีเวลา 'จบผ่าตัด' — ตั้งสถานะกำลังส่งกลับตึกไม่ได้")
                        return False
                    entry.state = "returning_to_ward"
                    entry.returning_started_at = _now_iso()
                    entry.postop_completed = False
                    entry.returned_to_ward_at = ""
                    entry.version = int(entry.version or 1) + 1
                    self.sched._save()
                    self._render_tree2()
                    self._flash_row_by_uid(uid)
                    self.toast.show_toast("ตั้งสถานะ 'กำลังส่งกลับตึก' แล้ว (เริ่มนับ 3 นาที)")
                    return True

                string_fields = {
                    "assist1",
                    "assist2",
                    "scrub",
                    "circulate",
                    "time_start",
                    "time_end",
                    "ward",
                    "doctor",
                    "state",
                    "returning_started_at",
                    "returned_to_ward_at",
                }

                for key in accepted_keys:
                    if key not in patch:
                        continue
                    value = patch.get(key)
                    if key == "version":
                        # version จะถูกปรับเพิ่มท้ายฟังก์ชัน
                        continue
                    if key == "postop_completed":
                        entry.postop_completed = bool(value)
                        continue
                    if key in string_fields:
                        setattr(entry, key, str(value or ""))
                        continue
                    setattr(entry, key, value)

                entry.version = int(entry.version or 1) + 1
                if entry.state == "returning_to_ward" and not entry.returning_started_at:
                    entry.returning_started_at = _now_iso()
                self.sched._save()
                self._render_tree2()
                self._flash_row_by_uid(uid)
                self.toast.show_toast("อัปเดตข้อมูลจาก Client สำเร็จ")
                return True
        return False

    # ---------- export ----------
    def _export_csv(self):
        path,_=QtWidgets.QFileDialog.getSaveFileName(self,"Export CSV","monitor.csv","CSV (*.csv)")
        if not path: return
        try:
            with open(path,"w",newline="",encoding="utf-8-sig") as f:
                w=csv.writer(f); w.writerow(["ID","Patient ID","Status","Timestamp","ETA(min)"])
                for r in self.rows_cache: w.writerow([r.get("id",""), r.get("patient_id",""), r.get("status",""), r.get("timestamp",""), r.get("eta_minutes","")])
            QtWidgets.QMessageBox.information(self,"ส่งออกแล้ว",path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self,"ผิดพลาด",str(e))

    def _export_deid_csv(self):
        """
        ส่งออกข้อมูลสำหรับวิเคราะห์แบบไม่ระบุตัวตน (de-identified)
        แหล่งข้อมูล: self.sched.entries (ตาราง Result Schedule ภายในเครื่อง)
        ฟิลด์สำคัญ: hn_hash, dept, or, queue, period, scheduled date/time, time_start, time_end, diags, ops, ward
        """
        path,_=QtWidgets.QFileDialog.getSaveFileName(self,"Export De-Identified CSV","cases_deid.csv","CSV (*.csv)")
        if not path: return
        try:
            rows=[]
            for e in self.sched.entries:
                rows.append({
                    "hn_hash": hn_hash(e.hn or ""),
                    "dept": e.dept or "",
                    "or": e.or_room or "",
                    "queue": int(e.queue or 0),
                    "period": e.period or "",
                    "scheduled_date": str(e.date or ""),
                    "scheduled_time": e.time or "",
                    "time_start": e.time_start or "",
                    "time_end": e.time_end or "",
                    "diag": " | ".join(e.diags or []),
                    "op": " | ".join(e.ops or []),
                    "ward": e.ward or "",
                    "case_size": e.case_size or "",
                    "urgency": e.urgency or "",
                    "doctor": e.doctor or "",
                    # หมายเหตุ: ไม่ส่งออก HN/ชื่อ
                })
            with open(path,"w",newline="",encoding="utf-8-sig") as f:
                cols=["hn_hash","dept","or","queue","period","scheduled_date","scheduled_time","time_start","time_end","diag","op","ward","case_size","urgency","doctor"]
                w=csv.DictWriter(f, fieldnames=cols)
                w.writeheader(); w.writerows(rows)
            QtWidgets.QMessageBox.information(self,"ส่งออกแล้ว",path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self,"ผิดพลาด",str(e))

    # ---------- notify ----------
    def _notify(self, title:str, msg:str):
        try: self.tray.showMessage(title, msg, QtWidgets.QSystemTrayIcon.Information, 3000)
        except Exception: pass

    def _set_result_title(self):
        now = datetime.now()
        txt = f"ตารางการผ่าตัด ประจำวัน ({now:%d/%m/%Y}) เวลา {now:%H:%M} น. ห้องผ่าตัดโรงพยาบาลหนองบัวลำภู"
        self.result_banner.set_icon("📁")
        self.result_banner.set_title(txt)
        self.result_banner.set_subtitle("ห้องผ่าตัดโรงพยาบาลหนองบัวลำภู")
        self.card_result.title_lbl.setText(txt)

    # ---------- seq watcher ----------
    def _check_seq(self):
        cur=self.sched.seq()
        if cur!=self.seq_seen:
            self.seq_seen=cur
            self.sched.entries=self.sched._load()
            self.sched.or_rooms=self.sched._load_or()
            self._refresh_or_cb(self.cb_or)
            self._render_tree2()

class WrapItemDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter, option, index):
        text = index.data(QtCore.Qt.DisplayRole)
        opt = QtWidgets.QStyleOptionViewItem(option); self.initStyleOption(opt, index); opt.text=""
        style = opt.widget.style() if opt.widget else QtWidgets.QApplication.style()
        style.drawControl(QtWidgets.QStyle.CE_ItemViewItem, opt, painter, opt.widget)
        rect = style.subElementRect(QtWidgets.QStyle.SE_ItemViewItemText, opt, opt.widget)
        doc = QtGui.QTextDocument(); doc.setDefaultFont(opt.font)
        topt = QtGui.QTextOption(); topt.setWrapMode(QtGui.QTextOption.WordWrap); doc.setDefaultTextOption(topt)
        doc.setTextWidth(rect.width()); doc.setPlainText(str(text) if text is not None else "")
        painter.save(); painter.translate(rect.topLeft()); doc.drawContents(painter, QtCore.QRectF(0,0,rect.width(),rect.height())); painter.restore()
    def sizeHint(self, option, index):
        text = index.data(QtCore.Qt.DisplayRole) or ""
        doc = QtGui.QTextDocument(); doc.setDefaultFont(option.font)
        topt = QtGui.QTextOption(); topt.setWrapMode(QtGui.QTextOption.WordWrap); doc.setDefaultTextOption(topt)
        # ใช้ความกว้างคอลัมน์จริงของ tree เพื่อลดปัญหาความสูงประเมินต่ำ
        tree = option.widget if isinstance(option.widget, QtWidgets.QTreeWidget) else None
        col_w = tree.columnWidth(index.column()) if tree else option.rect.width()
        # เผื่อระยะขอบนิดหน่อย
        w = max(120, int(col_w) - 12)
        doc.setTextWidth(w)
        doc.setPlainText(str(text))
        s = doc.size()
        return QtCore.QSize(w, int(s.height()) + 12)

class SearchSelectAdder(QtWidgets.QWidget):
    """Searchable selector with a multi-select list.

    - Enter / ปุ่ม "➕ เพิ่ม"  : เพิ่มลงรายการของเคส (ไม่แตะคลังหลัก)
    - ปุ่ม "💾 บันทึกเป็นรายการใหม่" : ส่งสัญญาณให้ภายนอกบันทึกเข้าคลังหลัก
    """

    itemsChanged = QtCore.Signal(list)
    requestPersist = QtCore.Signal(str)

    def __init__(self, placeholder="ค้นหา ICD-10...", suggestions=None, parent=None):
        super().__init__(parent)
        v = QtWidgets.QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)

        row = QtWidgets.QHBoxLayout()
        row.setSpacing(6)
        self.combo = NoWheelComboBox()  # กัน scroll เปลี่ยนค่าโดยไม่ตั้งใจ
        self.combo.setEditable(True)
        self.combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo.setMinimumWidth(280)
        self.combo.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.search_line = self.combo.lineEdit()
        if self.search_line:
            self.search_line.setPlaceholderText(placeholder)

        self.btn_add = QtWidgets.QPushButton("➕ เพิ่ม")
        self.btn_add.setProperty("variant", "ghost")
        self.btn_persist = QtWidgets.QPushButton("💾 บันทึกเป็นรายการใหม่")
        self.btn_persist.setProperty("variant", "ghost")

        row.addWidget(self.combo, 1)
        row.addWidget(self.btn_add)
        row.addWidget(self.btn_persist)
        v.addLayout(row)

        self.list = QtWidgets.QListWidget()
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.list.setStyleSheet("QListWidget{ border:1px dashed #e6eaf2; border-radius:12px; background:#fff; }")
        v.addWidget(self.list)

        self._completer: Optional[QtWidgets.QCompleter] = None

        self.set_suggestions(suggestions or [])

        # --- signals: รองรับทั้ง Enter / คลิกคอมโบ / เลือกจากคอมพลีทเตอร์ ---
        if self.search_line:
            self.search_line.returnPressed.connect(self._add_current)

        self.btn_add.clicked.connect(self._add_current)
        self.btn_persist.clicked.connect(self._persist_current)

        self.list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.list.customContextMenuRequested.connect(self._ctx_menu)
        model = self.list.model()
        model.rowsInserted.connect(lambda *_: self._emit_items_changed())
        model.rowsRemoved.connect(lambda *_: self._emit_items_changed())

    def _ctx_menu(self, pos):
        menu = QtWidgets.QMenu(self)
        a1 = menu.addAction("ลบรายการที่เลือก")
        a2 = menu.addAction("ลบทั้งหมด")
        act = menu.exec(self.list.mapToGlobal(pos))
        if act == a1:
            for it in self.list.selectedItems():
                self.list.takeItem(self.list.row(it))
        elif act == a2:
            self.list.clear()
        self._emit_items_changed()

    def _add_text(self, text: str):
        if not text:
            return
        current_lower = [self.list.item(i).text().lower().strip() for i in range(self.list.count())]
        if text.lower().strip() not in current_lower:
            self.list.addItem(text)
        self.combo.setCurrentIndex(0)
        self.combo.setEditText("")
        self._emit_items_changed()

    def _add_current(self):
        self._add_text(self.combo.currentText().strip())

    def _persist_current(self):
        text = self.combo.currentText().strip()
        if text:
            self.requestPersist.emit(text)

    def items(self) -> List[str]:
        return [self.list.item(i).text().strip() for i in range(self.list.count())]

    def clear(self):
        self.list.clear()
        self.combo.setCurrentIndex(0)
        self.combo.setEditText("")
        self._emit_items_changed()

    def set_suggestions(self, suggestions):
        seen = set()
        options: List[str] = []
        for value in suggestions or []:
            val = (value or "").strip()
            if not val or val in seen:
                continue
            seen.add(val)
            options.append(val)

        current_text = self.search_line.text() if self.search_line else ""

        self.combo.blockSignals(True)
        self.combo.clear()
        self.combo.addItem("")
        self.combo.addItems(options)
        self.combo.blockSignals(False)

        if self.search_line is not None:
            self.search_line.blockSignals(True)
            self.search_line.setText(current_text)
            self.search_line.setCursorPosition(len(current_text))
            self.search_line.blockSignals(False)

        self._completer = QtWidgets.QCompleter(options)
        self._completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self._completer.setFilterMode(QtCore.Qt.MatchContains)
        self.combo.setCompleter(self._completer)

        # ปิดการเลื่อนด้วยล้อเมาส์บนคอมโบ (กันเปลี่ยนค่าเวลาเลื่อนหน้า)
        self.combo.setFocusPolicy(QtCore.Qt.StrongFocus)

    def _emit_items_changed(self):
        self.itemsChanged.emit(self.items())


def main():
    QLocale.setDefault(QLocale("en_US"))
    app=QtWidgets.QApplication(sys.argv); app.setApplicationName("RegistryPatientConnect"); app.setOrganizationName(ORG_NAME); app.setWindowIcon(_load_app_icon())
    ap=argparse.ArgumentParser(); ap.add_argument("--host",default="127.0.0.1"); ap.add_argument("--port",type=int,default=DEFAULT_PORT); ap.add_argument("--token",default=DEFAULT_TOKEN)
    a=ap.parse_args()
    w=Main(a.host,a.port,a.token); w.show(); sys.exit(app.exec())

if __name__=="__main__":
    main()