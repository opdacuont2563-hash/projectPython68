# -*- coding: utf-8 -*-
"""
(à¸›à¸£à¸±à¸šà¸›à¸£à¸¸à¸‡à¸ˆà¸²à¸ registry_patient_connect.py â€” à¹à¸à¹‰ strike-through logic & à¸›à¸£à¸±à¸šà¸ªà¹„à¸•à¸¥à¹Œà¸•à¸²à¸£à¸²à¸‡)
"""
import os, sys, json, argparse, csv, base64, secrets, hashlib, unicodedata, re
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Set
from datetime import datetime, timedelta, time as dtime, date
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse

import requests

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import QSettings, QUrl, QLocale
from PySide6.QtGui import QIcon, QPixmap, QPainter, QLinearGradient
from PySide6.QtWebSockets import QWebSocket
from PySide6.QtWidgets import QDialog

from .config import CONFIG
from .logging_setup import get_logger
from .workers.io_worker import SESSION_MANAGER
from .icd10_catalog import (
    add_custom_entry,
    diagnosis_suggestions,
    get_custom_entries,
    get_diagnoses,
    get_operations,
)

try:
    from rapidfuzz import fuzz, process  # type: ignore

    _HAS_RAPIDFUZZ = True
except Exception:  # pragma: no cover - optional dependency
    _HAS_RAPIDFUZZ = False


logger = get_logger(__name__)


# ---------------------- Modern theme ----------------------
def apply_modern_theme(widget: QtWidgets.QWidget):
    widget.setStyleSheet("""
    QWidget{font-family:'Segoe UI','Inter','Noto Sans',system-ui;font-size:11pt;color:#0f172a;background:#f4f6fb;}
    QLineEdit, QDateEdit, QTimeEdit, QComboBox, QPlainTextEdit{
        padding:8px 12px;border-radius:12px;border:1px solid #e6eaf2;background:#fff;
    }
    QLineEdit:hover, QDateEdit:hover, QTimeEdit:hover, QComboBox:hover, QPlainTextEdit:hover{border-color:#cfd8e6;}
    QLineEdit:focus, QDateEdit:focus, QTimeEdit:focus, QComboBox:focus, QPlainTextEdit:focus{border:1px solid #7aa2ff;background:#fff;}
    QPushButton{padding:9px 16px;border-radius:14px;border:1px solid #e7ecf4;background:#fff;font-weight:800;}
    QPushButton:hover{background:#f7faff;border-color:#d9e4fb;}
    QPushButton:pressed{background:#eef4ff;}
    QPushButton[variant="primary"]{background:#2563eb;border-color:#2563eb;color:#fff;}
    QPushButton[variant="primary"]:hover{background:#1d4ed8;}
    QPushButton[variant="danger"]{background:#ef4444;border-color:#ef4444;color:#fff;}
    QPushButton[variant="danger"]:hover{background:#dc2626;}
    QPushButton[variant="ghost"]{background:transparent;border-color:#e7ecf4;color:#0f172a;}
    QTableWidget, QTreeWidget{background:#fff;border:1px solid #e6e6ef;border-radius:12px;gridline-color:#e6e6ef;selection-background-color:#e0f2fe;}
    QTreeWidget::item{height:36px;}
    QTreeWidget::item:hover{ background: rgba(2,132,199,0.06); }
    QHeaderView::section{background:#f1f5f9;border-bottom:1px solid #e6eaf2;padding:10px 12px;font-weight:900;color:#0f172a;}
    QLabel[role='t']{ font-weight:900; font-size:16pt; letter-spacing:.2px; }
    QLabel[role='s']{ color:#64748b; font-size:10pt; }
    QLabel[role='h']{
        font-weight:900; font-size:12.5pt; color:#0f2167; letter-spacing:.2px;
    }
    QFrame#SectionHeader{
        background:#eef4ff; border:1px solid #dbeafe; border-radius:12px; padding:6px 10px;
    }
    QLabel[hint="1"]{color:#64748b;}
    QLabel[warn="1"]{color:#b91c1c;font-weight:700;}
    TabWidget::pane{border:0;}
    QTabBar::tab{
        padding:10px 16px;border-radius:12px;margin:4px;background:#e9eef8;font-weight:700;color:#0f172a;
    }
    QTabBar::tab:hover{ background:#eef3ff; }
    QTabBar::tab:selected{ background:#2563eb;color:#fff; }
    """)


def add_shadow(widget: QtWidgets.QWidget, blur=28, x=0, y=8, color="#24000000"):
    eff = QtWidgets.QGraphicsDropShadowEffect(widget)
    eff.setBlurRadius(blur);
    eff.setOffset(x, y);
    eff.setColor(QtGui.QColor(color))
    widget.setGraphicsEffect(eff)


class NoWheelComboBox(QtWidgets.QComboBox):
    """à¸„à¸­à¸¡à¹‚à¸šà¸—à¸µà¹ˆà¹„à¸¡à¹ˆà¸¢à¸­à¸¡à¹ƒà¸«à¹‰à¹€à¸¡à¸²à¸ªà¹Œà¸ªà¸à¸£à¸­à¸¥à¸¥à¹Œà¹€à¸›à¸¥à¸µà¹ˆà¸¢à¸™à¸„à¹ˆà¸² (à¸à¸±à¸™à¹€à¸œà¸¥à¸­à¹€à¸¥à¸·à¹ˆà¸­à¸™)"""

    def wheelEvent(self, e: QtGui.QWheelEvent) -> None:
        e.ignore()  # à¹ƒà¸«à¹‰ scroll à¸—à¸µà¹ˆ parent à¹à¸—à¸™
        return


def make_search_combo(options: list[str]) -> QtWidgets.QComboBox:
    cb = NoWheelComboBox()
    cb.setEditable(True)
    cb.addItems([""] + options)
    cb.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
    comp = QtWidgets.QCompleter(options)
    comp.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
    comp.setFilterMode(QtCore.Qt.MatchContains)
    cb.setCompleter(comp)
    cb.setMinimumWidth(180)
    return cb


def section_header(text: str) -> QtWidgets.QFrame:
    wrap = QtWidgets.QFrame()
    wrap.setObjectName("SectionHeader")
    h = QtWidgets.QHBoxLayout(wrap)
    h.setContentsMargins(10, 6, 10, 6)
    lab = QtWidgets.QLabel(text)
    lab.setProperty("role", "h")
    h.addWidget(lab)
    h.addStretch(1)
    return wrap


class Card(QtWidgets.QFrame):
    def __init__(self, title="", subtitle=""):
        super().__init__()
        self.setObjectName("Card")
        self.setStyleSheet("""
            QFrame#Card { background:#ffffff; border:1px solid #e6eaf2; border-radius:22px; }
            QLabel[role='t'] { font-weight:900; font-size:16pt; letter-spacing:0.2px; }
            QLabel[role='s'] { color:#64748b; font-size:10pt; }
        """)
        v = QtWidgets.QVBoxLayout(self)
        v.setContentsMargins(20, 20, 20, 20);
        v.setSpacing(12)
        self.title_lbl = QtWidgets.QLabel(title);
        self.title_lbl.setProperty("role", "t");
        v.addWidget(self.title_lbl)
        if subtitle:
            s = QtWidgets.QLabel(subtitle);
            s.setProperty("role", "s");
            v.addWidget(s)
        self.body = QtWidgets.QWidget()
        self.grid = QtWidgets.QGridLayout(self.body)
        self.grid.setHorizontalSpacing(14);
        self.grid.setVerticalSpacing(12)
        v.addWidget(self.body);
        add_shadow(self)


class InfoBanner(QtWidgets.QFrame):
    def __init__(self, title: str = "", subtitle: str = "", variant: str = "blue", icon: str = "ðŸ“"):
        super().__init__()
        self.setObjectName("InfoBanner")
        self._variants = {
            "blue": {"bg": "#eaf6ff", "bd": "#cfe4ff", "accent": "#3b82f6"},
            "violet": {"bg": "#f4efff", "bd": "#e0d4ff", "accent": "#7c3aed"},
            "green": {"bg": "#eafaf3", "bd": "#cfeedd", "accent": "#10b981"},
        }
        pal = self._variants.get(variant, self._variants["blue"])

        self.setStyleSheet(
            f"""
        QFrame#InfoBanner {{
            background:{pal['bg']};
            border:1px solid {pal['bd']};
            border-radius:14px;
        }}
        QLabel[role='title']{{ font-weight:900; font-size:14pt; color:#0f172a; letter-spacing:.2px; }}
        QLabel[role='sub']  {{ color:#64748b; font-size:10pt; }}
        """
        )
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(12)

        accent = QtWidgets.QFrame()
        accent.setFixedWidth(8)
        accent.setStyleSheet(f"QFrame{{background:{pal['accent']}; border-radius:8px;}}")
        lay.addWidget(accent, 0)

        inner = QtWidgets.QVBoxLayout()
        inner.setSpacing(4)
        top = QtWidgets.QHBoxLayout()
        top.setSpacing(8)

        self.icon_lbl = QtWidgets.QLabel(icon)
        self.icon_lbl.setStyleSheet("font-size:16pt;")
        self.title_lbl = QtWidgets.QLabel(title)
        self.title_lbl.setProperty("role", "title")
        top.addWidget(self.icon_lbl, 0)
        top.addWidget(self.title_lbl, 1)
        top.addStretch(1)

        self.sub_lbl = QtWidgets.QLabel(subtitle)
        self.sub_lbl.setProperty("role", "sub")

        inner.addLayout(top)
        inner.addWidget(self.sub_lbl)
        lay.addLayout(inner, 1)

        add_shadow(self, blur=30, x=0, y=6, color="#2a000000")

    def set_title(self, text: str):
        self.title_lbl.setText(text or "")

    def set_subtitle(self, text: str):
        self.sub_lbl.setText(text or "")

    def set_icon(self, text: str):
        self.icon_lbl.setText(text or "ðŸ“")


# ---------------------- Diagnosis search helpers ----------------------
_COMBINE_RE = re.compile(r"[\u0300-\u036f\u0E31\u0E34-\u0E3A\u0E47-\u0E4E]")
_NON_ALNUM_RE = re.compile(r"[^0-9A-Za-z\u0E00-\u0E7F]+")


def normalize_text(text: str) -> str:
    if not text:
        return ""
    val = unicodedata.normalize("NFKD", text).lower()
    val = _COMBINE_RE.sub("", val)
    val = _NON_ALNUM_RE.sub(" ", val).strip()
    return " ".join(val.split())


class FastSearchIndex:
    def __init__(self, items: List[str], prefix_len: int = 3):
        self.items: List[str] = items or []
        self.norms: List[str] = [normalize_text(x) for x in self.items]
        self.prefix_len = max(1, int(prefix_len or 1))
        self.prefix_map: Dict[str, List[int]] = {}
        for idx, (raw, norm) in enumerate(zip(self.items, self.norms)):
            keys = set()
            left = (raw.split(" - ", 1)[0] if raw else "")
            for token in (raw, left, norm):
                token_norm = normalize_text(token)
                if not token_norm:
                    continue
                tokens = token_norm.split()
                if not tokens:
                    continue
                first = tokens[0]
                prefixes = {first[: self.prefix_len]}
                if len(first) >= 2:
                    prefixes.add(first[:2])
                for tk in tokens[:3]:
                    prefixes.add(tk[: self.prefix_len])
                keys.update({p for p in prefixes if p})
            for key in keys:
                self.prefix_map.setdefault(key, []).append(idx)

    def search(self, query: str, limit: int = 100) -> List[str]:
        if not self.items:
            return []
        limit = max(1, int(limit or 1))
        q = normalize_text(query)
        if not q:
            return self.items[:limit]

        q_parts = q.split()
        first_key = q_parts[0][: self.prefix_len]
        cand_idx = list(dict.fromkeys(self.prefix_map.get(first_key, [])))
        results: List[Tuple[int, float]] = []

        for idx in cand_idx:
            norm = self.norms[idx]
            if all(part in norm for part in q_parts):
                results.append((idx, 1.0))
                if len(results) >= limit:
                    break

        if len(results) < max(15, limit // 3):
            extra: List[Tuple[int, float]] = []
            cap = min(len(self.items), 5000)
            seen = {idx for idx, _ in results}
            for idx in range(cap):
                if idx in seen:
                    continue
                norm = self.norms[idx]
                if all(part in norm for part in q_parts):
                    extra.append((idx, 0.9))
                    if len(results) + len(extra) >= limit:
                        break
            for idx, score in extra:
                results.append((idx, score))

        if _HAS_RAPIDFUZZ and len(results) < limit:
            seen = {idx for idx, _ in results}
            corpus = {idx: norm for idx, norm in enumerate(self.norms) if idx not in seen}
            try:
                top_matches = process.extract(  # type: ignore[misc]
                    q,
                    corpus,
                    scorer=fuzz.partial_ratio,
                    score_cutoff=65,
                    limit=max(50, limit),
                )
            except Exception:
                top_matches = []
            for _, score, idx in top_matches:
                results.append((idx, score / 100.0))
                if len(results) >= limit:
                    break

        seen_idx = set()
        ordered: List[Tuple[int, float]] = []
        for idx, score in sorted(results, key=lambda item: (-item[1], len(self.items[item[0]]))):
            if idx in seen_idx:
                continue
            seen_idx.add(idx)
            ordered.append((idx, score))
            if len(ordered) >= limit:
                break

        return [self.items[idx] for idx, _ in ordered]


# ---------------------- Config ----------------------
DEFAULT_HOST = CONFIG.client_host
DEFAULT_PORT = CONFIG.client_port
DEFAULT_TOKEN = CONFIG.client_secret

# === Runner pickup service (FastAPI) ===
RUNNER_PORT = CONFIG.runner_port
RUNNER_BASE_DEFAULT = (CONFIG.runner_base_url or "").rstrip("/")
if not RUNNER_BASE_DEFAULT:
    RUNNER_BASE_DEFAULT = f"http://{DEFAULT_HOST}:{RUNNER_PORT}".rstrip("/")
_RUNNER_SCHEME_DEFAULT = urlparse(RUNNER_BASE_DEFAULT).scheme or "http"
RUNNER_UPDATE_API = "/runner/update"
RUNNER_HEALTH_API = "/health"
RUNNER_LIST_API = "/runner/list"
RUNNER_ACK_API = "/runner/ack"
RUNNER_ARRIVE_API = "/runner/arrive"
RUNNER_FINISH_API = "/runner/finish"


def _resolve_runner_base(base_url: Optional[str] = None) -> str:
    raw = (base_url or "").strip()
    if not raw:
        return RUNNER_BASE_DEFAULT
    candidate = raw
    if "://" not in candidate:
        candidate = f"{_RUNNER_SCHEME_DEFAULT}://{candidate}"
    parsed = urlparse(candidate)
    scheme = parsed.scheme or _RUNNER_SCHEME_DEFAULT
    host = parsed.hostname or ""
    if not host:
        return RUNNER_BASE_DEFAULT
    if host in {"0.0.0.0", ""}:
        host = "127.0.0.1"
    port = parsed.port or RUNNER_PORT
    if port:
        return f"{scheme}://{host}:{port}".rstrip("/")
    return f"{scheme}://{host}".rstrip("/")


def runner_health_ok(base_url: Optional[str] = None, timeout: float = 0.8) -> bool:
    try:
        sess = SESSION_MANAGER.get()
        base = _resolve_runner_base(base_url)
        r = sess.get(f"{base}{RUNNER_HEALTH_API}", timeout=timeout)
        return bool(r.ok)
    except requests.RequestException:
        return False


def _pickup_id_for_row(r: dict) -> str:
    day = str(r.get("date") or r.get("à¸§à¸±à¸™à¸—à¸µà¹ˆ") or date.today().isoformat()).strip()
    hn = str(r.get("HN") or r.get("hn") or r.get("patient_id") or "").strip()
    or_room = str(r.get("OR") or r.get("or") or r.get("or_room") or "").strip()
    return f"{day}:{hn}:{or_room}"


RUNNER_STATUS_LABELS = {
    "waiting": "à¸£à¸­à¸£à¸±à¸š",
    "picking": "à¸à¸³à¸¥à¸±à¸‡à¹„à¸›à¸£à¸±à¸š",
    "arrived": "à¸–à¸¶à¸‡ OR",
    "finished": "à¸œà¹ˆà¸²à¸•à¸±à¸”à¹€à¸ªà¸£à¹‡à¸ˆà¹à¸¥à¹‰à¸§",
}

RUNNER_STATUS_COLORS = {
    "waiting": "#64748b",
    "picking": "#f59e0b",
    "arrived": "#16a34a",
    "finished": "#0f172a",
}


def _fetch_runner_status_map(day: str, base_url: Optional[str] = None) -> Dict[str, dict]:
    try:
        base = _resolve_runner_base(base_url)
        resp = requests.get(
            f"{base}{RUNNER_LIST_API}",
            params={"date": day},
            timeout=2.0,
            headers={"Accept": "application/json"},
        )
        resp.raise_for_status()
        payload = resp.json()
        if isinstance(payload, dict):
            for key in ("items", "data", "rows", "list"):
                maybe = payload.get(key)
                if isinstance(maybe, list):
                    payload = maybe
                    break
            else:
                payload = [payload]
        if not isinstance(payload, list):
            return {}
        results: Dict[str, dict] = {}
        for row in payload:
            if not isinstance(row, dict):
                continue
            pid = row.get("pickup_id") or _pickup_id_for_row(row)
            if not pid:
                continue
            results[str(pid)] = row
        return results
    except (requests.RequestException, ValueError):
        return {}


API_HEALTH = "/api/health";
API_LIST = "/api/list";
API_LIST_FULL = "/api/list_full";
API_WS = "/api/ws"

STATUS_COLORS = {
    "à¸£à¸­à¸œà¹ˆà¸²à¸•à¸±à¸”": "#fde047", "à¸à¸³à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”": "#ef4444", "à¸à¸³à¸¥à¸±à¸‡à¸žà¸±à¸à¸Ÿà¸·à¹‰à¸™": "#22c55e",
    "à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸": "#a855f7", "à¹€à¸¥à¸·à¹ˆà¸­à¸™à¸à¸²à¸£à¸œà¹ˆà¸²à¸•à¸±à¸”": "#64748b",
}
PULSE_STATUS = {"à¸à¸³à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”", "à¸à¸³à¸¥à¸±à¸‡à¸žà¸±à¸à¸Ÿà¸·à¹‰à¸™", "à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸"}
DEFAULT_OR_ROOMS = ["OR1", "OR2", "OR3", "OR4", "OR5", "OR6", "OR8"]

# --- à¸ªà¸–à¸²à¸™à¸°à¸ˆà¸²à¸ monitor à¸—à¸µà¹ˆà¹ƒà¸Šà¹‰à¸ˆà¸±à¸šà¹€à¸§à¸¥à¸² / auto-complete ---
STATUS_OP_START = "à¸à¸³à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”"
STATUS_OP_END = "à¸à¸³à¸¥à¸±à¸‡à¸žà¸±à¸à¸Ÿà¸·à¹‰à¸™"
STATUS_RETURNING = "à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸"

WARD_LIST = [
    "â€” à¸à¸£à¸¸à¸“à¸²à¹€à¸¥à¸·à¸­à¸ â€”",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸­à¸ à¸´à¸šà¸²à¸¥à¸ªà¸‡à¸†à¹Œ",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡ à¸Šà¸±à¹‰à¸™ 4",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸à¹à¸¥à¸°à¸‚à¹‰à¸­",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸«à¸à¸´à¸‡",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸Šà¸²à¸¢",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸­à¸²à¸¢à¸¸à¸£à¸à¸£à¸£à¸¡ à¸Šà¸±à¹‰à¸™ 5",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸­à¸²à¸¢à¸¸à¸£à¸à¸£à¸£à¸¡ à¸Šà¸±à¹‰à¸™ 4",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢ICU-MED",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢ ICU à¸£à¸§à¸¡",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸­à¸²à¸¢à¸¸à¸£à¸à¸£à¸£à¸¡à¸«à¸à¸´à¸‡",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸­à¸²à¸¢à¸¸à¸£à¸à¸£à¸£à¸¡à¸Šà¸²à¸¢",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸£à¸§à¸¡à¸™à¹‰à¸³à¹ƒà¸ˆ",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸«à¸™à¸±à¸à¸à¸¸à¸¡à¸²à¸£à¹€à¸§à¸Š",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸«à¸¹ à¸•à¸² à¸„à¸­ à¸ˆà¸¡à¸¹à¸",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸à¸¸à¸¡à¸²à¸£à¹€à¸§à¸Š",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š à¸Šà¸±à¹‰à¸™ 5",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸žà¸´à¹€à¸¨à¸©à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š à¸Šà¸±à¹‰à¸™ 4",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸›à¸£à¸°à¸ªà¸²à¸—à¹à¸¥à¸°à¸ªà¸¡à¸­à¸‡",
    "à¸«à¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š",
    "à¸«à¹‰à¸­à¸‡à¸„à¸¥à¸­à¸”",
    "à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”",
    "à¹à¸œà¸™à¸à¸­à¸¸à¸šà¸±à¸•à¸´à¹€à¸«à¸•à¸¸à¹à¸¥à¸°à¸‰à¸¸à¸à¹€à¸‰à¸´à¸™",
]

WARD_PLACEHOLDER = WARD_LIST[0]

SCRUB_NURSES = [
    "à¸­à¸£à¸¸à¸“à¸µ", "à¸¨à¸´à¸§à¸”à¸²à¸•à¸´à¹Œ", "à¸à¸±à¸à¸à¸“à¸±à¸Š", "à¸Šà¸±à¸à¸à¸²à¸ à¸±à¸„", "à¸ªà¸¸à¸™à¸—à¸£à¸µ", "à¸žà¸´à¸¨à¸¡à¸±à¸¢", "à¹€à¸—à¸§à¸±à¸", "à¸à¸±à¸™à¸•à¹Œà¸žà¸‡à¸©à¹Œ",
    "à¸›à¸™à¸±à¸à¸à¸²", "à¸ªà¸¸à¸ˆà¸´à¸•à¸£à¸²", "à¸Šà¸±à¸¢à¸¢à¸‡à¸„à¹Œ", "à¸ªà¸¸à¸ à¸²à¸§à¸±à¸¥à¸¢à¹Œ", "à¸ˆà¸±à¸™à¸—à¸ˆà¸£", "à¸§à¸£à¸£à¸“à¸´à¸ à¸²", "à¸“à¸±à¸à¸žà¸‡à¸©à¹Œ", "à¸•à¸°à¸§à¸±à¸™",
    "à¸›à¸§à¸µà¸“à¸²", "à¸™à¸´à¸¤à¸¡à¸¥", "à¸›à¸£à¸´à¸à¸à¸²", "à¸ªà¸¢à¸¸à¸¡à¸žà¸£", "à¸ªà¸¸à¸£à¸ªà¸´à¸—à¸˜à¹Œ", "à¸šà¸¸à¸¨à¸£à¸´à¸™à¸—à¸£à¹Œ", "à¸¨à¸´à¸£à¸´à¸à¸±à¸à¸à¸²", "à¸™à¸£à¸²à¸§à¸±à¸•à¸™à¹Œ",
    "à¸šà¸±à¸“à¸‘à¸´à¸•à¸²", "à¸§à¸£à¸£à¸“à¸§à¸´à¸ªà¸²", "à¸Šà¸¥à¸”à¸²", "à¸§à¸£à¸µà¸ªà¸²",
]

ORG_NAME = "ORNBH";
APP_SHARED = "SurgiBotShared";
OR_KEY = "schedule/or_rooms";
ENTRIES_KEY = "schedule/entries";
SEQ_KEY = "schedule/seq"
APP_SETTINGS = "RegistryPatientConnect"
PDPA_ACK_KEY = "pdpa/ack"
SECRET_SALT_KEY = "sec/hn_salt"
FERNET_KEY = "sec/fernet_key"  # à¹€à¸œà¸·à¹ˆà¸­à¸ˆà¸°à¸•à¹ˆà¸­à¸¢à¸­à¸”à¹€à¸‚à¹‰à¸²à¸£à¸«à¸±à¸ªà¸‚à¹‰à¸­à¸„à¸§à¸²à¸¡à¹ƒà¸™à¸­à¸™à¸²à¸„à¸•

DEPT_DOCTORS = {
    "Surgery | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸—à¸±à¹ˆà¸§à¹„à¸›": ["à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™", "à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ à¸žà¸±à¸™à¸˜à¸¸à¹Œà¸žà¸£à¸«à¸¡", "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ",
                                 "à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£", "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ "],
    "Orthopedics | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸à¹à¸¥à¸°à¸‚à¹‰à¸­": ["à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥ à¸­à¸‡à¸„à¹Œà¹‚à¸†à¸©à¸´à¸•", "à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ à¸¨à¸£à¸µà¹‚à¸žà¸™à¸—à¸­à¸‡", "à¸™à¸ž.à¸­à¸³à¸™à¸²à¸ˆ à¸­à¸™à¸±à¸™à¸•à¹Œà¸§à¸±à¸’à¸™à¸à¸¸à¸¥",
                                           "à¸™à¸ž.à¸­à¸ à¸´à¸Šà¸²à¸•à¸´ à¸¥à¸±à¸à¸©à¸“à¸°", "à¸™à¸ž.à¸à¸¤à¸©à¸Žà¸² à¸­à¸´à¹‰à¸‡à¸­à¸³à¸žà¸£", "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ"],
    "Urology | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°": ["à¸žà¸.à¸ªà¸²à¸¢à¸à¸™ à¸šà¸£à¸£à¸“à¸ˆà¸´à¸•à¸£à¹Œ"],
    "ENT | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡ à¹‚à¸ªà¸• à¸¨à¸­ à¸™à¸²à¸ªà¸´à¸": ["à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸² à¹à¸ªà¸™à¸§à¸±à¸™à¸”à¸µ", "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£ à¸«à¸¡à¸§à¸”à¹„à¸˜à¸ªà¸‡", "à¸™à¸ž.à¸§à¸£à¸§à¸´à¸Š à¸žà¸¥à¹€à¸§à¸µà¸¢à¸‡à¸˜à¸£à¸£à¸¡"],
    "Obstetrics-Gynecology | à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š": ["à¸™à¸ž.à¸ªà¸¸à¸£à¸ˆà¸´à¸•à¸•à¹Œ à¸™à¸´à¸¡à¸´à¸•à¸£à¸§à¸‡à¸©à¹Œà¸ªà¸à¸¸à¸¥", "à¸žà¸.à¸‚à¸§à¸±à¸à¸•à¸² à¸—à¸¸à¸™à¸›à¸£à¸°à¹€à¸—à¸·à¸­à¸‡",
                                            "à¸žà¸.à¸§à¸±à¸Šà¸£à¸²à¸ à¸£à¸“à¹Œ à¸­à¸™à¸§à¸±à¸Šà¸Šà¸à¸¸à¸¥", "à¸žà¸.à¸£à¸¸à¹ˆà¸‡à¸¤à¸”à¸µ à¹‚à¸‚à¸¡à¸žà¸±à¸•à¸£", "à¸žà¸.à¸à¸´à¸•à¸´à¸¡à¸™ à¸Šà¸±à¸¢à¸Šà¸™à¸°à¸—à¸£à¸±à¸žà¸¢à¹Œ"],
    "Ophthalmology | à¸ˆà¸±à¸à¸©à¸¸": ["à¸™à¸ž.à¸ªà¸£à¸²à¸§à¸¸à¸˜ à¸ªà¸²à¸£à¸µà¸¢à¹Œ", "à¸žà¸.à¸”à¸§à¸´à¸©à¸² à¸­à¸±à¸‡à¸¨à¸£à¸µà¸›à¸£à¸°à¹€à¸ªà¸£à¸´à¸", "à¸žà¸.à¸ªà¸²à¸§à¸´à¸•à¸£à¸µ à¸–à¸™à¸­à¸¡à¸§à¸‡à¸¨à¹Œà¹„à¸—à¸¢",
                              "à¸žà¸.à¸ªà¸µà¸à¸Šà¸¡à¸žà¸¹ à¸•à¸±à¹‰à¸‡à¸ªà¸±à¸•à¸¢à¸²à¸˜à¸´à¸©à¸à¸²à¸™", "à¸žà¸.à¸™à¸±à¸™à¸—à¹Œà¸™à¸ à¸±à¸ª à¸Šà¸µà¸§à¸°à¹€à¸à¸£à¸µà¸¢à¸‡à¹„à¸à¸£"],
    "Maxillofacial | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£": ["à¸™à¸ž.à¸‰à¸¥à¸­à¸‡à¸£à¸±à¸ à¹€à¸”à¸Šà¸²", "à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸£à¸±à¸‡"],
}

DEPT_KEY_MAP = {
    "Surgery": "Surgery",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸—à¸±à¹ˆà¸§à¹„à¸›": "Surgery",
    "Surgery | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸—à¸±à¹ˆà¸§à¹„à¸›": "Surgery",
    "Orthopedics": "Orthopedics",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸à¹à¸¥à¸°à¸‚à¹‰à¸­": "Orthopedics",
    "Orthopedics | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸à¹à¸¥à¸°à¸‚à¹‰à¸­": "Orthopedics",
    "Urology": "Urology",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°": "Urology",
    "à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°": "Urology",
    "Urology | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°": "Urology",
    "ENT": "ENT",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡ à¹‚à¸ªà¸• à¸¨à¸­ à¸™à¸²à¸ªà¸´à¸": "ENT",
    "à¸«à¸¹ à¸„à¸­ à¸ˆà¸¡à¸¹à¸": "ENT",
    "ENT | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡ à¹‚à¸ªà¸• à¸¨à¸­ à¸™à¸²à¸ªà¸´à¸": "ENT",
    "Obstetrics-Gynecology": "OBGYN",
    "à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š": "OBGYN",
    "Obstetrics-Gynecology | à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š": "OBGYN",
    "Ophthalmology": "Ophthalmology",
    "à¸ˆà¸±à¸à¸©à¸¸": "Ophthalmology",
    "Ophthalmology | à¸ˆà¸±à¸à¸©à¸¸": "Ophthalmology",
    "Maxillofacial": "Maxillofacial",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£": "Maxillofacial",
    "à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸Šà¹ˆà¸­à¸‡à¸›à¸²à¸à¹à¸¥à¸°à¹ƒà¸šà¸«à¸™à¹‰à¸²": "Maxillofacial",
    "Maxillofacial | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£": "Maxillofacial",
}


def _dept_to_specialty_key(label: str) -> str:
    text = (label or "").strip()
    if not text:
        return ""
    parts = [p.strip() for p in text.split("|")]
    for part in parts + [text]:
        if part in DEPT_KEY_MAP:
            return DEPT_KEY_MAP[part]
        lowered = part.lower()
        for name, key in DEPT_KEY_MAP.items():
            if lowered == name.lower():
                return key
    lowered_all = text.lower()
    if "à¸à¸£à¸°à¸”à¸¹à¸" in lowered_all or "ortho" in lowered_all:
        return "Orthopedics"
    if "à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°" in lowered_all or "uro" in lowered_all:
        return "Urology"
    if "à¸ªà¸¹à¸•à¸´" in lowered_all or "ob" in lowered_all:
        return "OBGYN"
    if "à¹‚à¸ªà¸•" in lowered_all or "ent" in lowered_all or "à¸„à¸­" in lowered_all:
        return "ENT"
    if "à¸ˆà¸±à¸à¸©à¸¸" in lowered_all or "oph" in lowered_all:
        return "Ophthalmology"
    if "à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£" in lowered_all or "à¸Šà¹ˆà¸­à¸‡à¸›à¸²à¸" in lowered_all or "max" in lowered_all:
        return "Maxillofacial"
    return ""


class Toast(QtWidgets.QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            QFrame{background:#111827; color:#fff; border-radius:12px; padding:10px 14px;}
            QLabel{color:#fff;}
        """)
        add_shadow(self, blur=30, x=0, y=8, color="#40000000")
        self.lab = QtWidgets.QLabel("", self)
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(14, 10, 14, 10)
        lay.addWidget(self.lab)
        self._anim: Optional[QtCore.QPropertyAnimation] = None
        self.hide()

    def show_toast(self, text: str, msec: int = 2200):
        self.lab.setText(text)
        if self.parent():
            p = self.parent()
            geo = p.geometry()
            self.adjustSize()
            w, h = self.width(), self.height()
            self.setGeometry(geo.width() - w - 24, 24, w, h)

        if self._anim is not None:
            try:
                self._anim.stop()
            except Exception:
                pass
            self._anim.deleteLater()
            self._anim = None

        self.setWindowOpacity(0.0)
        self.show()
        self.raise_()

        anim = QtCore.QPropertyAnimation(self, b"windowOpacity", self)
        anim.setDuration(msec)
        anim.setStartValue(0.0)
        anim.setKeyValueAt(0.1, 1.0)
        anim.setKeyValueAt(0.9, 1.0)
        anim.setEndValue(0.0)

        anim.finished.connect(self.hide)
        self._anim = anim
        anim.start()


class SweetAlert:
    @staticmethod
    def info(parent: QtWidgets.QWidget, title: str, text: str) -> None:
        QtWidgets.QMessageBox.information(parent, title, text)

    @staticmethod
    def success(
            parent: QtWidgets.QWidget,
            title: str,
            text: str,
            auto_close_msec: Optional[int] = None,
    ) -> None:
        box = QtWidgets.QMessageBox(parent)
        box.setIcon(QtWidgets.QMessageBox.Information)
        box.setWindowTitle(title)
        box.setText(f"âœ… {text}")
        box.setStandardButtons(QtWidgets.QMessageBox.Ok)
        box.setDefaultButton(QtWidgets.QMessageBox.Ok)
        if auto_close_msec and auto_close_msec > 0:
            QtCore.QTimer.singleShot(auto_close_msec, box.accept)
        box.exec()

    @staticmethod
    def warning(parent: QtWidgets.QWidget, title: str, text: str) -> None:
        QtWidgets.QMessageBox.warning(parent, title, text)

    @staticmethod
    def confirm(parent: QtWidgets.QWidget, title: str, text: str) -> bool:
        box = QtWidgets.QMessageBox(parent)
        box.setIcon(QtWidgets.QMessageBox.Question)
        box.setWindowTitle(title)
        box.setText(text)
        box.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        box.setDefaultButton(QtWidgets.QMessageBox.No)
        return box.exec() == QtWidgets.QMessageBox.Yes

    @staticmethod
    def loading(parent: QtWidgets.QWidget, text: str = "à¸à¸³à¸¥à¸±à¸‡à¸”à¸³à¹€à¸™à¸´à¸™à¸à¸²à¸£...") -> QtWidgets.QProgressDialog:
        dlg = QtWidgets.QProgressDialog(text, None, 0, 0, parent)
        dlg.setWindowModality(QtCore.Qt.ApplicationModal)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setWindowTitle("à¹‚à¸›à¸£à¸”à¸£à¸­à¸ªà¸±à¸à¸„à¸£à¸¹à¹ˆ")
        return dlg


class StatusChipWidget(QtWidgets.QWidget):
    def __init__(self, text: str, color: str, pulse: bool = False, parent=None):
        super().__init__(parent)
        self._text = text;
        self._color = color;
        self._pulse = pulse
        if pulse:
            self.eff = QtWidgets.QGraphicsOpacityEffect(self);
            self.setGraphicsEffect(self.eff)
            self.anim = QtCore.QPropertyAnimation(self.eff, b"opacity", self)
            self.anim.setDuration(1200);
            self.anim.setStartValue(0.5);
            self.anim.setEndValue(1.0)
            self.anim.setEasingCurve(QtCore.QEasingCurve.InOutQuad);
            self.anim.setLoopCount(-1);
            self.anim.start()

    def minimumSizeHint(self):
        fm = QtGui.QFontMetrics(self.font())
        w = fm.horizontalAdvance(self._text) + 22 + 16
        h = fm.height() + 10
        return QtCore.QSize(w, h)

    def paintEvent(self, e):
        p = QtGui.QPainter(self);
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        rect = self.rect().adjusted(2, 2, -2, -2)
        bg = QtGui.QColor(self._color);
        bg.setAlpha(205)
        p.setPen(QtCore.Qt.NoPen);
        p.setBrush(bg)
        p.drawRoundedRect(rect, 10, 10)
        p.setPen(QtGui.QColor("#ffffff"))
        p.drawText(rect.adjusted(12, 0, -8, 0), QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft, self._text)


class PeriodBadge(QtWidgets.QWidget):
    def __init__(self, text: str, bg: str, parent=None):
        super().__init__(parent)
        self._text = text
        self._bg = bg
        self.setMinimumHeight(28)

    def sizeHint(self):
        fm = QtGui.QFontMetrics(self.font())
        return QtCore.QSize(fm.horizontalAdvance(self._text) + 22, fm.height() + 10)

    def paintEvent(self, _event):
        painter = QtGui.QPainter(self);
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
        rect = self.rect().adjusted(2, 2, -2, -2)
        color = QtGui.QColor(self._bg);
        color.setAlpha(210)
        painter.setPen(QtCore.Qt.NoPen);
        painter.setBrush(color)
        painter.drawRoundedRect(rect, 10, 10)
        painter.setPen(QtGui.QColor("#ffffff"))
        painter.drawText(rect.adjusted(10, 0, -8, 0), QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft, self._text)


class ScheduleEntry(QtCore.QObject):
    def __init__(
            self,
            or_room="",
            dt=None,
            time_str="",
            hn="",
            name="",
            age=0,
            dept="",
            doctor="",
            diags=None,
            ops=None,
            ward="",
            case_size="",
            queue=0,
            period="in",
            urgency="Elective",
            assist1="",
            assist2="",
            scrub="",
            circulate="",
            time_start="",
            time_end="",
            case_uid: str = "",
            version: int = 1,
            state: str = "scheduled",
            returning_started_at: str = "",
            returned_to_ward_at: str = "",
            postop_completed: bool = False,
    ):
        super().__init__()
        self.or_room = or_room
        self.date = dt or datetime.now().date()
        self.time = time_str
        self.hn = (hn or "").strip()
        self.name = (name or "").strip()
        self.age = int(age) if str(age).isdigit() else 0
        self.dept = dept
        self.doctor = doctor
        self.diags = diags or []
        self.ops = ops or []
        self.ward = (ward or "").strip()
        self.case_size = (case_size or "").strip()  # Minor/Major
        self.queue = int(queue) if str(queue).isdigit() else 0
        self.period = period  # "in" | "off"
        self.urgency = (urgency or "Elective")
        self.assist1 = assist1
        self.assist2 = assist2
        self.scrub = scrub
        self.circulate = circulate
        self.time_start = time_start
        self.time_end = time_end
        self.case_uid = case_uid or self._gen_case_uid()
        self.version = int(version or 1)
        self.state = state or "scheduled"
        self.returning_started_at = returning_started_at or ""
        self.returned_to_ward_at = returned_to_ward_at or ""
        self.postop_completed = bool(postop_completed)

    def _gen_case_uid(self) -> str:
        base = f"{self.or_room}|{self.hn}|{self.time}|{self.date}"
        return hashlib.sha1(base.encode("utf-8", "ignore")).hexdigest()

    def to_dict(self):
        return {
            "or": self.or_room,
            "date": str(self.date),
            "time": self.time,
            "hn": self.hn,
            "name": self.name,
            "age": self.age,
            "dept": self.dept,
            "doctor": self.doctor,
            "diags": self.diags,
            "ops": self.ops,
            "ward": self.ward,
            "case_size": self.case_size,
            "queue": self.queue,
            "period": self.period,
            "urgency": self.urgency,
            "assist1": self.assist1,
            "assist2": self.assist2,
            "scrub": self.scrub,
            "circulate": self.circulate,
            "time_start": self.time_start,
            "time_end": self.time_end,
            "case_uid": self.case_uid,
            "version": self.version,
            "state": self.state,
            "returning_started_at": self.returning_started_at,
            "returned_to_ward_at": self.returned_to_ward_at,
            "postop_completed": self.postop_completed,
        }

    @staticmethod
    def from_dict(d: dict):
        try:
            fromiso = datetime.fromisoformat(d.get("date")).date()
        except Exception:
            fromiso = datetime.now().date()
        return ScheduleEntry(
            d.get("or", ""),
            fromiso,
            d.get("time", ""),
            d.get("hn", ""),
            d.get("name", ""),
            d.get("age", 0),
            d.get("dept", ""),
            d.get("doctor", ""),
            d.get("diags", []) or [],
            d.get("ops", []) or [],
            d.get("ward", ""),
            d.get("case_size", ""),
            d.get("queue", 0),
            d.get("period", "in"),
            d.get("urgency", "Elective"),
            d.get("assist1", ""),
            d.get("assist2", ""),
            d.get("scrub", ""),
            d.get("circulate", ""),
            d.get("time_start", ""),
            d.get("time_end", ""),
            d.get("case_uid", ""),
            d.get("version", 1),
            d.get("state", "scheduled"),
            d.get("returning_started_at", ""),
            d.get("returned_to_ward_at", ""),
            bool(d.get("postop_completed", False)),
        )

    def uid(self) -> str:
        return f"{self.or_room}|{self.hn}|{self.time}|{self.date}"


class SharedScheduleModel:
    def __init__(self):
        self.s = QSettings(ORG_NAME, APP_SHARED)
        self.entries = self._load();
        self.or_rooms = self._load_or()
        if not self.s.contains(SEQ_KEY): self.s.setValue(SEQ_KEY, 0)

    def _load(self) -> List[ScheduleEntry]:
        raw = self.s.value(ENTRIES_KEY, []);
        out = []
        if isinstance(raw, list):
            for d in raw:
                if isinstance(d, dict): out.append(ScheduleEntry.from_dict(d))
        return out

    def _save(self):
        self.s.setValue(ENTRIES_KEY, [e.to_dict() for e in self.entries])
        self.s.setValue(SEQ_KEY, int(self.s.value(SEQ_KEY, 0)) + 1);
        self.s.sync()

    def _load_or(self) -> List[str]:
        lst = self.s.value(OR_KEY)
        if not isinstance(lst, list) or not lst: lst = DEFAULT_OR_ROOMS[:]; self.s.setValue(OR_KEY, lst)
        return [str(x) for x in lst]

    def set_or_rooms(self, rooms: List[str]):
        norm = []
        for r in rooms:
            r = r.strip().upper()
            if r and r.startswith("OR") and r != "OR7" and r not in norm: norm.append(r)
        if not norm: norm = DEFAULT_OR_ROOMS[:]
        self.or_rooms = norm;
        self.s.setValue(OR_KEY, norm);
        self.s.setValue(SEQ_KEY, int(self.s.value(SEQ_KEY, 0)) + 1);
        self.s.sync()

    def add(self, e: ScheduleEntry):
        self.entries.append(e); self._save()

    def update(self, idx: int, e: ScheduleEntry):
        if 0 <= idx < len(self.entries): self.entries[idx] = e; self._save()

    def delete(self, idx: int):
        if 0 <= idx < len(self.entries): self.entries.pop(idx); self._save()

    def seq(self) -> int:
        return int(self.s.value(SEQ_KEY, 0))

    def all(self) -> List[ScheduleEntry]:
        return list(self.entries)

    def clear(self) -> int:
        removed = len(self.entries)
        if removed:
            self.entries.clear()
            self._save()
        return removed

    def remove_by_date(self, day: date) -> int:
        before = len(self.entries)
        self.entries = [e for e in self.entries if getattr(e, "date", None) != day]
        removed = before - len(self.entries)
        if removed:
            self._save()
        return removed

    def replace_all(self, items: List[ScheduleEntry]) -> None:
        self.entries = list(items)
        self._save()


class LocalDBLogger:
    def __init__(self, elective_path="schedule_elective.db", emergency_path="schedule_emergency.db"):
        import sqlite3
        self.sqlite3 = sqlite3
        self.conn_e = sqlite3.connect(elective_path)
        self.conn_x = sqlite3.connect(emergency_path)
        self._init(self.conn_e)
        self._init(self.conn_x)

    def _init(self, conn):
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS schedule(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                urgency TEXT,
                period TEXT,
                or_room TEXT,
                date TEXT,
                time TEXT,
                hn TEXT,
                name TEXT,
                age INTEGER,
                dept TEXT,
                doctor TEXT,
                diagnosis TEXT,
                operation TEXT,
                ward TEXT,
                queue INTEGER,
                time_start TEXT,
                time_end TEXT,
                case_size TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS surgery_events(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                case_uid TEXT,
                event TEXT,
                at TEXT,
                details TEXT
            )
            """
        )
        conn.commit()

    def append_entry(self, e: 'ScheduleEntry'):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = (
            ts,
            e.urgency,
            e.period,
            e.or_room,
            str(e.date),
            e.time,
            e.hn,
            e.name,
            int(e.age or 0),
            e.dept,
            e.doctor,
            " with ".join(e.diags),
            " with ".join(e.ops),
            e.ward,
            int(e.queue or 0),
            e.time_start or "",
            e.time_end or "",
            e.case_size or "",
        )

        conn = self.conn_x if str(e.urgency).lower() == "emergency" else self.conn_e
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO schedule(
                timestamp, urgency, period, or_room, date, time,
                hn, name, age, dept, doctor,
                diagnosis, operation, ward, queue,
                time_start, time_end, case_size
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            row,
        )
        conn.commit()

    def log_event(self, case_uid: str, event: str, details: Optional[dict] = None, emergency: bool = False):
        conn = self.conn_x if emergency else self.conn_e
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO surgery_events(case_uid,event,at,details) VALUES(?,?,?,?)",
            (case_uid, event, _now_iso(), json.dumps(details or {})),
        )
        conn.commit()


# ---------------------- Security helpers (salt & hash) ----------------------
def _app_settings() -> QSettings:
    # à¹ƒà¸Šà¹‰ settings à¸Šà¸¸à¸”à¹€à¸”à¸µà¸¢à¸§à¸à¸±à¸šà¸•à¸±à¸§à¹à¸­à¸› à¹€à¸žà¸·à¹ˆà¸­à¹€à¸à¹‡à¸š salt/ack
    return QSettings(ORG_NAME, APP_SETTINGS)


def _get_or_create_secret(key: str, nbytes: int = 32) -> str:
    s = _app_settings()
    if not s.contains(key):
        # à¹ƒà¸Šà¹‰ urlsafe token à¹€à¸žà¸·à¹ˆà¸­ copy/backup à¹„à¸”à¹‰à¸‡à¹ˆà¸²à¸¢
        tok = secrets.token_urlsafe(nbytes)
        s.setValue(key, tok);
        s.sync()
    return str(s.value(key))


def hn_hash(hn: str) -> str:
    """De-identified hash à¸‚à¸­à¸‡ HN: SHA-256(HN + salt)"""
    salt = _get_or_create_secret(SECRET_SALT_KEY, 32)
    x = (str(hn) + salt).encode("utf-8", "ignore")
    return hashlib.sha256(x).hexdigest()


# (à¸žà¸·à¹‰à¸™à¸—à¸µà¹ˆà¸•à¹ˆà¸­à¸¢à¸­à¸”: à¸–à¹‰à¸²à¸•à¹‰à¸­à¸‡à¸à¸²à¸£à¹€à¸‚à¹‰à¸²à¸£à¸«à¸±à¸ªà¸Šà¸·à¹ˆà¸­/à¸«à¸¡à¸²à¸¢à¹€à¸¥à¸‚)
# from cryptography.fernet import Fernet
# def _fernet() -> Fernet:
#     key = _get_or_create_secret(FERNET_KEY, 32)
#     # Fernet key à¸•à¹‰à¸­à¸‡à¹€à¸›à¹‡à¸™ base64 32 bytes â†’ à¹à¸›à¸¥à¸‡à¹ƒà¸«à¹‰à¹€à¸›à¹‡à¸™ 32 bytes à¹à¸¥à¹‰à¸§ b64
#     k = hashlib.sha256(key.encode()).digest()
#     return Fernet(base64.urlsafe_b64encode(k))
# def enc(txt:str)->str: return _fernet().encrypt(txt.encode()).decode()
# def dec(tok:str)->str: return _fernet().decrypt(tok.encode()).decode()

# ---------------------- PDPA / Consent ----------------------
class PDPANoticeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("à¹à¸ˆà¹‰à¸‡à¹€à¸•à¸·à¸­à¸™ PDPA / à¸‚à¹‰à¸­à¸à¸³à¸à¸±à¸šà¸à¸²à¸£à¹ƒà¸Šà¹‰à¸‚à¹‰à¸­à¸¡à¸¹à¸¥")
        self.setModal(True)
        lay = QtWidgets.QVBoxLayout(self)
        text = QtWidgets.QTextEdit(self)
        text.setReadOnly(True)
        text.setMinimumHeight(220)
        text.setStyleSheet("QTextEdit{background:#fff;border:1px solid #e6eaf2;border-radius:12px;padding:10px;}")
        text.setText(
            "à¸§à¸±à¸•à¸–à¸¸à¸›à¸£à¸°à¸ªà¸‡à¸„à¹Œà¸à¸²à¸£à¹ƒà¸Šà¹‰à¸‚à¹‰à¸­à¸¡à¸¹à¸¥:\n"
            "- à¹ƒà¸Šà¹‰à¹€à¸žà¸·à¹ˆà¸­à¸à¸²à¸£à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™/à¸šà¸£à¸´à¸«à¸²à¸£à¸ˆà¸±à¸”à¸à¸²à¸£à¸„à¸´à¸§à¸œà¹ˆà¸²à¸•à¸±à¸” à¹à¸¥à¸°à¸ªà¸·à¹ˆà¸­à¸ªà¸²à¸£à¸à¸²à¸£à¸—à¸³à¸‡à¸²à¸™à¹ƒà¸™à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”\n"
            "- à¹ƒà¸Šà¹‰à¸ªà¸–à¸´à¸•à¸´à¸ à¸²à¸žà¸£à¸§à¸¡à¹à¸šà¸šà¹„à¸¡à¹ˆà¸£à¸°à¸šà¸¸à¸•à¸±à¸§à¸•à¸™ (de-identified) à¹€à¸žà¸·à¹ˆà¸­à¸›à¸£à¸±à¸šà¸›à¸£à¸¸à¸‡à¸„à¸¸à¸“à¸ à¸²à¸žà¸šà¸£à¸´à¸à¸²à¸£ (QI)\n\n"
            "à¸à¸²à¸£à¸„à¸¸à¹‰à¸¡à¸„à¸£à¸­à¸‡à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ªà¹ˆà¸§à¸™à¸šà¸¸à¸„à¸„à¸¥ (PDPA):\n"
            "- à¹€à¸à¹‡à¸šà¹€à¸—à¹ˆà¸²à¸—à¸µà¹ˆà¸ˆà¸³à¹€à¸›à¹‡à¸™ (data minimization)\n"
            "- à¸¡à¸µà¸›à¸¸à¹ˆà¸¡ Export à¹à¸šà¸šà¹„à¸¡à¹ˆà¸£à¸°à¸šà¸¸à¸•à¸±à¸§à¸•à¸™ (à¹à¸®à¸Š HN) à¸ªà¸³à¸«à¸£à¸±à¸šà¸‡à¸²à¸™à¸§à¸´à¹€à¸„à¸£à¸²à¸°à¸«à¹Œ\n"
            "- à¸«à¹‰à¸²à¸¡à¸ªà¹ˆà¸‡à¸­à¸­à¸/à¸–à¹ˆà¸²à¸¢à¹‚à¸­à¸™à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸—à¸µà¹ˆà¸£à¸°à¸šà¸¸à¸•à¸±à¸§à¸šà¸¸à¸„à¸„à¸¥à¹‚à¸”à¸¢à¹„à¸¡à¹ˆà¹„à¸”à¹‰à¸£à¸±à¸šà¸­à¸™à¸¸à¸à¸²à¸•\n"
            "- à¸à¸²à¸£à¸šà¸±à¸™à¸—à¸¶à¸ Log à¸ˆà¸°à¹„à¸¡à¹ˆà¹€à¸à¹‡à¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸—à¸µà¹ˆà¸£à¸°à¸šà¸¸à¸•à¸±à¸§à¸šà¸¸à¸„à¸„à¸¥à¹‚à¸”à¸¢à¹„à¸¡à¹ˆà¸ˆà¸³à¹€à¸›à¹‡à¸™\n\n"
            "à¸à¸²à¸£à¸”à¸³à¹€à¸™à¸´à¸™à¸à¸²à¸£à¸•à¹ˆà¸­à¸–à¸·à¸­à¸§à¹ˆà¸²à¸—à¹ˆà¸²à¸™à¹€à¸‚à¹‰à¸²à¹ƒà¸ˆà¹à¸¥à¸°à¸¢à¸­à¸¡à¸£à¸±à¸šà¸•à¸²à¸¡à¸‚à¹‰à¸²à¸‡à¸•à¹‰à¸™"
        )
        chk = QtWidgets.QCheckBox("à¸‰à¸±à¸™à¸­à¹ˆà¸²à¸™à¹à¸¥à¸°à¸¢à¸­à¸¡à¸£à¸±à¸šà¸à¸²à¸£à¹ƒà¸Šà¹‰à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸•à¸²à¸¡ PDPA à¹à¸¥à¹‰à¸§")
        btn = QtWidgets.QPushButton("à¸•à¸à¸¥à¸‡");
        btn.setProperty("variant", "primary");
        btn.setEnabled(False)
        chk.toggled.connect(lambda b: btn.setEnabled(b))
        btn.clicked.connect(self.accept)
        lay.addWidget(text);
        lay.addWidget(chk);
        lay.addWidget(btn)


def _fmt_td(td: timedelta) -> str:
    total = int(abs(td.total_seconds()));
    h = total // 3600;
    m = (total % 3600) // 60;
    s = total % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _parse_iso(ts: str):
    if not ts: return None
    try:
        return datetime.fromisoformat(ts.replace("Z", ""))
    except Exception:
        return None


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _is_postop_complete_entry(e: "ScheduleEntry") -> bool:
    if not (e.time_start and e.time_end):
        return False
    try:
        hh1, mm1 = e.time_start.split(":")
        hh2, mm2 = e.time_end.split(":")
        t1 = int(hh1) * 60 + int(mm1)
        t2 = int(hh2) * 60 + int(mm2)
        if t2 < t1:
            return False
    except Exception:
        return False
    if not (e.scrub or e.circulate or e.assist1 or e.assist2):
        return False
    if not (e.ops or e.diags):
        return False
    return True


def _app_icon() -> QIcon:
    pm = QtGui.QPixmap(64, 64);
    pm.fill(QtCore.Qt.transparent)
    pa = QPainter(pm);
    pa.setRenderHint(QtGui.QPainter.Antialiasing, True)
    gr = QLinearGradient(0, 0, 64, 64)
    gr.setColorAt(0, "#d9ecff");
    gr.setColorAt(.55, "#e1f5ff");
    gr.setColorAt(1, "#e6fff5")
    pa.setBrush(gr);
    pa.setPen(QtCore.Qt.NoPen);
    pa.drawEllipse(6, 6, 52, 52)
    pa.setBrush(QtGui.QColor("#0f172a"))
    pa.drawEllipse(26, 20, 12, 12)
    pa.end()
    return QIcon(pm)


def _load_app_icon() -> QIcon:
    p = Path("assets/app.ico")
    if p.exists():
        ico = QIcon(str(p))
        if not ico.isNull():
            return ico
    return _app_icon()


def _now_period(dt_val: datetime) -> str:
    start = dtime(8, 30);
    end = dtime(16, 30)
    return "in" if (start <= dt_val.time() < end) else "off"


def _period_label(code: str) -> str: return "à¹ƒà¸™à¹€à¸§à¸¥à¸²à¸£à¸²à¸Šà¸à¸²à¸£" if code == "in" else "à¸™à¸­à¸à¹€à¸§à¸¥à¸²à¸£à¸²à¸Šà¸à¸²à¸£"


def _period_badge(period_code: str) -> PeriodBadge:
    label = "à¹ƒà¸™à¹€à¸§à¸¥à¸²à¸£à¸²à¸Šà¸à¸²à¸£" if (period_code or "").lower() == "in" else "à¸™à¸­à¸à¹€à¸§à¸¥à¸²à¸£à¸²à¸Šà¸à¸²à¸£"
    color = "#2563eb" if (period_code or "").lower() == "in" else "#64748b"
    return PeriodBadge(label, color)


# ---------------------- Wednesday OR ownership helpers ----------------------
OWNER_WED_DOCTOR2OR = {
    "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™": "OR1",
    "à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£": "OR6",
}


def _owner_variants(name: str) -> Set[str]:
    """Return a set of alias variants that should map to the canonical name."""
    variants: Set[str] = {name}
    normalized_target = " ".join(str(name or "").split())
    aliases = globals().get("DOCTOR_ALIASES", {})
    if isinstance(aliases, dict):
        for alias, canonical in aliases.items():
            if " ".join(str(canonical or "").split()) == normalized_target:
                variants.add(alias)
    return variants


def _infer_doctor_from_entry(entry: "ScheduleEntry") -> str:
    """Extract the best-effort normalized doctor name from the entry."""
    raw = getattr(entry, "doctor", "") or ""
    who = normalize_doctor_name(raw)
    if who:
        return who

    blobs: List[str] = []
    ops = getattr(entry, "ops", None)
    diags = getattr(entry, "diags", None)
    if isinstance(ops, (list, tuple)):
        blobs.append(" ".join(str(x) for x in ops))
    elif isinstance(ops, str):
        blobs.append(ops)
    if isinstance(diags, (list, tuple)):
        blobs.append(" ".join(str(x) for x in diags))
    elif isinstance(diags, str):
        blobs.append(diags)

    text = " ".join(blobs)
    for canonical in OWNER_WED_DOCTOR2OR.keys():
        variants = _owner_variants(canonical)
        if any(var and var in text for var in variants):
            return normalize_doctor_name(canonical)
    return ""


def normalize_owner_for_wednesday(entries: List["ScheduleEntry"], dt: date) -> List["ScheduleEntry"]:
    """Ensure Wednesday cases stay with their designated room owners."""
    if not dt or dt.weekday() != 2:
        return entries

    for entry in entries:
        who = _infer_doctor_from_entry(entry)
        for owner_name, target_or in OWNER_WED_DOCTOR2OR.items():
            if normalize_doctor_name(owner_name) == normalize_doctor_name(who):
                if getattr(entry, "or_room", None) != target_or:
                    setattr(entry, "or_room", target_or)
                break
    return entries


def _span_first_column(item: Optional[QtWidgets.QTreeWidgetItem]) -> None:
    """Helper to span the first column on a tree item (PySide6-compatible)."""
    if item is None:
        return
    try:
        item.setFirstColumnSpanned(True)
    except AttributeError:
        pass


class ClientHTTP:
    def __init__(self, host=DEFAULT_HOST, port=DEFAULT_PORT, token=DEFAULT_TOKEN, timeout: float | None = None):
        self.base = f"http://{host}:{port}"
        self.token = token
        self.timeout = timeout or CONFIG.request_timeout
        self.s = SESSION_MANAGER.get()

    def health(self):
        r = self.s.get(self.base + API_HEALTH, timeout=self.timeout, headers={"Accept": "application/json"})
        r.raise_for_status();
        return r.json()

    def list_items(self):
        try:
            r = self.s.get(f"{self.base}{API_LIST_FULL}?token={self.token}", timeout=self.timeout,
                           headers={"Accept": "application/json"})
            if r.status_code == 200: return self._wrap(r.json())
        except Exception:
            pass
        try:
            r = self.s.get(self.base + API_LIST, timeout=self.timeout, headers={"Accept": "application/json"})
            if r.status_code == 200: return self._wrap(r.json())
        except Exception:
            pass
        return {"items": []}

    @staticmethod
    def _wrap(d):
        if isinstance(d, list): return {"items": d}
        if isinstance(d, dict):
            for k in ("items", "data", "table", "rows", "list"):
                if k in d and isinstance(d[k], list): return {"items": d[k]}
            for v in d.values():
                if isinstance(v, list): return {"items": v}
            return d
        return {"items": []}


def extract_rows(payload):
    if isinstance(payload, list):
        src = payload
    elif isinstance(payload, dict):
        for k in ("items", "data", "table", "rows", "list"):
            if k in payload and isinstance(payload[k], list): src = payload[k]; break
        else:
            src = next((v for v in payload.values() if isinstance(v, list)), [])
    else:
        src = []
    rows = []
    for i, it in enumerate(src, 1):
        if not isinstance(it, dict): continue
        hn = str(it.get("hn_full") or it.get("hn") or "").strip()
        pid = str(it.get("patient_id") or it.get("pid") or it.get("queue_id") or "")
        if not pid:
            orr = str(it.get("or") or it.get("or_room") or "");
            q = str(it.get("queue") or it.get("q") or "")
            if orr and q: pid = f"{orr}-{q}"
        status = str(it.get("status") or "")
        ts = (it.get("timestamp") or it.get("ts") or it.get("updated_at") or it.get("created_at") or it.get("time"))
        eta = it.get("eta_minutes", it.get("eta", it.get("eta_min")))
        if isinstance(eta, str) and eta.isdigit():
            eta = int(eta)
        elif not isinstance(eta, int):
            eta = None
        rows.append({"id": hn if hn else i, "hn_full": hn or None, "patient_id": pid, "status": status, "timestamp": ts,
                     "eta_minutes": eta})
    return rows


class QueueSelectWidget(QtWidgets.QWidget):
    changed = QtCore.Signal(int)

    def __init__(self, value: int = 0, parent=None):
        super().__init__(parent)
        h = QtWidgets.QHBoxLayout(self);
        h.setContentsMargins(2, 0, 2, 0);
        h.setSpacing(6)
        self.combo = QtWidgets.QComboBox()
        self.combo.addItem("à¸•à¸²à¸¡à¹€à¸§à¸¥à¸²", 0)
        for i in range(1, 10): self.combo.addItem(str(i), i)
        self.combo.setStyleSheet("""
            QComboBox{
                padding:6px 10px;border-radius:10px;border:1px solid #e6eaf2;background:#ffffff;
                min-width: 110px; font-weight:700;
            }
            QComboBox:hover{border-color:#cfd8e6;}
            QComboBox:focus{border:1px solid #7aa2ff;background:#ffffff;}
        """)
        idx = 0
        for i in range(self.combo.count()):
            if int(self.combo.itemData(i)) == int(value): idx = i; break
        self.combo.setCurrentIndex(idx)
        self.combo.currentIndexChanged.connect(self._emit_changed)
        h.addWidget(self.combo, 1)
        self.setMinimumWidth(120)
        self.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)

    def _emit_changed(self, _i):
        self.changed.emit(int(self.combo.currentData() or 0))

    def value(self) -> int:
        return int(self.combo.currentData() or 0)


# --------------------------- Fixed Excel import helpers ---------------------------
FIXED_MAPPING_TH = {
    "time": "à¸ªà¸±à¹ˆà¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¹€à¸§à¸¥à¸²",
    "hn": "HN",
    "name": "à¸Šà¸·à¹ˆà¸­",
    "age": "à¸­à¸²à¸¢à¸¸",
    "diags": "ICD Name",
    "ops": "à¸Šà¸·à¹ˆà¸­à¸à¸²à¸£à¸œà¹ˆà¸²à¸•à¸±à¸”",
    "doctor": "à¹à¸žà¸—à¸¢à¹Œà¸œà¸¹à¹‰à¸ªà¸±à¹ˆà¸‡",
    "ward": "Ward",
}

_re_hhmm = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*$")
_re_hhmmss = re.compile(r"^\s*(\d{1,2}):(\d{2}):(\d{2})\s*$")
_re_dt_hhmm = re.compile(r".*?(\d{1,2}):(\d{2})(?::\d{2})?\s*$")
_year_re = re.compile(r"(\d+)\s*à¸›à¸µ")


def _excel_time_to_hhmm(value: float) -> str:
    try:
        fraction = float(value)
    except Exception:
        return ""
    if not (0 <= fraction < 1.1):
        return ""
    total_seconds = int(round(fraction * 24 * 3600))
    hours = (total_seconds // 3600) % 24
    minutes = (total_seconds // 60) % 60
    return f"{hours:02d}:{minutes:02d}"


def parse_time_hhmm_or_tf(raw_value) -> str:
    """à¹à¸›à¸¥à¸‡à¸„à¹ˆà¸²à¸ˆà¸²à¸ Excel/à¸‚à¹‰à¸­à¸„à¸§à¸²à¸¡à¹ƒà¸«à¹‰à¹€à¸›à¹‡à¸™ HH:MM à¸«à¸£à¸·à¸­ 'TF' à¸«à¸²à¸à¹„à¸¡à¹ˆà¸—à¸£à¸²à¸šà¹€à¸§à¸¥à¸²"""
    if raw_value is None:
        return "TF"

    if isinstance(raw_value, (int, float)):
        return _excel_time_to_hhmm(raw_value) or "TF"

    if isinstance(raw_value, datetime):
        return f"{raw_value.hour:02d}:{raw_value.minute:02d}"
    if isinstance(raw_value, dtime):
        return f"{raw_value.hour:02d}:{raw_value.minute:02d}"

    text = str(raw_value).strip()
    if not text:
        return "TF"

    match = _re_hhmmss.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    match = _re_hhmm.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    match = _re_dt_hhmm.match(text)
    if match:
        hh, mm = int(match.group(1)), int(match.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    return "TF"


def parse_age_years(txt: str) -> int:
    """à¸”à¸¶à¸‡à¹€à¸‰à¸žà¸²à¸°à¸ˆà¸³à¸™à¸§à¸™ 'à¸›à¸µ' à¹€à¸Šà¹ˆà¸™ '23 à¸›à¸µ 4 à¹€à¸”à¸·à¸­à¸™ 30 à¸§à¸±à¸™' -> 23; '23' -> 23; à¸­à¸·à¹ˆà¸™à¹† -> 0"""
    if not txt:
        return 0
    s = str(txt)
    m = _year_re.search(s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return 0
    try:
        return int(float(s))
    except Exception:
        return 0


def normalize_doctor(txt: str) -> str:
    """à¸„à¸‡à¸„à¸³à¸™à¸³à¸«à¸™à¹‰à¸²à¹à¸¥à¸°à¸ˆà¸±à¸”à¸à¸²à¸£ alias à¸Šà¸·à¹ˆà¸­à¹à¸žà¸—à¸¢à¹Œà¹ƒà¸«à¹‰à¸ªà¸°à¸­à¸²à¸”"""
    return normalize_doctor_name(txt)


def map_to_known_ward(src: str, known_wards: List[str]) -> str:
    """
    à¸ˆà¸±à¸šà¸„à¸¹à¹ˆà¸Šà¸·à¹ˆà¸­à¸§à¸­à¸£à¹Œà¸”à¹ƒà¸«à¹‰à¸•à¸£à¸‡à¸à¸±à¸šà¸£à¸²à¸¢à¸à¸²à¸£à¸—à¸µà¹ˆà¹à¸­à¸›à¸¡à¸µà¸­à¸¢à¸¹à¹ˆà¹à¸¥à¹‰à¸§
    à¸à¸¥à¹„à¸: à¹€à¸—à¸µà¸¢à¸šà¹à¸šà¸š case-insensitive + à¸•à¸±à¸”à¸Šà¹ˆà¸­à¸‡à¸§à¹ˆà¸²à¸‡à¹€à¸à¸´à¸™ + à¸«à¸² 'à¸„à¸µà¸¢à¹Œà¹€à¸§à¸´à¸£à¹Œà¸”à¸«à¸¥à¸±à¸'
    à¸„à¸¸à¸“à¸ªà¸²à¸¡à¸²à¸£à¸–à¸›à¸£à¸±à¸š synonyms à¹„à¸”à¹‰à¸•à¸²à¸¡à¸Šà¸·à¹ˆà¸­à¹ƒà¸™à¸£à¸°à¸šà¸šà¸ˆà¸£à¸´à¸‡
    """

    s = " ".join((src or "").lower().split())
    synonyms = {
        "à¸«à¸¹à¸„à¸­à¸ˆà¸¡à¸¹à¸": ["à¸«à¸¹ à¸„à¸­ à¸ˆà¸¡à¸¹à¸", "ent", "à¹‚à¸ªà¸•à¸¨à¸­à¸™à¸²à¸ªà¸´à¸"],
    }

    for w in known_wards:
        if s == " ".join(w.lower().split()):
            return w

    for canonical, words in synonyms.items():
        for kw in words:
            if kw.replace(" ", "") in s.replace(" ", ""):
                for w in known_wards:
                    if canonical.replace(" ", "") in w.replace(" ", "").lower():
                        return w
                return src

    for w in known_wards:
        lowered = w.lower()
        if any(token and token in s for token in lowered.split()):
            return w

    return src


# ================== CONFIG: à¸•à¸²à¸£à¸²à¸‡à¹à¸žà¸—à¸¢à¹Œà¸›à¸£à¸°à¸ˆà¸³ OR ==================
# à¸Ÿà¸­à¸£à¹Œà¹à¸¡à¸•:
# WEEKLY_DOCTOR_OR_PLAN[weekday]["ORx"] = [
#   {"doctor": "à¸Šà¸·à¹ˆà¸­à¹à¸žà¸—à¸¢à¹Œ", "when": "ALLDAY|AM|PM", "weeks": [1,2,3,4]},
#   {"doctor": ["à¸”à¸£.à¸", "à¸”à¸£.à¸‚"], ...}
# ]
# à¸«à¸¡à¸²à¸¢à¹€à¸«à¸•à¸¸: Monday=0 ... Sunday=6
WEEKLY_DOCTOR_OR_PLAN: Dict[int, Dict[str, List[Dict[str, object]]]] = {
  0: {
    'OR1': [{'doctor': 'à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™', 'when': 'ALLDAY', 'weeks': [1]}, {'doctor': 'à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£', 'when': 'ALLDAY', 'weeks': [2]}, {'doctor': 'à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ', 'when': 'ALLDAY', 'weeks': [3]}, {'doctor': 'à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ à¸žà¸±à¸™à¸˜à¸¸à¹Œà¸žà¸£à¸«à¸¡', 'when': 'ALLDAY', 'weeks': [4]}],
    'OR2': [{'doctor': 'à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ à¸¨à¸£à¸µà¹‚à¸žà¸™à¸—à¸­à¸‡', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR3': [{'doctor': 'à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸² à¹à¸ªà¸™à¸§à¸±à¸™à¸”à¸µ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR5': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR6': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR8': [{'doctor': 'à¸žà¸.à¸ªà¸µà¸Šà¸¡à¸žà¸¹ à¸•à¸±à¹‰à¸‡à¸ªà¸±à¸•à¸¢à¸²à¸˜à¸´à¸©à¸à¸²à¸™', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
  },
  1: {
    'OR1': [{'doctor': 'à¸žà¸.à¸ªà¸²à¸¢à¸à¸™ à¸šà¸£à¸£à¸“à¸ˆà¸´à¸•à¸£à¹Œ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR2': [{'doctor': 'à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥ à¸­à¸‡à¸„à¹Œà¹‚à¸†à¸©à¸´à¸•', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR3': [{'doctor': 'à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ', 'when': 'AM', 'weeks': [1, 2, 3, 4]}, {'doctor': 'à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸£à¸±à¸‡', 'when': 'PM', 'weeks': [1, 2, 3, 4]}],
    'OR5': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR6': [{'doctor': 'à¸™à¸ž.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR8': [{'doctor': 'à¸žà¸.à¸ªà¸²à¸§à¸´à¸•à¸£à¸µ à¸–à¸™à¸­à¸¡à¸§à¸‡à¸¨à¹Œà¹„à¸—à¸¢', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
  },
  2: {
    'OR1': [{'doctor': 'à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR2': [{'doctor': 'à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR3': [{'doctor': 'CLOSED', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR5': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR6': [{'doctor': 'à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR8': [{'doctor': 'à¸žà¸.à¸™à¸±à¸™à¸—à¹Œà¸™à¸ à¸±à¸ª à¸Šà¸µà¸§à¸°à¹€à¸à¸£à¸µà¸¢à¸‡à¹„à¸à¸£', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
  },
  3: {
    'OR1': [{'doctor': 'à¸žà¸.à¸ªà¸²à¸¢à¸à¸™ à¸šà¸£à¸£à¸“à¸ˆà¸´à¸•à¸£à¹Œ', 'when': 'AM', 'weeks': [1, 2, 3, 4]}, {'doctor': 'à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥ à¸­à¸‡à¸„à¹Œà¹‚à¸†à¸©à¸´à¸•', 'when': 'PM', 'weeks': [1, 3]}, {'doctor': ['à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ à¸¨à¸£à¸µà¹‚à¸žà¸™à¸—à¸­à¸‡', 'à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ'], 'when': 'PM', 'weeks': [2, 4]}],
    'OR2': [{'doctor': 'à¸™à¸ž.à¸­à¸³à¸™à¸²à¸ˆ à¸­à¸™à¸±à¸™à¸•à¹Œà¸§à¸±à¸’à¸™à¸à¸¸à¸¥', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR3': [{'doctor': 'à¸™à¸ž.à¸§à¸£à¸§à¸´à¸Š à¸žà¸¥à¹€à¸§à¸µà¸¢à¸‡à¸˜à¸£à¸£à¸¡', 'when': 'AM', 'weeks': [1, 2, 3, 4]}, {'doctor': 'à¸—à¸ž.à¸‰à¸¥à¸­à¸‡à¸£à¸±à¸ à¹€à¸”à¸Šà¸²', 'when': 'PM', 'weeks': [1, 2, 3, 4]}],
    'OR5': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR6': [{'doctor': 'à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ à¸žà¸±à¸™à¸˜à¸¸à¹Œà¸žà¸£à¸«à¸¡', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR8': [{'doctor': 'à¸žà¸.à¸”à¸§à¸´à¸©à¸² à¸­à¸±à¸‡à¸¨à¸£à¸µà¸›à¸£à¸°à¹€à¸ªà¸£à¸´à¸', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
  },
  4: {
    'OR1': [{'doctor': 'à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR2': [{'doctor': 'à¸™à¸ž.à¸à¸¤à¸©à¸Žà¸² à¸­à¸´à¹‰à¸‡à¸­à¸³à¸žà¸£', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR3': [{'doctor': 'à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£ à¸«à¸¡à¸§à¸”à¹„à¸˜à¸ªà¸‡', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR5': [{'doctor': 'OBGYN_ANY', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR6': [{'doctor': 'CLOSED', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
    'OR8': [{'doctor': 'à¸™à¸ž.à¸ªà¸£à¸²à¸§à¸¸à¸˜ à¸ªà¸²à¸£à¸µà¸¢à¹Œ', 'when': 'ALLDAY', 'weeks': [1, 2, 3, 4]}],
  },
}

GROUPS: Dict[str, List[str]] = {
    # Surgery | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸—à¸±à¹ˆà¸§à¹„à¸›
    "SUR_ANY": [
        "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™",
        "à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ à¸žà¸±à¸™à¸˜à¸¸à¹Œà¸žà¸£à¸«à¸¡",
        "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ",
        "à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£",
        "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ",
    ],

    # Orthopedics | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸à¹à¸¥à¸°à¸‚à¹‰à¸­
    "ORTHO_ANY": [
        "à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥ à¸­à¸‡à¸„à¹Œà¹‚à¸†à¸©à¸´à¸•",
        "à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ à¸¨à¸£à¸µà¹‚à¸žà¸™à¸—à¸­à¸‡",
        "à¸™à¸ž.à¸­à¸³à¸™à¸²à¸ˆ à¸­à¸™à¸±à¸™à¸•à¹Œà¸§à¸±à¸’à¸™à¸à¸¸à¸¥",
        "à¸™à¸ž.à¸­à¸ à¸´à¸Šà¸²à¸•à¸´ à¸¥à¸±à¸à¸©à¸“à¸°",
        "à¸™à¸ž.à¸à¸¤à¸©à¸Žà¸² à¸­à¸´à¹‰à¸‡à¸­à¸³à¸žà¸£",
        "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ",
    ],

    # Urology | à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°
    "URO_ANY": [
        "à¸žà¸.à¸ªà¸²à¸¢à¸à¸™ à¸šà¸£à¸£à¸“à¸ˆà¸´à¸•à¸£à¹Œ",
    ],

    # ENT | à¹‚à¸ªà¸• à¸¨à¸­ à¸™à¸²à¸ªà¸´à¸
    "ENT_ANY": [
        "à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸² à¹à¸ªà¸™à¸§à¸±à¸™à¸”à¸µ",
        "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£ à¸«à¸¡à¸§à¸”à¹„à¸˜à¸ªà¸‡",
        "à¸™à¸ž.à¸§à¸£à¸§à¸´à¸Š à¸žà¸¥à¹€à¸§à¸µà¸¢à¸‡à¸˜à¸£à¸£à¸¡",
    ],

    # Obstetrics-Gynecology | à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š
    "OBGYN_ANY": [
        "à¸™à¸ž.à¸ªà¸¸à¸£à¸ˆà¸´à¸•à¸•à¹Œ à¸™à¸´à¸¡à¸´à¸•à¸£à¸§à¸‡à¸©à¹Œà¸ªà¸à¸¸à¸¥",
        "à¸žà¸.à¸‚à¸§à¸±à¸à¸•à¸² à¸—à¸¸à¸™à¸›à¸£à¸°à¹€à¸—à¸·à¸­à¸‡",
        "à¸žà¸.à¸§à¸±à¸Šà¸£à¸²à¸ à¸£à¸“à¹Œ à¸­à¸™à¸§à¸±à¸Šà¸Šà¸à¸¸à¸¥",
        "à¸žà¸.à¸£à¸¸à¹ˆà¸‡à¸¤à¸”à¸µ à¹‚à¸‚à¸¡à¸žà¸±à¸•à¸£",
        "à¸žà¸.à¸à¸´à¸•à¸´à¸¡à¸™ à¸Šà¸±à¸¢à¸Šà¸™à¸°à¸—à¸£à¸±à¸žà¸¢à¹Œ",
    ],

    # Ophthalmology | à¸ˆà¸±à¸à¸©à¸¸
    "EYE_ANY": [
        "à¸™à¸ž.à¸ªà¸£à¸²à¸§à¸¸à¸˜ à¸ªà¸²à¸£à¸µà¸¢à¹Œ",
        "à¸žà¸.à¸”à¸§à¸´à¸©à¸² à¸­à¸±à¸‡à¸¨à¸£à¸µà¸›à¸£à¸°à¹€à¸ªà¸£à¸´à¸",
        "à¸žà¸.à¸ªà¸²à¸§à¸´à¸•à¸£à¸µ à¸–à¸™à¸­à¸¡à¸§à¸‡à¸¨à¹Œà¹„à¸—à¸¢",
        "à¸žà¸.à¸ªà¸µà¸à¸Šà¸¡à¸žà¸¹ à¸•à¸±à¹‰à¸‡à¸ªà¸±à¸•à¸¢à¸²à¸˜à¸´à¸©à¸à¸²à¸™",
        "à¸žà¸.à¸™à¸±à¸™à¸—à¹Œà¸™à¸ à¸±à¸ª à¸Šà¸µà¸§à¸°à¹€à¸à¸£à¸µà¸¢à¸‡à¹„à¸à¸£",
    ],

    # Maxillofacial | à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£
    "MAXILO_ANY": [
        "à¸—à¸ž.à¸‰à¸¥à¸­à¸‡à¸£à¸±à¸ à¹€à¸”à¸Šà¸²",
        "à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸£à¸±à¸‡",
    ],
}

DOCTOR_ALIASES: Dict[str, str] = {
    # à¸¢à¸à¸•à¸±à¸§à¸­à¸¢à¹ˆà¸²à¸‡à¸ªà¸°à¸à¸”/à¸§à¸£à¸£à¸„à¸•à¹ˆà¸²à¸‡à¸à¸±à¸™/à¸žà¸´à¸¡à¸žà¹Œà¸œà¸´à¸”à¸—à¸µà¹ˆà¹€à¸ˆà¸­à¸šà¹ˆà¸­à¸¢
    "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸° à¸„à¸¸à¸“à¸²à¸Šà¸™": "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™",
    "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸²": "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™",
    "à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ": "à¸™à¸ž.à¸˜à¸™à¸§à¸±à¸’à¸™à¹Œ à¸žà¸±à¸™à¸˜à¸¸à¹Œà¸žà¸£à¸«à¸¡",
    "à¸žà¸.à¸£à¸±à¸à¸žà¸£": "à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£",
    "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢": "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ",
    "à¸žà¸.à¸žà¸´à¸£à¸´à¸¢à¸²": "à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸² à¹à¸ªà¸™à¸§à¸±à¸™à¸”à¸µ",
    "à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸²": "à¸žà¸.à¸žà¸´à¸£à¸¸à¸“à¸¢à¸² à¹à¸ªà¸™à¸§à¸±à¸™à¸”à¸µ",
    "à¸žà¸.à¸ªà¸²à¸¢à¸à¸™": "à¸žà¸.à¸ªà¸²à¸¢à¸à¸™ à¸šà¸£à¸£à¸“à¸ˆà¸´à¸•à¸£à¹Œ",
    "à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥": "à¸™à¸ž.à¸Šà¸±à¸Šà¸žà¸¥ à¸­à¸‡à¸„à¹Œà¹‚à¸†à¸©à¸´à¸•",
    "à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ": "à¸™à¸ž.à¸“à¸±à¸à¸žà¸‡à¸¨à¹Œ à¸¨à¸£à¸µà¹‚à¸žà¸™à¸—à¸­à¸‡",
    "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸": "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ",
    "à¸™à¸ž.à¸à¸¤à¸©à¸Žà¸²": "à¸™à¸ž.à¸à¸¤à¸©à¸Žà¸² à¸­à¸´à¹‰à¸‡à¸­à¸³à¸žà¸£",
    "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ": "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ",
    "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£": "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£ à¸«à¸¡à¸§à¸”à¹„à¸˜à¸ªà¸‡",
    "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—": "à¸žà¸.à¸ªà¸¸à¸ à¸²à¸ à¸£à¸“à¹Œ à¸žà¸´à¸“à¸žà¸²à¸—à¸¢à¹Œ",
    "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ": "à¸žà¸.à¸žà¸´à¸Šà¸±à¸¢ à¸ªà¸¸à¸§à¸±à¸’à¸™à¸žà¸¹à¸™à¸¥à¸²à¸ ",
    "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™": "à¸™à¸ž.à¸§à¸´à¸©à¸“à¸¸ à¸œà¸¹à¸à¸žà¸±à¸™à¸˜à¹Œ",
    "à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸¥à¸±à¸‡": "à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸£à¸±à¸‡",
    "à¸žà¸.à¸ªà¸µà¸à¸Šà¸¡à¸žà¸¹ à¸•à¸±à¹‰à¸‡à¸ªà¸±à¸•à¸¢à¸²à¸˜": "à¸žà¸.à¸ªà¸µà¸à¸Šà¸¡à¸žà¸¹ à¸•à¸±à¹‰à¸‡à¸ªà¸±à¸•à¸¢à¸²à¸˜à¸´à¸©à¸à¸²à¸™",
}

DOCTOR_ALIASES.update(
    {
        "à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸²": "à¸—à¸žà¸.à¸­à¸£à¸¸à¸“à¸™à¸ à¸² à¸„à¸´à¸ªà¸²à¸£à¸±à¸‡",
        "à¸—à¸ž.à¸‰à¸¥à¸­à¸‡à¸£à¸±à¸": "à¸—à¸ž.à¸‰à¸¥à¸­à¸‡à¸£à¸±à¸ à¹€à¸”à¸Šà¸²",
        "à¸™à¸ž.à¸§à¸£à¸§à¸´à¸Š": "à¸™à¸ž.à¸§à¸£à¸§à¸´à¸Š à¸žà¸¥à¹€à¸§à¸µà¸¢à¸‡à¸˜à¸£à¸£à¸¡",
        "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£": "à¸žà¸.à¸ªà¸¸à¸—à¸˜à¸´à¸žà¸£ à¸«à¸¡à¸§à¸”à¹„à¸˜à¸ªà¸‡",
    }
)

TOKEN_DISPLAY_NAMES: Dict[str, str] = {
    "SUR_ANY": "à¸—à¸µà¸¡à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸—à¸±à¹ˆà¸§à¹„à¸›",
    "ORTHO_ANY": "à¸—à¸µà¸¡à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸à¸£à¸°à¸”à¸¹à¸",
    "URO_ANY": "à¸—à¸µà¸¡à¸£à¸°à¸šà¸šà¸—à¸²à¸‡à¹€à¸”à¸´à¸™à¸›à¸±à¸ªà¸ªà¸²à¸§à¸°",
    "ENT_ANY": "à¸—à¸µà¸¡à¹‚à¸ªà¸• à¸¨à¸­ à¸™à¸²à¸ªà¸´à¸",
    "OBGYN_ANY": "à¸—à¸µà¸¡à¸ªà¸¹à¸•à¸´-à¸™à¸£à¸µà¹€à¸§à¸Š",
    "EYE_ANY": "à¸—à¸µà¸¡à¸ˆà¸±à¸à¸©à¸¸",
    "MAXILO_ANY": "à¸—à¸µà¸¡à¸¨à¸±à¸¥à¸¢à¸à¸£à¸£à¸¡à¸‚à¸²à¸à¸£à¸£à¹„à¸à¸£",
    "CLOSE": "à¸›à¸´à¸”à¸«à¹‰à¸­à¸‡",
    "CLOSED": "à¸›à¸´à¸”à¸«à¹‰à¸­à¸‡",
}

CLOSED_TOKENS = {"CLOSE", "CLOSED"}


# ================================================================


def normalize_doctor_name(name: str) -> str:
    s = " ".join(str(name or "").split())
    return DOCTOR_ALIASES.get(s, s)


GROUP_MEMBER_LOOKUP: Dict[str, Set[str]] = {
    token: {normalize_doctor_name(member) for member in members}
    for token, members in GROUPS.items()
}


def week_of_month(d: date) -> int:
    first = d.replace(day=1)
    return ((d.day + first.weekday() - 1) // 7) + 1


def time_to_period(hhmm_or_tf: str) -> str:
    if hhmm_or_tf == "TF":
        return "ANY"
    try:
        hour = int(hhmm_or_tf.split(":", 1)[0])
    except Exception:
        return "ANY"
    return "AM" if hour < 12 else "PM"


def doctor_in_group(doc: str, token: str) -> bool:
    normalized = normalize_doctor_name(doc)
    return normalized in GROUP_MEMBER_LOOKUP.get(token, set())


def match_doctor(token_or_name: str, doctor_name: str) -> bool:
    token = token_or_name
    if not token:
        return False
    if token in CLOSED_TOKENS:
        return False
    if token in GROUPS or token in {"SUR_ANY", "ORTHO_ANY", "ENT_ANY", "EYE_ANY", "MAXILO_ANY", "OBGYN_ANY"}:
        return doctor_in_group(doctor_name, token)
    return normalize_doctor_name(token) == normalize_doctor_name(doctor_name)


def doctor_service_token(doctor_name: str) -> str:
    normalized = normalize_doctor_name(doctor_name)
    for token, members in GROUP_MEMBER_LOOKUP.items():
        if normalized in members:
            return token
    return ""


def _rule_tokens(rule: Dict[str, object]) -> List[str]:
    doctor_token = rule.get("doctor")
    if isinstance(doctor_token, list):
        return [str(tok or "") for tok in doctor_token]
    return [str(doctor_token or "")]


def _rule_matches_service(rule: Dict[str, object], service_token: str) -> bool:
    if not service_token:
        return False
    members = GROUP_MEMBER_LOOKUP.get(service_token, set())
    if not members and service_token in GROUPS:
        members = {normalize_doctor_name(name) for name in GROUPS.get(service_token, [])}

    for token in _rule_tokens(rule):
        if not token or token in CLOSED_TOKENS:
            continue
        if token == service_token:
            return True
        if token in GROUP_MEMBER_LOOKUP and token == service_token:
            return True
        if normalize_doctor_name(token) in members:
            return True
    return False


def _describe_doctor_token(token: str) -> str:
    if not token:
        return ""
    label = TOKEN_DISPLAY_NAMES.get(token)
    if label:
        return label
    if token in GROUPS:
        # à¹„à¸¡à¹ˆà¸¡à¸µà¸‚à¹‰à¸­à¸„à¸§à¸²à¸¡à¹€à¸‰à¸žà¸²à¸° à¹ƒà¸«à¹‰à¹ƒà¸Šà¹‰à¸Šà¸·à¹ˆà¸­à¸à¸¥à¸¸à¹ˆà¸¡à¹à¸£à¸à¹€à¸›à¹‡à¸™à¸•à¸±à¸§à¹à¸—à¸™
        members = GROUPS.get(token, [])
        if members:
            return f"{TOKEN_DISPLAY_NAMES.get(token, '') or members[0]}"
    return normalize_doctor_name(token)


def describe_or_plan_label(case_date: date, or_room: str) -> str:
    if not or_room or or_room == "-":
        return ""

    weekday = case_date.weekday()
    plan = WEEKLY_DOCTOR_OR_PLAN.get(weekday, {})
    rules = plan.get(or_room, []) or []
    if not rules:
        return ""

    current_week = week_of_month(case_date)

    def label_for_rule(rule: Dict[str, object]) -> str:
        doctor_token = rule.get("doctor")
        if isinstance(doctor_token, list):
            tokens = [str(tok) for tok in doctor_token]
        else:
            tokens = [str(doctor_token or "")]
        parts: List[str] = []
        for tok in tokens:
            desc = _describe_doctor_token(tok)
            if desc:
                parts.append(desc)
        label = ", ".join(parts)
        when = (str(rule.get("when") or "ALLDAY").upper())
        if when == "AM":
            return f"à¹€à¸Šà¹‰à¸²: {label}" if label else "à¹€à¸Šà¹‰à¸²"
        if when == "PM":
            return f"à¸šà¹ˆà¸²à¸¢: {label}" if label else "à¸šà¹ˆà¸²à¸¢"
        if label:
            return label
        return ""

    filtered = [rule for rule in rules if current_week in rule.get("weeks", [1, 2, 3, 4, 5])]
    if not filtered:
        filtered = rules

    labels = [label_for_rule(rule) for rule in filtered]
    labels = [lbl for lbl in labels if lbl]
    return " â€¢ ".join(labels)


_WED = 2
_OWNER_WED: Dict[str, str] = {
    "OR1": "à¸™à¸ž.à¸ªà¸¸à¸£à¸´à¸¢à¸² à¸„à¸¸à¸“à¸²à¸Šà¸™",
    "OR6": "à¸žà¸.à¸£à¸±à¸à¸žà¸£ à¸•à¸±à¹‰à¸‡à¹€à¸žà¸µà¸¢à¸£",
}


def resolve_or_owner(or_room: str, dt: date, fallback: str | None = None) -> str:
    """Return the resolved owner for an OR room, overriding Wednesday defaults."""
    room_key = (or_room or "").strip()
    if dt and room_key and dt.weekday() == _WED and room_key in _OWNER_WED:
        return _OWNER_WED[room_key]
    fallback_name = normalize_doctor_name(fallback) if fallback else ""
    return fallback_name or "-"


def pick_or_by_doctor(case_date: date, time_str: str, doctor_name: str) -> str:
    if not doctor_name:
        return ""

    weekday = case_date.weekday()
    plan = WEEKLY_DOCTOR_OR_PLAN.get(weekday, {})
    if not plan:
        return ""

    current_week = week_of_month(case_date)
    period = time_to_period(time_str)

    def iter_rules():
        for or_room, rules in plan.items():
            for rule in rules or []:
                yield or_room, rule

    for or_room, rule in iter_rules():
        doctor_token = rule.get("doctor")
        if not doctor_token:
            continue
        weeks = rule.get("weeks", [1, 2, 3, 4, 5])
        when = (rule.get("when") or "ALLDAY").upper()
        doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
        if current_week in weeks and (when == "ALLDAY" or period == "ANY" or period == when):
            if any(match_doctor(doc, doctor_name) for doc in doctors):
                return or_room

    if period == "ANY":
        for or_room, rule in iter_rules():
            doctor_token = rule.get("doctor")
            if not doctor_token:
                continue
            weeks = rule.get("weeks", [1, 2, 3, 4, 5])
            if current_week not in weeks:
                continue
            doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
            if any(tok in CLOSED_TOKENS for tok in doctors):
                continue
            if any(match_doctor(doc, doctor_name) for doc in doctors):
                return or_room

    for or_room, rule in iter_rules():
        doctor_token = rule.get("doctor")
        if not doctor_token:
            continue
        doctors = doctor_token if isinstance(doctor_token, list) else [doctor_token]
        if any(tok in CLOSED_TOKENS for tok in doctors):
            continue
        # à¸•à¸£à¸‡à¸Šà¸·à¹ˆà¸­à¹à¸žà¸—à¸¢à¹Œà¹à¸šà¸šà¹€à¸‰à¸žà¸²à¸°à¹€à¸ˆà¸²à¸°à¸ˆà¸‡à¹€à¸—à¹ˆà¸²à¸™à¸±à¹‰à¸™ (à¹„à¸¡à¹ˆà¸ˆà¸±à¸šà¸„à¸¹à¹ˆ token à¸£à¸°à¸”à¸±à¸šà¹à¸œà¸™à¸)
        explicit_tokens = [tok for tok in doctors if tok not in GROUPS]
        if not explicit_tokens:
            continue
        if any(match_doctor(doc, doctor_name) for doc in explicit_tokens):
            return or_room

    service_token = doctor_service_token(doctor_name)
    if service_token:
        for or_room, rule in iter_rules():
            weeks = rule.get("weeks", [1, 2, 3, 4, 5])
            when = (rule.get("when") or "ALLDAY").upper()
            if current_week in weeks and (when == "ALLDAY" or period == "ANY" or period == when):
                if _rule_matches_service(rule, service_token):
                    return or_room

        for or_room, rule in iter_rules():
            if _rule_matches_service(rule, service_token):
                return or_room

    return ""


class Main(QtWidgets.QWidget):
    def __init__(self, host, port, token):
        super().__init__()
        self.cli = ClientHTTP(host, port, token)
        self.sched = SharedScheduleModel()
        self.db_logger = LocalDBLogger()
        self.ws = None;
        self.rows_cache = []
        self.seq_seen = self.sched.seq()
        icon = _load_app_icon()
        self.setWindowIcon(icon)
        self.tray = QtWidgets.QSystemTrayIcon(icon, self);
        self.tray.show()

        self._last_status_by_hn: dict[str, str] = {}
        self._runner_status_cache: Dict[str, dict] = {}
        self._last_runner_user: str = ""
        self._runner_finished_sent: Set[str] = set()

        # form edit mode
        self._edit_idx: Optional[int] = None
        self._last_focus_uid: Optional[str] = None  # à¹ƒà¸Šà¹‰à¹„à¸®à¹„à¸¥à¸•à¹Œà¸«à¸¥à¸±à¸‡à¸šà¸±à¸™à¸—à¸¶à¸

        self.toast = Toast(self)
        self._current_specialty_key = ""
        self._last_snapshot: Optional[List[Dict[str, object]]] = None

        self._diag_base_catalog: List[str] = []
        self._diag_catalog_full: List[str] = []
        self._op_catalog_full: List[str] = []
        self._dx_index: Optional[FastSearchIndex] = None
        self._op_index: Optional[FastSearchIndex] = None
        self._search_executor = ThreadPoolExecutor(max_workers=1)
        self._search_timer = QtCore.QTimer(self)
        self._search_timer.setSingleShot(True)
        try:
            self._dx_search_limit = max(1, int(os.getenv("DX_SEARCH_LIMIT", "100")))
        except ValueError:
            self._dx_search_limit = 100
        try:
            debounce_ms = max(0, int(os.getenv("DX_SEARCH_DEBOUNCE_MS", "150")))
        except ValueError:
            debounce_ms = 150
        self._search_timer.setInterval(debounce_ms)
        self._search_timer.timeout.connect(self._on_diag_search_timeout)
        self._latest_diag_query = ""
        try:
            self._op_search_limit = max(1, int(os.getenv("OP_SEARCH_LIMIT", "100")))
        except ValueError:
            self._op_search_limit = 100
        try:
            op_debounce = max(0, int(os.getenv("OP_SEARCH_DEBOUNCE_MS", "150")))
        except ValueError:
            op_debounce = 150
        self._op_search_timer = QtCore.QTimer(self)
        self._op_search_timer.setSingleShot(True)
        self._op_search_timer.setInterval(op_debounce)
        self._op_search_timer.timeout.connect(self._on_op_search_timeout)
        self._latest_op_query = ""

        self.setWindowTitle("Registry Patient Connect â€” ORNBH")
        self.resize(1360, 900)
        apply_modern_theme(self)
        self._build_ui();
        self._load_settings();
        self._pdpa_gate();
        self._start_timers()

    # ---------- UI ----------
    def _build_ui(self):
        outer = QtWidgets.QVBoxLayout(self);
        outer.setSpacing(12);
        outer.setContentsMargins(14, 14, 14, 14)
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setStyleSheet(
            "QTabWidget::pane{border:0;} QTabBar::tab{padding:10px 16px;border-radius:12px;margin:4px;background:#e9eef8;} QTabBar::tab:selected{background:#2563eb;color:#fff;}")
        outer.addWidget(self.tabs)

        # TAB 1 â€” à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™ (à¸«à¹ˆà¸­à¸”à¹‰à¸§à¸¢ ScrollArea à¹€à¸žà¸·à¹ˆà¸­à¸›à¹‰à¸­à¸‡à¸à¸±à¸™à¸„à¸­à¸™à¹‚à¸—à¸£à¸¥à¸«à¸”)
        tab1_inner = QtWidgets.QWidget()
        t1 = QtWidgets.QVBoxLayout(tab1_inner);
        t1.setSpacing(12);
        t1.setContentsMargins(0, 0, 0, 0)
        t1_banner = InfoBanner(
            title="à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢ (Schedule â€” Private)",
            subtitle="à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹€à¸à¹‡à¸šà¹ƒà¸™à¹€à¸„à¸£à¸·à¹ˆà¸­à¸‡ à¹à¸¥à¸°à¹à¸Šà¸£à¹Œà¹ƒà¸«à¹‰à¹‚à¸›à¸£à¹à¸à¸£à¸¡à¸«à¸¥à¸±à¸à¹à¸šà¸šà¹€à¸£à¸µà¸¢à¸¥à¹„à¸—à¸¡à¹Œ",
            variant="blue",
            icon="ðŸ“",
        )
        t1.addWidget(t1_banner)
        form = Card("à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢ (Schedule â€” Private)", "à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹€à¸à¹‡à¸šà¹ƒà¸™à¹€à¸„à¸£à¸·à¹ˆà¸­à¸‡ à¹à¸¥à¸°à¹à¸Šà¸£à¹Œà¹ƒà¸«à¹‰à¹‚à¸›à¸£à¹à¸à¸£à¸¡à¸«à¸¥à¸±à¸à¹à¸šà¸šà¹€à¸£à¸µà¸¢à¸¥à¹„à¸—à¸¡à¹Œ")
        form.title_lbl.hide()
        g = form.grid;
        r = 0
        g.setColumnStretch(0, 0);
        g.setColumnStretch(1, 2);
        g.setColumnStretch(2, 0);
        g.setColumnStretch(3, 1)
        g.setColumnStretch(4, 0);
        g.setColumnStretch(5, 2)
        g.addWidget(QtWidgets.QLabel("OR"), r, 0)
        self.cb_or = QtWidgets.QComboBox();
        self._refresh_or_cb(self.cb_or);
        self.cb_or.setMinimumWidth(140)
        g.addWidget(self.cb_or, r, 1)
        self.btn_manage_or = QtWidgets.QPushButton("à¸ˆà¸±à¸”à¸à¸²à¸£ OR");
        self.btn_manage_or.setProperty("variant", "ghost")
        g.addWidget(self.btn_manage_or, r, 2, 1, 2)
        r += 1
        g.addWidget(QtWidgets.QLabel("à¸Šà¸·à¹ˆà¸­-à¸ªà¸à¸¸à¸¥"), r, 0)
        self.ent_name = QtWidgets.QLineEdit()
        g.addWidget(self.ent_name, r, 1, 1, 3)
        g.addWidget(QtWidgets.QLabel("à¸­à¸²à¸¢à¸¸"), r, 4)
        self.ent_age = QtWidgets.QLineEdit();
        self.ent_age.setValidator(QtGui.QIntValidator(0, 150, self))
        g.addWidget(self.ent_age, r, 5)
        r += 1
        g.addWidget(QtWidgets.QLabel("HN"), r, 0)
        self.ent_hn = QtWidgets.QLineEdit();
        self.ent_hn.setMaxLength(9);
        self.ent_hn.setValidator(QtGui.QIntValidator(0, 999999999, self))
        g.addWidget(self.ent_hn, r, 1)
        g.addWidget(QtWidgets.QLabel("Ward"), r, 2)
        self.cb_ward = QtWidgets.QComboBox();
        self.cb_ward.setEditable(True)
        self.cb_ward.blockSignals(True)
        self.cb_ward.clear()

        def _safe_ward_list() -> list[str]:
            try:
                src = WARD_LIST() if callable(WARD_LIST) else WARD_LIST
                items = list(src)
            except Exception:
                items = []
            results: list[str] = []
            for entry in items:
                if isinstance(entry, (str, bytes)):
                    text = str(entry).strip()
                    if text:
                        results.append(text)
            return results

        wards = _safe_ward_list()
        self.cb_ward.addItems(wards)
        if self.cb_ward.count() > 0:
            self.cb_ward.setCurrentIndex(0)
        self.cb_ward.blockSignals(False)
        self.cb_ward.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        ward_options = [w for w in wards if w and w != WARD_PLACEHOLDER]
        comp = QtWidgets.QCompleter(ward_options)
        comp.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        comp.setFilterMode(QtCore.Qt.MatchContains)
        self.cb_ward.setCompleter(comp)
        self.cb_ward.setCurrentIndex(0)
        self.cb_ward.setEditText(WARD_PLACEHOLDER)
        g.addWidget(self.cb_ward, r, 3)
        g.addWidget(QtWidgets.QLabel("à¸‚à¸™à¸²à¸”à¹€à¸„à¸ª"), r, 4)
        self.cb_case = NoWheelComboBox();
        self.cb_case.addItems(["", "Minor", "Major"])
        self.cb_case.setMinimumWidth(120)
        g.addWidget(self.cb_case, r, 5)
        r += 1
        g.addWidget(QtWidgets.QLabel("à¸„à¸§à¸²à¸¡à¹€à¸£à¹ˆà¸‡à¸”à¹ˆà¸§à¸™"), r, 0)
        self.cb_urgency = NoWheelComboBox();
        self.cb_urgency.addItems(["Elective", "Emergency"])
        self.cb_urgency.setMinimumWidth(180)
        g.addWidget(self.cb_urgency, r, 1)
        self.lbl_period_info = QtWidgets.QLabel("")
        self.lbl_period_info.setProperty("hint", "1")
        g.addWidget(self.lbl_period_info, r, 2, 1, 4)
        r += 1
        g.addWidget(QtWidgets.QLabel("à¸§à¸±à¸™à¸—à¸µà¹ˆ"), r, 0)
        self.date = QtWidgets.QDateEdit(QtCore.QDate.currentDate());
        self.date.setCalendarPopup(True);
        self.date.setDisplayFormat("dd/MM/yyyy");
        self.date.setLocale(QLocale("en_US"))
        g.addWidget(self.date, r, 1)
        g.addWidget(QtWidgets.QLabel("à¹€à¸§à¸¥à¸²"), r, 2)
        self.time = QtWidgets.QTimeEdit(QtCore.QTime.currentTime());
        self.time.setDisplayFormat("HH:mm");
        self.time.setLocale(QLocale("en_US"))
        g.addWidget(self.time, r, 3)
        g.addWidget(QtWidgets.QLabel("à¹à¸œà¸™à¸"), r, 4)
        self.cb_dept = QtWidgets.QComboBox();
        self.cb_dept.addItems(["â€” à¹€à¸¥à¸·à¸­à¸à¹à¸œà¸™à¸ â€”"] + list(DEPT_DOCTORS.keys()))
        g.addWidget(self.cb_dept, r, 5)
        r += 1
        self.lbl_warn = QtWidgets.QLabel("");
        self.lbl_warn.setProperty("warn", "1")
        g.addWidget(self.lbl_warn, r, 0, 1, 6)
        r += 1
        self.row_doctor_label = QtWidgets.QLabel("à¹à¸žà¸—à¸¢à¹Œà¸œà¸¹à¹‰à¸œà¹ˆà¸²à¸•à¸±à¸”");
        g.addWidget(self.row_doctor_label, r, 0)
        self.cb_doctor = QtWidgets.QComboBox();
        g.addWidget(self.cb_doctor, r, 1, 1, 5)
        r += 1

        g.addWidget(section_header("Diagnosis"), r, 0, 1, 6)
        r += 1
        self.diag_adder = SearchSelectAdder("à¸„à¹‰à¸™à¸«à¸² ICD-10 / ICD-10-TM...", suggestions=[])
        self.diag_adder.requestPersist.connect(self._on_diagnosis_persist_requested)
        if self.diag_adder.search_line:
            self.diag_adder.search_line.textChanged.connect(self._on_diag_query_changed)
        g.addWidget(self.diag_adder, r, 0, 1, 6)
        r += 1

        g.addWidget(section_header("Operation"), r, 0, 1, 6)
        r += 1
        self.op_adder = SearchSelectAdder("à¸„à¹‰à¸™à¸«à¸²/à¹€à¸¥à¸·à¸­à¸ Operation...", suggestions=[])
        self.op_adder.itemsChanged.connect(self._on_operations_changed)
        self.op_adder.requestPersist.connect(self._on_operation_persist_requested)
        if self.op_adder.search_line:
            self.op_adder.search_line.textChanged.connect(self._on_op_query_changed)
        g.addWidget(self.op_adder, r, 0, 1, 6)
        r += 1

        g.addWidget(section_header("Scrub Nurse / à¸—à¸µà¸¡à¸žà¸¢à¸²à¸šà¸²à¸¥"), r, 0, 1, 6)
        r += 1
        row_n = QtWidgets.QHBoxLayout();
        row_n.setSpacing(8)

        def _hint(txt: str) -> QtWidgets.QLabel:
            lab = QtWidgets.QLabel(txt)
            lab.setProperty("hint", "1")
            return lab

        self.cb_assist1 = make_search_combo(SCRUB_NURSES)
        self.cb_assist2 = make_search_combo(SCRUB_NURSES)
        self.cb_scrub = make_search_combo(SCRUB_NURSES)
        self.cb_circulate = make_search_combo(SCRUB_NURSES)

        row_n.addWidget(_hint("Assist 1"));
        row_n.addWidget(self.cb_assist1, 1)
        row_n.addWidget(_hint("Assist 2"));
        row_n.addWidget(self.cb_assist2, 1)
        row_n.addWidget(_hint("Scrub"));
        row_n.addWidget(self.cb_scrub, 1)
        row_n.addWidget(_hint("Circulate"));
        row_n.addWidget(self.cb_circulate, 1)
        g.addLayout(row_n, r, 0, 1, 6)
        r += 1

        g.addWidget(section_header("à¹€à¸§à¸¥à¸²à¹€à¸£à¸´à¹ˆà¸¡â€“à¸ˆà¸šà¸œà¹ˆà¸²à¸•à¸±à¸” (à¹ƒà¸ªà¹ˆà¸«à¸£à¸·à¸­à¹„à¸¡à¹ˆà¹ƒà¸ªà¹ˆà¸à¹‡à¹„à¸”à¹‰)"), r, 0, 1, 6)
        r += 1
        row_t = QtWidgets.QHBoxLayout();
        row_t.setSpacing(10)
        self.ck_time_start = QtWidgets.QCheckBox("à¸£à¸°à¸šà¸¸à¹€à¸§à¸¥à¸²à¹€à¸£à¸´à¹ˆà¸¡")
        self.time_start = QtWidgets.QTimeEdit(QtCore.QTime.currentTime())
        self.time_start.setDisplayFormat("HH:mm")
        self.time_start.setEnabled(False)
        self.ck_time_end = QtWidgets.QCheckBox("à¸£à¸°à¸šà¸¸à¹€à¸§à¸¥à¸²à¸ˆà¸š")
        self.time_end = QtWidgets.QTimeEdit(QtCore.QTime.currentTime())
        self.time_end.setDisplayFormat("HH:mm")
        self.time_end.setEnabled(False)

        self.ck_time_start.toggled.connect(lambda ch: self.time_start.setEnabled(ch))
        self.ck_time_end.toggled.connect(lambda ch: self.time_end.setEnabled(ch))

        row_t.addWidget(self.ck_time_start)
        row_t.addWidget(self.time_start)
        row_t.addSpacing(16)
        row_t.addWidget(self.ck_time_end)
        row_t.addWidget(self.time_end)
        row_t.addStretch(1)
        g.addLayout(row_t, r, 0, 1, 6)
        r += 1

        self.btn_add = QtWidgets.QPushButton("âž• à¹€à¸žà¸´à¹ˆà¸¡");
        self.btn_add.setProperty("variant", "primary")
        self.btn_cancel_edit = QtWidgets.QPushButton("à¸¢à¸à¹€à¸¥à¸´à¸à¹à¸à¹‰à¹„à¸‚");
        self.btn_cancel_edit.setProperty("variant", "ghost");
        self.btn_cancel_edit.hide()
        self.btn_clear = QtWidgets.QPushButton("ðŸ§¹ à¹€à¸„à¸¥à¸µà¸¢à¸£à¹Œ");
        self.btn_clear.setProperty("variant", "ghost")
        rowb = QtWidgets.QHBoxLayout();
        rowb.setSpacing(10);
        rowb.addWidget(self.btn_add);
        rowb.addWidget(self.btn_cancel_edit);
        rowb.addWidget(self.btn_clear);
        rowb.addStretch(1)
        g.addLayout(rowb, r, 0, 1, 6)
        r += 1

        t1.addWidget(form);
        t1.addStretch(1)
        tab1_scroll = QtWidgets.QScrollArea()
        tab1_scroll.setWidgetResizable(True)
        tab1_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab1_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        tab1_scroll.setWidget(tab1_inner)
        self.tabs.addTab(tab1_scroll, "à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢")

        # TAB 2 â€” Result Schedule
        tab2 = QtWidgets.QWidget();
        t2 = QtWidgets.QVBoxLayout(tab2);
        t2.setSpacing(12)
        self.result_banner = InfoBanner("", "à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¹‚à¸£à¸‡à¸žà¸¢à¸²à¸šà¸²à¸¥à¸«à¸™à¸­à¸‡à¸šà¸±à¸§à¸¥à¸³à¸ à¸¹")
        t2.addWidget(self.result_banner)
        self.card_result = Card("à¸•à¸²à¸£à¸²à¸‡à¸à¸²à¸£à¸œà¹ˆà¸²à¸•à¸±à¸” à¸›à¸£à¸°à¸ˆà¸³à¸§à¸±à¸™", "à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¹‚à¸£à¸‡à¸žà¸¢à¸²à¸šà¸²à¸¥à¸«à¸™à¸­à¸‡à¸šà¸±à¸§à¸¥à¸³à¸ à¸¹")
        self.card_result.title_lbl.hide()
        gr2 = self.card_result.grid
        self.tree2 = QtWidgets.QTreeWidget()
        self.tree2.setColumnCount(18)
        self.tree2.setHeaderLabels([
            "OR/à¹€à¸§à¸¥à¸²", "HN", "à¸Šà¸·à¹ˆà¸­-à¸ªà¸à¸¸à¸¥", "à¸­à¸²à¸¢à¸¸", "Diagnosis", "Operation",
            "à¹à¸žà¸—à¸¢à¹Œ", "Ward", "à¸‚à¸™à¸²à¸”à¹€à¸„à¸ª", "à¹à¸œà¸™à¸", "à¹€à¸£à¸´à¹ˆà¸¡", "à¸ˆà¸š", "à¸Šà¹ˆà¸§à¸‡à¹€à¸§à¸¥à¸²",
            "Assist 1", "Assist 2", "Scrub", "Cir", "à¸ªà¸–à¸²à¸™à¸°"
        ])
        self.tree2.setUniformRowHeights(False)
        self.tree2.setAlternatingRowColors(True)
        self.tree2.setRootIsDecorated(True)
        self.tree2.setIndentation(12)
        self.tree2.setWordWrap(False)
        self.tree2.setTextElideMode(QtCore.Qt.ElideNone)
        self.tree2.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)

        self.tree2.setStyleSheet("""
            QTreeWidget{
                background:#ffffff;
                border:1px solid #dfe6f0;
                border-radius:12px;
                gridline-color:#e8edf5;
            }
            QHeaderView{ background:transparent; border:none; margin:0; padding:0; }
            QHeaderView::section{
                background:#f6f9ff; color:#0f172a; font-weight:900; letter-spacing:.2px;
                padding:12px 14px;
                border-top:1px solid #dfe6f0; border-bottom:1px solid #dfe6f0;
                border-right:1px solid #dfe6f0;
            }
            QHeaderView::section:first{ border-top-left-radius:12px; border-left:1px solid #dfe6f0; }
            QHeaderView::section:last{ border-top-right-radius:12px; border-right:1px solid #dfe6f0; }
            QTreeWidget::item{ height:36px; }
            QTreeWidget::item:alternate{ background:#fbfdff; }
            QTreeWidget::item:selected{ background:rgba(37,99,235,0.12); border-radius:8px; }
            QTreeWidget::item:hover{ background:rgba(2,132,199,0.06); }
        """)
        hdr = self.tree2.header()
        hdr.setStretchLastSection(False)
        hdr.setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        hdr.setFixedHeight(42)
        for i in range(18):
            hdr.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
        for i in (2, 4, 5):
            hdr.setSectionResizeMode(i, QtWidgets.QHeaderView.Interactive)
        self.tree2.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tree2.customContextMenuRequested.connect(self._result_ctx_menu)
        gr2.addWidget(self.tree2, 0, 0, 1, 1)

        import_bar = QtWidgets.QHBoxLayout()
        import_bar.setContentsMargins(0, 6, 0, 0)
        import_bar.setSpacing(10)
        self.btn_send_runner = QtWidgets.QPushButton("ðŸšš à¸ªà¹ˆà¸‡à¹ƒà¸«à¹‰ Runner (à¸§à¸±à¸™à¸™à¸µà¹‰)")
        self.btn_send_runner.setProperty("variant", "primary")
        import_bar.addWidget(self.btn_send_runner, 0)
        self.btn_import_excel = QtWidgets.QPushButton("ðŸ“¥ à¸™à¸³à¹€à¸‚à¹‰à¸²à¸ˆà¸²à¸ Excel")
        self.btn_import_excel.setProperty("variant", "ghost")
        import_bar.addWidget(self.btn_import_excel, 0)
        self.btn_clear_board = QtWidgets.QPushButton("ðŸ§¹ à¸¥à¹‰à¸²à¸‡à¸à¸£à¸°à¸”à¸²à¸™")
        self.btn_clear_board.setProperty("variant", "destructive")
        import_bar.addWidget(self.btn_clear_board, 0)
        self.btn_undo_clear = QtWidgets.QPushButton("â†©ï¸ à¸¢à¹‰à¸­à¸™à¸à¸¥à¸±à¸šà¸à¸²à¸£à¸¥à¹‰à¸²à¸‡")
        self.btn_undo_clear.setProperty("variant", "ghost")
        self.btn_undo_clear.setEnabled(False)
        import_bar.addWidget(self.btn_undo_clear, 0)
        import_bar.addStretch(1)
        gr2.addLayout(import_bar, 1, 0, 1, 1)
        gr2.setRowStretch(0, 1)
        gr2.setRowStretch(1, 0)
        t2.addWidget(self.card_result, 1)
        self.tabs.addTab(tab2, "Result Schedule")

        self._clear_shortcut = QtGui.QShortcut(QtGui.QKeySequence("Ctrl+Del"), self)
        self._clear_shortcut.activated.connect(self._on_clear_board_clicked)

        # TAB 3 â€” Monitor
        tab3 = QtWidgets.QWidget();
        t3 = QtWidgets.QVBoxLayout(tab3);
        t3.setSpacing(12);
        t3.setContentsMargins(0, 0, 0, 0)
        t3_banner = InfoBanner(
            title="Result (Monitor) â€” à¸ˆà¸²à¸à¹€à¸‹à¸´à¸£à¹Œà¸Ÿà¹€à¸§à¸­à¸£à¹Œ",
            subtitle="",
            variant="violet",
            icon="ðŸ—“ï¸",
        )
        t3.addWidget(t3_banner)
        server_bar = QtWidgets.QFrame();
        server_bar.setStyleSheet("QFrame{background:#fff;border:1px solid #e6eaf2;border-radius:14px;padding:8px;}");
        add_shadow(server_bar)
        hb = QtWidgets.QHBoxLayout(server_bar);
        hb.setContentsMargins(8, 8, 8, 8)
        self.ent_host = QtWidgets.QLineEdit("127.0.0.1");
        self.ent_host.setMaximumWidth(180);
        self.ent_host.setEchoMode(QtWidgets.QLineEdit.Password)
        self.ent_port = QtWidgets.QLineEdit(str(DEFAULT_PORT));
        self.ent_port.setMaximumWidth(90)
        self.ent_token = QtWidgets.QLineEdit(DEFAULT_TOKEN);
        self.ent_token.setEchoMode(QtWidgets.QLineEdit.Password)
        self.btn_health = QtWidgets.QPushButton("Health");
        self.btn_health.setProperty("variant", "ghost");
        self.btn_health.clicked.connect(self._on_health)
        for w, lbl in [(self.ent_host, "Host"), (self.ent_port, "Port"), (self.ent_token, "Token")]:
            box = QtWidgets.QHBoxLayout();
            lab = QtWidgets.QLabel(lbl);
            lab.setProperty("hint", "1");
            box.addWidget(lab);
            box.addWidget(w);
            hb.addLayout(box)
        hb.addWidget(self.btn_health);
        hb.addStretch(1)
        self.status_chip = QtWidgets.QLabel("â— Offline")
        self.status_chip.setStyleSheet(
            "color:#ef4444;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")
        hb.addWidget(self.status_chip)
        t3.addWidget(server_bar)

        mon = Card("Monitor Realtime (à¸ˆà¸²à¸à¹€à¸‹à¸´à¸£à¹Œà¸Ÿà¹€à¸§à¸­à¸£à¹Œ)", "à¸ªà¸–à¸²à¸™à¸°à¸ªà¸” (à¸”à¸±à¸šà¹€à¸šà¸´à¸¥à¸„à¸¥à¸´à¸à¹€à¸žà¸·à¹ˆà¸­à¸ªà¹ˆà¸‡ HN à¹„à¸›à¸¥à¸‡à¸—à¸°à¹€à¸šà¸µà¸¢à¸™/à¹à¸à¹‰à¹„à¸‚)")
        gm = mon.grid
        self.table = QtWidgets.QTableWidget(0, 4);
        self.table.setHorizontalHeaderLabels(["ID", "Patient ID", "à¸ªà¸–à¸²à¸™à¸°", "à¹€à¸§à¸¥à¸²"])
        hdr2 = self.table.horizontalHeader();
        hdr2.setStretchLastSection(True)
        hdr2.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        hdr2.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(34)
        gm.addWidget(self.table, 0, 0, 1, 3)
        self.btn_refresh = QtWidgets.QPushButton("à¸£à¸µà¹€à¸Ÿà¸£à¸Š");
        self.btn_refresh.setProperty("variant", "ghost")
        self.btn_export = QtWidgets.QPushButton("Export CSV");
        self.btn_export.setProperty("variant", "ghost")
        self.btn_export_deid = QtWidgets.QPushButton("Export De-Identified (CSV)");
        self.btn_export_deid.setProperty("variant", "ghost")
        gm.addWidget(self.btn_refresh, 1, 0)
        gm.addWidget(self.btn_export, 1, 1)
        gm.addWidget(self.btn_export_deid, 1, 2)
        gm.setColumnStretch(0, 0);
        gm.setColumnStretch(1, 0);
        gm.setColumnStretch(2, 1)
        t3.addWidget(mon, 1)
        self.tabs.addTab(tab3, "Monitor Realtime")

        # signals
        self.btn_refresh.clicked.connect(lambda: self._refresh(True))
        self.btn_export.clicked.connect(self._export_csv)
        self.btn_export_deid.clicked.connect(self._export_deid_csv)
        self.btn_send_runner.clicked.connect(self._on_send_runner_today)
        self.btn_import_excel.clicked.connect(self._on_import_excel)
        self.btn_clear_board.clicked.connect(self._on_clear_board_clicked)
        self.btn_undo_clear.clicked.connect(self._on_undo_clear_clicked)
        self.btn_manage_or.clicked.connect(self._manage_or)
        self.cb_dept.currentTextChanged.connect(self._on_dept_changed)
        self.btn_add.clicked.connect(self._on_add_or_update)
        self.btn_cancel_edit.clicked.connect(self._cancel_edit_mode)
        self.btn_clear.clicked.connect(self._clear_form)
        self.table.itemDoubleClicked.connect(self._on_monitor_double_click)
        self.tree2.itemDoubleClicked.connect(self._on_result_double_click)

        # default period info (auto-calculated)
        self._update_period_info()
        self.date.dateChanged.connect(lambda *_: self._update_period_info())
        self.time.timeChanged.connect(lambda *_: self._update_period_info())

        self._set_doctor_visibility(False)
        self._on_dept_changed(self.cb_dept.currentText())
        self._render_tree2()

    # ---------- settings / timers ----------
    def _current_specialty_key_safe(self) -> str:
        return (getattr(self, "_current_specialty_key", "") or "").strip()

    def _load_settings(self):
        self.cfg = QSettings(ORG_NAME, APP_SETTINGS)
        self.tabs.setCurrentIndex(0)

    def _save_settings(self):
        pass

    def closeEvent(self, e):
        try:
            if self.ws: self.ws.close()
        except Exception:
            pass
        try:
            self._search_executor.shutdown(wait=False)
        except Exception:
            pass
        super().closeEvent(e)

    def _start_timers(self):
        self._pull = QtCore.QTimer(self);
        self._pull.timeout.connect(lambda: self._refresh(True));
        self._pull.start(3000)
        self._seq_timer = QtCore.QTimer(self);
        self._seq_timer.timeout.connect(self._check_seq);
        self._seq_timer.start(1000)
        QtCore.QTimer.singleShot(200, lambda: self._refresh(True))
        QtCore.QTimer.singleShot(600, self._start_ws)
        self._returning_cron = QtCore.QTimer(self)
        self._returning_cron.timeout.connect(self._tick_returning_cron)
        self._returning_cron.start(30_000)

    def _tick_returning_cron(self):
        now = datetime.now()
        changed = False
        alerts: List[Tuple[str, ScheduleEntry]] = []

        for entry in self.sched.entries:
            if entry.state == "returning_to_ward" and entry.returning_started_at:
                t0 = _parse_iso(entry.returning_started_at)
                if not t0 or not entry.time_end:
                    continue
                if (now - t0) >= timedelta(minutes=3):
                    if self._is_entry_completed(entry):
                        entry.postop_completed = True
                        entry.state = "returned_to_ward"
                        entry.returned_to_ward_at = now.strftime("%Y-%m-%dT%H:%M:%S")
                        self._db_insert_case(entry)
                        alerts.append(("ok", entry))
                    else:
                        entry.postop_completed = False
                        entry.state = "postop_pending"
                        entry.returned_to_ward_at = now.strftime("%Y-%m-%dT%H:%M:%S")
                        alerts.append(("warn", entry))
                    entry.version = int(entry.version or 1) + 1
                    changed = True

        if changed:
            self.sched._save()
            self._render_tree2()
            if alerts:
                kind, entry = alerts[-1]
                if kind == "ok":
                    self._banner_returned_ok(entry)
                else:
                    self._banner_incomplete(entry)

    def _db_insert_case(self, entry: "ScheduleEntry"):
        try:
            self.db_logger.log_event(
                case_uid=entry.case_uid,
                event="returned_to_ward",
                details={
                    "hn": entry.hn,
                    "or": entry.or_room,
                    "time_start": entry.time_start,
                    "time_end": entry.time_end,
                    "assist1": entry.assist1,
                    "assist2": entry.assist2,
                    "scrub": entry.scrub,
                    "circulate": entry.circulate,
                    "diags": entry.diags,
                    "ops": entry.ops,
                },
                emergency=str(entry.urgency).lower() == "emergency",
            )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "DB error", str(exc))

    def _banner_incomplete(self, entry: "ScheduleEntry"):
        self.result_banner.set_icon("âš ï¸")
        self.result_banner.set_title("âš ï¸ à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸«à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸„à¸£à¸š â€” à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸šà¸±à¸™à¸—à¸¶à¸à¸¥à¸‡à¸à¸²à¸™à¸‚à¹‰à¸­à¸¡à¸¹à¸¥")
        self.result_banner.set_subtitle(
            f"HN {entry.hn} | OR {entry.or_room} | à¹‚à¸›à¸£à¸”à¸à¸£à¸­à¸à¸—à¸µà¸¡à¸žà¸¢à¸²à¸šà¸²à¸¥/Diagnosis/Operation à¹à¸¥à¸°à¸•à¸£à¸§à¸ˆà¹€à¸§à¸¥à¸²à¹€à¸£à¸´à¹ˆà¸¡â€“à¸ˆà¸š"
        )

    def _banner_returned_ok(self, entry: "ScheduleEntry"):
        self.result_banner.set_icon("âœ…")
        self.result_banner.set_title("à¸šà¸±à¸™à¸—à¸¶à¸à¸¥à¸‡à¸à¸²à¸™à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ªà¸³à¹€à¸£à¹‡à¸ˆ (Returned)")
        self.result_banner.set_subtitle(
            f"HN {entry.hn} | OR {entry.or_room} | à¹€à¸§à¸¥à¸² {entry.time_start or '-'}â€“{entry.time_end or '-'}"
        )

    # ---------- PDPA first-run gate ----------
    def _pdpa_gate(self):
        # à¹€à¸•à¸£à¸µà¸¢à¸¡ salt à¸—à¸±à¸™à¸—à¸µ (à¹ƒà¸Šà¹‰à¸ªà¸³à¸«à¸£à¸±à¸š export à¹à¸šà¸š de-id)
        _get_or_create_secret(SECRET_SALT_KEY, 32)
        # à¹à¸ªà¸”à¸‡ PDPA à¹à¸„à¹ˆà¸„à¸£à¸±à¹‰à¸‡à¹à¸£à¸
        if not self.cfg.value(PDPA_ACK_KEY, False, type=bool):
            dlg = PDPANoticeDialog(self)
            dlg.exec()
            self.cfg.setValue(PDPA_ACK_KEY, True);
            self.cfg.sync()

    # ---------- helpers ----------
    def _client(self):
        try:
            return ClientHTTP(self.ent_host.text().strip() or "127.0.0.1",
                              int(self.ent_port.text().strip() or DEFAULT_PORT),
                              self.ent_token.text().strip() or DEFAULT_TOKEN)
        except Exception:
            return ClientHTTP()

    def _runner_base(self) -> str:
        override = os.getenv("SURGIBOT_RUNNER_BASE_URL")
        if override and override.strip():
            return _resolve_runner_base(override)

        host_text = self.ent_host.text().strip() or DEFAULT_HOST
        parsed = urlparse(host_text if "://" in host_text else f"{_RUNNER_SCHEME_DEFAULT}://{host_text}")
        host = parsed.hostname or host_text.split(":", 1)[0]
        if not host:
            host = DEFAULT_HOST
        if host in {"0.0.0.0", ""}:
            host = "127.0.0.1"
        scheme = parsed.scheme or _RUNNER_SCHEME_DEFAULT
        return _resolve_runner_base(f"{scheme}://{host}")

    def _on_health(self):
        try:
            self._client().health(); self._chip(True)
        except Exception:
            self._chip(False)

    def _chip(self, ok: bool):
        if ok:
            self.status_chip.setText("â— Online");
            self.status_chip.setStyleSheet(
                "color:#10b981;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")
        else:
            self.status_chip.setText("â— Offline");
            self.status_chip.setStyleSheet(
                "color:#ef4444;font-weight:800;padding:6px 10px;border:1px solid #e5e7eb;border-radius:999px;background:#fff;")

    def _refresh(self, prefer_server=True):
        self.btn_refresh.setEnabled(False)
        try:
            data = self._client().list_items()
            rows = extract_rows(data)
            # à¸­à¸±à¸›à¹€à¸”à¸• historical monitor seen à¸à¹ˆà¸­à¸™ render (à¹€à¸à¹‡à¸š HN à¸—à¸µà¹ˆ monitor à¸£à¸²à¸¢à¸‡à¸²à¸™à¸¡à¸²)
            self._scan_monitor_status_transitions(rows)
            self._rebuild_table(rows);
            self._chip(True)
        except Exception:
            self._chip(False);
            self._rebuild_table([])
        finally:
            self.btn_refresh.setEnabled(True)

    def _rebuild_table(self, rows):
        self.rows_cache = rows;
        self.table.setRowCount(0)
        if not rows:
            self.table.setRowCount(1);
            self.table.setSpan(0, 0, 1, 4)
            empty = QtWidgets.QTableWidgetItem("à¹„à¸¡à¹ˆà¸¡à¸µà¸‚à¹‰à¸­à¸¡à¸¹à¸¥ (à¸à¸”à¸£à¸µà¹€à¸Ÿà¸£à¸Š)")
            empty.setFlags(QtCore.Qt.ItemIsEnabled);
            empty.setForeground(QtGui.QBrush(QtGui.QColor("#64748b")))
            self.table.setItem(0, 0, empty)
            # à¸­à¸±à¸›à¹€à¸”à¸• Result tree à¹ƒà¸«à¹‰à¸›à¸£à¸±à¸šà¸•à¸±à¸§à¸à¸£à¸­à¸‡à¸à¸£à¸“à¸µ HN à¸«à¸²à¸¢à¹„à¸›
            self._render_tree2()
            return
        for r in rows:
            i = self.table.rowCount();
            self.table.insertRow(i)
            self.table.setItem(i, 0, QtWidgets.QTableWidgetItem(str(r.get("id", ""))))
            self.table.setItem(i, 1, QtWidgets.QTableWidgetItem(str(r.get("patient_id", ""))))
            status = str(r.get("status", ""))
            col = STATUS_COLORS.get(status, "#64748b")
            chip = StatusChipWidget(status or "-", col, pulse=(status in PULSE_STATUS))
            self.table.setCellWidget(i, 2, chip)
            ts = _parse_iso(r.get("timestamp"));
            txt = ""
            if ts: txt = _fmt_td(datetime.now() - ts)
            self.table.setItem(i, 3, QtWidgets.QTableWidgetItem(txt))
        # à¹ƒà¸«à¹‰ Result tree à¸£à¸µà¹€à¸Ÿà¸£à¸Šà¹€à¸‡à¸·à¹ˆà¸­à¸™à¹„à¸‚à¹à¸ªà¸”à¸‡à¸œà¸¥à¸”à¹‰à¸§à¸¢ à¹€à¸¡à¸·à¹ˆà¸­ monitor à¹€à¸›à¸¥à¸µà¹ˆà¸¢à¸™
        self._render_tree2()

    def _ws_url(self):
        return f"ws://{self.ent_host.text().strip() or '127.0.0.1'}:{int(self.ent_port.text().strip() or DEFAULT_PORT)}{API_WS}?token={self.ent_token.text().strip() or DEFAULT_TOKEN}"

    def _start_ws(self):
        try:
            self.ws = QWebSocket()
            self.ws.errorOccurred.connect(lambda _e: self._ws_disc())
            self.ws.connected.connect(lambda: (self._chip(True), self._pull.stop()))
            self.ws.disconnected.connect(self._ws_disc)
            self.ws.textMessageReceived.connect(self._on_ws_msg)
            self.ws.open(QUrl(self._ws_url()))
        except Exception:
            self._ws_disc()

    def _ws_disc(self):
        if self._pull.isActive() == False: self._pull.start(3000)

    def _on_ws_msg(self, msg):
        try:
            rows = extract_rows(json.loads(msg))
            self._scan_monitor_status_transitions(rows)
            self._rebuild_table(rows)
        except Exception:
            pass

    # ---------- schedule ----------
    def _refresh_or_cb(self, cb: QtWidgets.QComboBox):
        cb.clear();
        cb.addItems(self.sched.or_rooms)

    def _manage_or(self):
        dlg = QtWidgets.QDialog(self);
        dlg.setWindowTitle("à¸ˆà¸±à¸”à¸à¸²à¸£ OR");
        v = QtWidgets.QVBoxLayout(dlg)
        lst = QtWidgets.QListWidget();
        lst.addItems(self.sched.or_rooms);
        v.addWidget(lst)
        h = QtWidgets.QHBoxLayout();
        ent = QtWidgets.QLineEdit();
        ent.setPlaceholderText("à¹€à¸Šà¹ˆà¸™ OR9");
        btn_add = QtWidgets.QPushButton("à¹€à¸žà¸´à¹ˆà¸¡");
        btn_del = QtWidgets.QPushButton("à¸¥à¸š");
        h.addWidget(ent, 1);
        h.addWidget(btn_add);
        h.addWidget(btn_del);
        v.addLayout(h)
        ok = QtWidgets.QPushButton("à¸šà¸±à¸™à¸—à¸¶à¸");
        ok.setProperty("variant", "primary");
        v.addWidget(ok)
        btn_add.clicked.connect(
            lambda: (ent.text().strip().upper() and lst.addItem(ent.text().strip().upper()), ent.setText("")))
        btn_del.clicked.connect(lambda: [lst.takeItem(lst.row(x)) for x in lst.selectedItems()])

        def save():
            rooms = [lst.item(i).text() for i in range(lst.count())];
            self.sched.set_or_rooms(rooms);
            self._refresh_or_cb(self.cb_or);
            dlg.accept()

        ok.clicked.connect(save);
        dlg.exec()

    def _entries_of_selected_date(self) -> List["ScheduleEntry"]:
        try:
            qdate = self.date.date()
            if hasattr(qdate, "toPython"):
                day = qdate.toPython()
            else:
                day = date(qdate.year(), qdate.month(), qdate.day())
        except Exception:
            day = datetime.now().date()

        matches: List[ScheduleEntry] = []

        for entry in list(getattr(self.sched, "entries", [])):
            entry_day = getattr(entry, "date", None)
            if isinstance(entry_day, datetime):
                entry_day = entry_day.date()
            elif isinstance(entry_day, str):
                try:
                    entry_day = datetime.fromisoformat(entry_day).date()
                except Exception:
                    continue
            if entry_day == day:
                matches.append(entry)
        return matches

    def _pickup_id_for_entry(self, entry: "ScheduleEntry", override_or: Optional[str] = None) -> str:
        entry_or = override_or if override_or is not None else getattr(entry, "or_room", "")
        payload = {
            "date": str(getattr(entry, "date", date.today())),
            "HN": getattr(entry, "hn", ""),
            "OR": entry_or,
        }
        return _pickup_id_for_row(payload)

    def _coerce_time_value(self, value) -> str:
        if value in (None, "", "TF"):
            return ""
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, dtime):
            return value.strftime("%H:%M")
        text = str(value).strip()
        if not text or text.upper() == "TF":
            return ""
        parts = text.split(":")
        if len(parts) >= 2 and all(part.isdigit() for part in parts[:2]):
            try:
                hh, mm = int(parts[0]), int(parts[1])
                if 0 <= hh <= 23 and 0 <= mm <= 59:
                    return f"{hh:02d}:{mm:02d}"
            except Exception:
                return ""
        return text

    def _entry_to_runner_payload(self, entry: "ScheduleEntry", override_or: Optional[str] = None) -> Optional[dict]:
        hn = (entry.hn or "").strip()
        or_room = (override_or if override_or is not None else entry.or_room or "").strip()
        if not hn or not or_room:
            return None

        start_value = getattr(entry, "time_start", "") or getattr(entry, "time", "")
        start_time = self._coerce_time_value(start_value)

        pickup_id = self._pickup_id_for_entry(entry, or_room)

        return {
            "pickup_id": pickup_id,
            "date": str(getattr(entry, "date", date.today())),
            "hn": hn,
            "name": getattr(entry, "name", ""),
            "ward_from": getattr(entry, "ward", ""),
            "or_to": or_room,
            "call_time": datetime.now().strftime("%H:%M"),
            "due_time": "",
            "status": "waiting",
            "assignee": "",
            "ack_time": "",
            "start_time": start_time,
            "arrive_time": "",
            "note": getattr(entry, "note", "") if hasattr(entry, "note") else "",
        }

    def _push_rows_to_runner(
            self,
            entries: List["ScheduleEntry"],
            *,
            runner_ready: Optional[bool] = None,
            collect_failures: bool = False,
    ) -> Tuple[int, List[str]]:
        if not entries:
            return (0, [])

        base = self._runner_base()
        if runner_ready is None:
            runner_ready = runner_health_ok(base)
        if not runner_ready:
            return (0, [])

        ok = 0
        failed: List[str] = []
        url = f"{base}{RUNNER_UPDATE_API}"

        for entry in entries:
            payload = self._entry_to_runner_payload(entry)
            if not payload:
                continue
            try:
                resp = requests.post(url, json=payload, timeout=2.0, headers={"Accept": "application/json"})
                resp.raise_for_status()
                ok += 1
            except requests.RequestException:
                if collect_failures:
                    failed.append(payload.get("hn") or payload.get("pickup_id") or "-")
        return ok, failed

    def _runner_status_label(self, status: str) -> str:
        status = (status or "").strip()
        return RUNNER_STATUS_LABELS.get(status, status)

    def _runner_status_tooltip(self, payload: dict) -> str:
        hints: List[str] = []
        mapping = [
            ("status", "à¸ªà¸–à¸²à¸™à¸°"),
            ("assignee", "à¸œà¸¹à¹‰à¸£à¸±à¸šà¹€à¸„à¸ª"),
            ("ack_time", "à¹€à¸§à¸¥à¸²à¸£à¸±à¸šà¹€à¸„à¸ª"),
            ("start_time", "à¹€à¸§à¸¥à¸²à¹€à¸£à¸´à¹ˆà¸¡à¸ªà¹ˆà¸‡"),
            ("arrive_time", "à¸–à¸¶à¸‡ OR"),
            ("note", "à¸«à¸¡à¸²à¸¢à¹€à¸«à¸•à¸¸"),
        ]
        for key, label in mapping:
            value = payload.get(key)
            if value:
                hints.append(f"{label}: {value}")
        return "\n".join(hints)

    def _ask_runner_name(self) -> str:
        text, ok = QtWidgets.QInputDialog.getText(
            self,
            "à¸Šà¸·à¹ˆà¸­à¸œà¸¹à¹‰à¹„à¸›à¸£à¸±à¸šà¹€à¸„à¸ª",
            "à¸à¸£à¸¸à¸“à¸²à¸£à¸°à¸šà¸¸à¸Šà¸·à¹ˆà¸­à¹€à¸ˆà¹‰à¸²à¸«à¸™à¹‰à¸²à¸—à¸µà¹ˆ Runner:",
            QtWidgets.QLineEdit.Normal,
            self._last_runner_user,
        )
        if not ok:
            return ""
        text = str(text).strip()
        if text:
            self._last_runner_user = text
        return text

    def _runner_ack(self, pickup_id: str, user: str) -> bool:
        try:
            resp = requests.post(
                f"{self._runner_base()}{RUNNER_ACK_API}",
                json={"pickup_id": pickup_id, "user": user},
                timeout=2.0,
                headers={"Accept": "application/json"},
            )
            return bool(resp.ok)
        except requests.RequestException:
            return False

    def _runner_arrive(self, pickup_id: str, user: str) -> bool:
        try:
            resp = requests.post(
                f"{self._runner_base()}{RUNNER_ARRIVE_API}",
                json={"pickup_id": pickup_id, "user": user},
                timeout=2.0,
                headers={"Accept": "application/json"},
            )
            return bool(resp.ok)
        except requests.RequestException:
            return False

    def _runner_finish(self, pickup_id: str, user: str = "à¸£à¸°à¸šà¸š") -> bool:
        try:
            resp = requests.post(
                f"{self._runner_base()}{RUNNER_FINISH_API}",
                json={"pickup_id": pickup_id, "user": user},
                timeout=2.0,
                headers={"Accept": "application/json"},
            )
            return bool(resp.ok)
        except requests.RequestException:
            return False

    def _auto_finish_runner_cases(self, entries: List["ScheduleEntry"], status_map: Dict[str, dict]) -> None:
        if not entries or not status_map:
            return
        for entry in entries:
            if not self._is_entry_completed(entry):
                continue
            pickup_id = self._pickup_id_for_entry(entry)
            if not pickup_id:
                continue
            row = status_map.get(pickup_id) or {}
            if str(row.get("status") or "").strip() == "finished":
                self._runner_finished_sent.discard(pickup_id)
                continue
            if pickup_id in self._runner_finished_sent:
                continue
            if self._runner_finish(pickup_id, user="à¸£à¸°à¸šà¸š"):
                self._runner_finished_sent.add(pickup_id)

    def _handle_runner_action(self, entry: "ScheduleEntry", action: str) -> None:
        pid = self._pickup_id_for_entry(entry)
        if not pid:
            self.toast.show_toast("à¹„à¸¡à¹ˆà¸žà¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥ OR/HN à¸ªà¸³à¸«à¸£à¸±à¸š Runner")
            return
        base = self._runner_base()
        if not runner_health_ok(base):
            self.toast.show_toast("à¹„à¸¡à¹ˆà¸ªà¸²à¸¡à¸²à¸£à¸–à¹€à¸Šà¸·à¹ˆà¸­à¸¡à¸•à¹ˆà¸­ Runner à¹„à¸”à¹‰")
            return
        user = self._ask_runner_name()
        if not user:
            return
        if action == "ack":
            ok = self._runner_ack(pid, user)
            success_msg = "à¸£à¸±à¸šà¹€à¸„à¸ªà¹€à¸£à¸µà¸¢à¸šà¸£à¹‰à¸­à¸¢"
        else:
            ok = self._runner_arrive(pid, user)
            success_msg = "à¸šà¸±à¸™à¸—à¸¶à¸à¸–à¸¶à¸‡ OR à¹à¸¥à¹‰à¸§"
        if ok:
            self.toast.show_toast(success_msg)
            self._render_tree2()
        else:
            self.toast.show_toast("à¸ªà¹ˆà¸‡à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹„à¸› Runner à¹„à¸¡à¹ˆà¸ªà¸³à¹€à¸£à¹‡à¸ˆ")

    def _on_send_runner_today(self):
        rows = self._entries_of_selected_date()
        if not rows:
            SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£", "à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µà¹€à¸„à¸ªà¸‚à¸­à¸‡à¸§à¸±à¸™à¸—à¸µà¹ˆà¹€à¸¥à¸·à¸­à¸à¸—à¸µà¹ˆà¸ˆà¸°à¸ªà¹ˆà¸‡à¹ƒà¸«à¹‰ Runner")
            return

        base = self._runner_base()
        runner_ready = runner_health_ok(base)
        if not runner_ready:
            SweetAlert.warning(
                self,
                "à¹„à¸¡à¹ˆà¸ªà¸²à¸¡à¸²à¸£à¸–à¹€à¸Šà¸·à¹ˆà¸­à¸¡à¸•à¹ˆà¸­",
                f"à¹€à¸Šà¸·à¹ˆà¸­à¸¡à¸•à¹ˆà¸­ Runner à¹„à¸¡à¹ˆà¹„à¸”à¹‰ (à¸•à¸£à¸§à¸ˆà¸ªà¸­à¸š {base})",
            )
            return

        dlg = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹„à¸› Runner ...")
        dlg.show()
        QtWidgets.QApplication.processEvents()
        try:
            ok, failed = self._push_rows_to_runner(rows, runner_ready=runner_ready, collect_failures=True)
        finally:
            dlg.close()

        if ok > 0 and not failed:
            SweetAlert.success(self, "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", f"à¸ªà¹ˆà¸‡à¹ƒà¸«à¹‰ Runner à¹à¸¥à¹‰à¸§ {ok} à¸£à¸²à¸¢à¸à¸²à¸£", auto_close_msec=1600)
        elif ok > 0 and failed:
            SweetAlert.success(
                self,
                "à¸ªà¸³à¹€à¸£à¹‡à¸ˆà¸šà¸²à¸‡à¸ªà¹ˆà¸§à¸™",
                f"à¸ªà¸³à¹€à¸£à¹‡à¸ˆ {ok} â€¢ à¸¥à¹‰à¸¡à¹€à¸«à¸¥à¸§ {len(failed)}\n(HN: {', '.join(failed[:10])}{' â€¦' if len(failed) > 10 else ''})",
            )
        else:
            SweetAlert.warning(
                self,
                "à¹„à¸¡à¹ˆà¸ªà¸³à¹€à¸£à¹‡à¸ˆ",
                f"à¸ªà¹ˆà¸‡à¹ƒà¸«à¹‰ Runner à¹„à¸¡à¹ˆà¹„à¸”à¹‰à¹€à¸¥à¸¢ â€” à¸•à¸£à¸§à¸ˆà¸ªà¸­à¸šà¸§à¹ˆà¸²à¹€à¸‹à¸´à¸£à¹Œà¸Ÿà¹€à¸§à¸­à¸£à¹Œ Runner à¹€à¸›à¸´à¸”à¸­à¸¢à¸¹à¹ˆà¸—à¸µà¹ˆ {base} à¸«à¸£à¸·à¸­à¹„à¸¡à¹ˆ",
            )

        self._render_tree2()

    def _on_import_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "à¹€à¸¥à¸·à¸­à¸à¹„à¸Ÿà¸¥à¹Œà¸™à¸³à¹€à¸‚à¹‰à¸²à¸•à¸²à¸£à¸²à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”",
            str(Path.home()),
            "Excel/CSV (*.xlsx *.xlsm *.xls *.csv)"
        )
        if not path:
            return

        loader = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¸­à¹ˆà¸²à¸™à¹„à¸Ÿà¸¥à¹Œà¸™à¸³à¹€à¸‚à¹‰à¸²...")
        loader.setLabelText("à¸à¸³à¸¥à¸±à¸‡à¸­à¹ˆà¸²à¸™à¹„à¸Ÿà¸¥à¹Œà¸™à¸³à¹€à¸‚à¹‰à¸²...")
        loader.show()
        QtWidgets.QApplication.processEvents()

        error_title: Optional[str] = None
        error_message: Optional[str] = None
        empty_rows = False
        ok = 0
        skipped: List[Tuple[str, str]] = []

        try:
            rows = self._load_fixed_excel_rows(path)
            if not rows:
                empty_rows = True
            else:
                loader.setLabelText("à¸à¸³à¸¥à¸±à¸‡à¸™à¸³à¹€à¸‚à¹‰à¸²à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹€à¸‚à¹‰à¸²à¸ªà¸¹à¹ˆà¸•à¸²à¸£à¸²à¸‡...")
                QtWidgets.QApplication.processEvents()
                ok, skipped = self._import_from_fixed_excel_rows(rows)
        except ImportError as exc:
            error_title = "à¸™à¸³à¹€à¸‚à¹‰à¸²à¹„à¸¡à¹ˆà¹„à¸”à¹‰"
            error_message = str(exc)
        except Exception as exc:
            error_title = "à¸™à¸³à¹€à¸‚à¹‰à¸²à¸¥à¹‰à¸¡à¹€à¸«à¸¥à¸§"
            error_message = str(exc)
        finally:
            loader.close()

        if error_message:
            QtWidgets.QMessageBox.critical(self, error_title or "à¸œà¸´à¸”à¸žà¸¥à¸²à¸”", error_message)
            return

        if empty_rows:
            SweetAlert.warning(self, "à¹€à¸•à¸·à¸­à¸™", "à¹„à¸¡à¹ˆà¸žà¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¹ƒà¸™à¹„à¸Ÿà¸¥à¹Œà¸—à¸µà¹ˆà¹€à¸¥à¸·à¸­à¸")
            return

        if ok <= 0:
            SweetAlert.warning(self, "à¹€à¸•à¸·à¸­à¸™", "à¹„à¸¡à¹ˆà¸ªà¸²à¸¡à¸²à¸£à¸–à¸™à¸³à¹€à¸‚à¹‰à¸²à¹à¸–à¸§à¹ƒà¸”à¹„à¸”à¹‰")
            if skipped:
                QtWidgets.QMessageBox.information(
                    self,
                    "à¸£à¸²à¸¢à¸¥à¸°à¹€à¸­à¸µà¸¢à¸”à¹à¸–à¸§à¸—à¸µà¹ˆà¸‚à¹‰à¸²à¸¡",
                    "\n".join([f"HN {hn}: {reason}" for hn, reason in skipped[:20]])
                    + ("\nâ€¦ (à¸¡à¸µà¸¡à¸²à¸à¸à¸§à¹ˆà¸²à¸™à¸µà¹‰)" if len(skipped) > 20 else ""),
                )
            return

        msg = f"à¸™à¸³à¹€à¸‚à¹‰à¸²à¸ªà¸³à¹€à¸£à¹‡à¸ˆ {ok} à¹à¸–à¸§"
        if skipped:
            msg += f" â€¢ à¸‚à¹‰à¸²à¸¡ {len(skipped)} à¹à¸–à¸§"
        SweetAlert.success(self, "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", msg, auto_close_msec=1500)

        if skipped:
            QtWidgets.QMessageBox.information(
                self,
                "à¸£à¸²à¸¢à¸¥à¸°à¹€à¸­à¸µà¸¢à¸”à¹à¸–à¸§à¸—à¸µà¹ˆà¸‚à¹‰à¸²à¸¡",
                "\n".join([f"HN {hn}: {reason}" for hn, reason in skipped[:20]])
                + ("\nâ€¦ (à¸¡à¸µà¸¡à¸²à¸à¸à¸§à¹ˆà¸²à¸™à¸µà¹‰)" if len(skipped) > 20 else ""),
            )

    def _load_fixed_excel_rows(self, path: str) -> List[dict]:
        suffix = Path(path).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as exc:  # pragma: no cover - runtime dependency
                raise ImportError("à¸•à¹‰à¸­à¸‡à¸•à¸´à¸”à¸•à¸±à¹‰à¸‡ openpyxl à¹€à¸žà¸·à¹ˆà¸­à¸­à¹ˆà¸²à¸™à¹„à¸Ÿà¸¥à¹Œ Excel") from exc

            wb = load_workbook(path, data_only=True)
            sheet = wb.active
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                return []

            headers = [str(col).strip() if col is not None else "" for col in header]
            results: List[dict] = []
            for row in rows_iter:
                row_dict: Dict[str, object] = {}
                has_value = False
                for key, value in zip(headers, row):
                    if not key:
                        continue
                    row_dict[key] = value if value is not None else ""
                    if not has_value and str(value or "").strip():
                        has_value = True
                if row_dict and has_value:
                    results.append(row_dict)
            return results

        if suffix == ".csv":
            results: List[dict] = []
            with open(path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                if reader.fieldnames is None:
                    return []
                for row in reader:
                    row_dict: Dict[str, object] = {}
                    has_value = False
                    for key, value in row.items():
                        if not key:
                            continue
                        key_clean = str(key).strip()
                        row_dict[key_clean] = value if value is not None else ""
                        if not has_value and str(value or "").strip():
                            has_value = True
                    if row_dict and has_value:
                        results.append(row_dict)
            return results

        raise ValueError("à¸£à¸­à¸‡à¸£à¸±à¸šà¹€à¸‰à¸žà¸²à¸°à¹„à¸Ÿà¸¥à¹Œ Excel (.xlsx/.xlsm) à¸«à¸£à¸·à¸­ CSV")

    def _import_from_fixed_excel_rows(self, rows: List[dict]) -> Tuple[int, List[Tuple[str, str]]]:
        ok = 0
        skipped: List[Tuple[str, str]] = []

        known_wards: List[str] = []
        if hasattr(self, "cb_ward") and isinstance(self.cb_ward, QtWidgets.QComboBox):
            for i in range(self.cb_ward.count()):
                text = self.cb_ward.itemText(i).strip()
                if text:
                    known_wards.append(text)
        else:
            known_wards = [w for w in WARD_LIST if w and w != WARD_PLACEHOLDER]

        qdate = self.date.date() if hasattr(self, "date") else QtCore.QDate.currentDate()
        if hasattr(qdate, "toPython"):
            base_date = qdate.toPython()
        else:  # pragma: no cover - fallback for older Qt bindings
            base_date = datetime(qdate.year(), qdate.month(), qdate.day()).date()

        default_period = self._update_period_info()

        for raw in rows:
            if not isinstance(raw, dict):
                continue

            lookup: Dict[str, object] = {}
            for key, value in raw.items():
                key_str = str(key or "").strip()
                if not key_str:
                    continue
                lookup[key_str] = value
                lookup[key_str.lower()] = value

            def get(field: str) -> str:
                header = FIXED_MAPPING_TH[field]
                val = lookup.get(header)
                if val is None:
                    val = lookup.get(header.strip())
                if val is None:
                    val = lookup.get(header.lower())
                return str(val or "").strip()

            time_raw = lookup.get(FIXED_MAPPING_TH["time"])
            time_str = parse_time_hhmm_or_tf(time_raw)
            hn = get("hn")
            name = get("name")
            doctor = get("doctor")

            if not (hn and name and doctor):
                skipped.append((hn or "-", "à¸•à¹‰à¸­à¸‡à¸¡à¸µ HN, à¸Šà¸·à¹ˆà¸­, à¹à¸žà¸—à¸¢à¹Œà¸œà¸¹à¹‰à¸ªà¸±à¹ˆà¸‡"))
                continue

            if time_str != "TF":
                try:
                    hh, mm = [int(x) for x in time_str.split(":", 1)]
                    period_code = _now_period(datetime(base_date.year, base_date.month, base_date.day, hh, mm))
                except Exception:
                    period_code = default_period
            else:
                period_code = default_period

            diag_txt = get("diags")
            op_txt = get("ops")

            or_room = pick_or_by_doctor(base_date, time_str, doctor)

            entry = ScheduleEntry(
                or_room=or_room or "-",
                dt=base_date,
                time_str=time_str,
                hn=hn,
                name=" ".join(name.split()),
                age=parse_age_years(get("age")),
                dept="",
                doctor=normalize_doctor(doctor),
                diags=[diag_txt] if diag_txt else [],
                ops=[op_txt] if op_txt else [],
                ward=map_to_known_ward(get("ward"), known_wards),
                case_size="",
                period=period_code,
                urgency="Elective",
                assist1="",
                assist2="",
                scrub="",
                circulate="",
                time_start="",
                time_end="",
            )

            self.sched.add(entry)
            try:
                self.db_logger.append_entry(entry)
            except Exception:
                pass
            ok += 1

        self._set_result_title()
        self._render_tree2()

        return ok, skipped

    def _snapshot_sched(self) -> None:
        try:
            self._last_snapshot = [entry.to_dict() for entry in self.sched.all()]
        except Exception:
            self._last_snapshot = None
        if hasattr(self, "btn_undo_clear"):
            self.btn_undo_clear.setEnabled(bool(self._last_snapshot))

    def _restore_snapshot(self) -> None:
        if not self._last_snapshot:
            SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸¡à¸µà¸ªà¸³à¹€à¸™à¸²", "à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µà¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ªà¸³à¸«à¸£à¸±à¸šà¸¢à¹‰à¸­à¸™à¸à¸¥à¸±à¸š")
            return

        restored: List[ScheduleEntry] = []
        for payload in self._last_snapshot:
            try:
                restored.append(ScheduleEntry.from_dict(payload))
            except Exception:
                continue

        self._last_snapshot = None
        if hasattr(self, "btn_undo_clear"):
            self.btn_undo_clear.setEnabled(False)

        self.sched.replace_all(restored)
        self._set_result_title()
        self._render_tree2()
        SweetAlert.success(self, "à¹€à¸£à¸µà¸¢à¸šà¸£à¹‰à¸­à¸¢", "à¸à¸¹à¹‰à¸„à¸·à¸™à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸¥à¹ˆà¸²à¸ªà¸¸à¸”à¹à¸¥à¹‰à¸§", auto_close_msec=1500)

    def _on_clear_board_clicked(self) -> None:
        entries = self.sched.all()
        if not entries:
            SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸¡à¸µà¸‚à¹‰à¸­à¸¡à¸¹à¸¥", "à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸™à¸•à¸²à¸£à¸²à¸‡à¹ƒà¸«à¹‰à¸¥à¹‰à¸²à¸‡")
            return

        box = QtWidgets.QMessageBox(self)
        box.setWindowTitle("à¸¢à¸·à¸™à¸¢à¸±à¸™à¸à¸²à¸£à¸¥à¹‰à¸²à¸‡à¸à¸£à¸°à¸”à¸²à¸™")
        box.setIcon(QtWidgets.QMessageBox.Question)
        box.setText("à¸•à¹‰à¸­à¸‡à¸à¸²à¸£à¸¥à¹‰à¸²à¸‡à¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸™à¸•à¸²à¸£à¸²à¸‡à¸ªà¸³à¸«à¸£à¸±à¸šà¸—à¸”à¸ªà¸­à¸šà¹ƒà¸Šà¹ˆà¹„à¸«à¸¡?")
        btn_today = box.addButton("à¸¥à¹‰à¸²à¸‡à¹€à¸‰à¸žà¸²à¸°à¸§à¸±à¸™à¸›à¸±à¸ˆà¸ˆà¸¸à¸šà¸±à¸™", QtWidgets.QMessageBox.AcceptRole)
        btn_all = box.addButton("à¸¥à¹‰à¸²à¸‡à¸—à¸±à¹‰à¸‡à¸«à¸¡à¸”", QtWidgets.QMessageBox.DestructiveRole)
        btn_cancel = box.addButton("à¸¢à¸à¹€à¸¥à¸´à¸", QtWidgets.QMessageBox.RejectRole)
        box.setDefaultButton(btn_today)
        box.exec()

        clicked = box.clickedButton()
        if clicked is None or clicked is btn_cancel:
            return

        if clicked is btn_today:
            qdate = self.date.date() if hasattr(self, "date") else QtCore.QDate.currentDate()
            day = qdate.toPython()
            todays = [e for e in entries if getattr(e, "date", None) == day]
            if not todays:
                SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸žà¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥", f"à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£à¸§à¸±à¸™à¸—à¸µà¹ˆ {day.strftime('%d/%m/%Y')} à¹ƒà¸«à¹‰à¸¥à¹‰à¸²à¸‡")
                return

            self._snapshot_sched()
            removed = self.sched.remove_by_date(day)
            if removed <= 0:
                self._last_snapshot = None
                if hasattr(self, "btn_undo_clear"):
                    self.btn_undo_clear.setEnabled(False)
                SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸žà¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥", f"à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£à¸§à¸±à¸™à¸—à¸µà¹ˆ {day.strftime('%d/%m/%Y')} à¹ƒà¸«à¹‰à¸¥à¹‰à¸²à¸‡")
                return

            message = f"à¸¥à¸š {removed} à¸£à¸²à¸¢à¸à¸²à¸£à¸‚à¸­à¸‡à¸§à¸±à¸™à¸—à¸µà¹ˆ {day.strftime('%d/%m/%Y')} à¹à¸¥à¹‰à¸§"
        else:
            self._snapshot_sched()
            removed = self.sched.clear()
            if removed <= 0:
                self._last_snapshot = None
                if hasattr(self, "btn_undo_clear"):
                    self.btn_undo_clear.setEnabled(False)
                SweetAlert.info(self, "à¹„à¸¡à¹ˆà¸¡à¸µà¸‚à¹‰à¸­à¸¡à¸¹à¸¥", "à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸™à¸•à¸²à¸£à¸²à¸‡à¹ƒà¸«à¹‰à¸¥à¹‰à¸²à¸‡")
                return

            message = f"à¸¥à¸šà¸—à¸±à¹‰à¸‡à¸«à¸¡à¸” {removed} à¸£à¸²à¸¢à¸à¸²à¸£à¹à¸¥à¹‰à¸§"

        self._set_result_title()
        self._render_tree2()
        SweetAlert.success(self, "à¸¥à¹‰à¸²à¸‡à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", message, auto_close_msec=1500)

    def _on_undo_clear_clicked(self) -> None:
        self._restore_snapshot()

    def _update_period_info(self):
        qd = self.date.date()
        qtime = self.time.time()
        dt = datetime(qd.year(), qd.month(), qd.day(), qtime.hour(), qtime.minute())
        auto = _now_period(dt)
        if hasattr(self, "lbl_period_info"):
            self.lbl_period_info.setText(
                f"à¸£à¸°à¸šà¸šà¸à¸³à¸«à¸™à¸”à¸Šà¹ˆà¸§à¸‡à¹€à¸§à¸¥à¸²à¸­à¸±à¸•à¹‚à¸™à¸¡à¸±à¸•à¸´: {_period_label(auto)} (à¸­à¹‰à¸²à¸‡à¸­à¸´à¸‡ {dt:%d/%m/%Y %H:%M})"
            )
        return auto

    def _on_dept_changed(self, dept_label: str):
        if dept_label and not dept_label.startswith("â€”"):
            self._set_doctor_visibility(True)
            self.cb_doctor.clear()
            self.cb_doctor.addItems(DEPT_DOCTORS.get(dept_label, []))
        else:
            self._set_doctor_visibility(False)
            self.cb_doctor.clear()

        specialty = _dept_to_specialty_key(dept_label or "")
        self._current_specialty_key = specialty

        if not specialty:
            self._diag_base_catalog = []
            self._diag_catalog_full = []
            self._op_catalog_full = []
            self._dx_index = None
            self._op_index = None
            if self.diag_adder.search_line:
                self.diag_adder.search_line.clear()
            if self.op_adder.search_line:
                self.op_adder.search_line.clear()
            self.diag_adder.clear()
            self.op_adder.clear()
            self.diag_adder.set_suggestions([])
            self.op_adder.set_suggestions([])
            self.diag_adder.setEnabled(False)
            self.op_adder.setEnabled(False)
            return

        loader = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¹€à¸•à¸£à¸µà¸¢à¸¡à¸£à¸²à¸¢à¸à¸²à¸£à¸ªà¸³à¸«à¸£à¸±à¸šà¹à¸œà¸™à¸à¸™à¸µà¹‰...")
        QtWidgets.QApplication.processEvents()
        try:
            base_ops = get_operations(specialty) or []
            base_dx_list = get_diagnoses(specialty) or []
            user_op = get_custom_entries("operation", specialty) or []
            user_dx = get_custom_entries("diagnosis", specialty) or []
        finally:
            loader.close()

        merged_ops: List[str] = []
        for bucket in (user_op, base_ops):
            for value in bucket:
                val = (value or "").strip()
                if val and val not in merged_ops:
                    merged_ops.append(val)

        merged_dx: List[str] = []
        for bucket in (user_dx, base_dx_list):
            for value in bucket:
                val = (value or "").strip()
                if val and val not in merged_dx:
                    merged_dx.append(val)

        self._op_catalog_full = merged_ops
        self._diag_base_catalog = merged_dx
        self._diag_catalog_full = list(merged_dx)

        self._op_index = FastSearchIndex(merged_ops, prefix_len=3) if merged_ops else None
        self._dx_index = FastSearchIndex(merged_dx, prefix_len=3) if merged_dx else None

        if self.op_adder.search_line:
            self.op_adder.search_line.clear()
        if self.diag_adder.search_line:
            self.diag_adder.search_line.clear()

        self.op_adder.clear()
        self.diag_adder.clear()

        self.op_adder.setEnabled(True)
        self.diag_adder.setEnabled(True)

        self._latest_op_query = ""
        self._latest_diag_query = ""

        initial_ops = (
            self._op_index.search("", self._op_search_limit) if self._op_index else merged_ops[: self._op_search_limit]
        )
        initial_dx = (
            self._dx_index.search("", self._dx_search_limit) if self._dx_index else merged_dx[: self._dx_search_limit]
        )

        self.op_adder.set_suggestions(initial_ops)
        self.diag_adder.set_suggestions(initial_dx)

        self._refresh_diag_suggestions()

    def _on_operation_persist_requested(self, text: str):
        item = (text or "").strip()
        if not item:
            return
        specialty = self._current_specialty_key_safe()
        if not specialty:
            SweetAlert.warning(self, "à¹€à¸•à¸·à¸­à¸™", "à¸à¸£à¸¸à¸“à¸²à¹€à¸¥à¸·à¸­à¸à¹à¸œà¸™à¸à¸à¹ˆà¸­à¸™")
            return
        loader = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¸šà¸±à¸™à¸—à¸¶à¸ Operation ...")
        QtWidgets.QApplication.processEvents()
        try:
            added = add_custom_entry("operation", specialty, item)
        finally:
            loader.close()
        self._on_dept_changed(self.cb_dept.currentText())
        SweetAlert.success(
            self,
            "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ" if added else "à¸‹à¹‰à¸³",
            "à¸šà¸±à¸™à¸—à¸¶à¸ Operation à¹€à¸žà¸´à¹ˆà¸¡à¹€à¸‚à¹‰à¸²à¸„à¸¥à¸±à¸‡à¹à¸¥à¹‰à¸§" if added else "à¸¡à¸µ Operation à¸™à¸µà¹‰à¸­à¸¢à¸¹à¹ˆà¹à¸¥à¹‰à¸§",
        )

    def _on_diagnosis_persist_requested(self, text: str):
        item = (text or "").strip()
        if not item:
            return
        specialty = self._current_specialty_key_safe()
        if not specialty:
            SweetAlert.warning(self, "à¹€à¸•à¸·à¸­à¸™", "à¸à¸£à¸¸à¸“à¸²à¹€à¸¥à¸·à¸­à¸à¹à¸œà¸™à¸à¸à¹ˆà¸­à¸™")
            return
        loader = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¸šà¸±à¸™à¸—à¸¶à¸ Diagnosis ...")
        QtWidgets.QApplication.processEvents()
        try:
            added = add_custom_entry("diagnosis", specialty, item)
        finally:
            loader.close()
        self._on_dept_changed(self.cb_dept.currentText())
        SweetAlert.success(
            self,
            "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ" if added else "à¸‹à¹‰à¸³",
            "à¸šà¸±à¸™à¸—à¸¶à¸ Diagnosis à¹€à¸žà¸´à¹ˆà¸¡à¹€à¸‚à¹‰à¸²à¸„à¸¥à¸±à¸‡à¹à¸¥à¹‰à¸§" if added else "à¸¡à¸µ Diagnosis à¸™à¸µà¹‰à¸­à¸¢à¸¹à¹ˆà¹à¸¥à¹‰à¸§",
        )

    def _refresh_diag_suggestions(self):
        specialty = self._current_specialty_key_safe()
        if not specialty:
            self._diag_catalog_full = []
            self._dx_index = None
            self.diag_adder.set_suggestions([])
            return

        base_suggestions = diagnosis_suggestions(specialty)

        merged: List[str] = []

        def _append(values: List[str]) -> None:
            for value in values or []:
                val = (value or "").strip()
                if val and val not in merged:
                    merged.append(val)

        _append(self._diag_base_catalog)
        _append(list(base_suggestions))

        self._diag_catalog_full = merged
        self._dx_index = FastSearchIndex(self._diag_catalog_full, prefix_len=3) if self._diag_catalog_full else None
        if self.diag_adder.search_line:
            self._latest_diag_query = self.diag_adder.search_line.text()
        if self._dx_index:
            initial = self._dx_index.search(self._latest_diag_query, self._dx_search_limit)
        else:
            initial = []
        self.diag_adder.set_suggestions(initial)

    def _on_operations_changed(self, _items: List[str]):
        self._refresh_diag_suggestions()

    def _on_diag_query_changed(self, text: str):
        self._latest_diag_query = text or ""
        if not self._dx_index:
            return
        self._search_timer.stop()
        self._search_timer.start()

    def _on_diag_search_timeout(self):
        self._run_diag_search(self._latest_diag_query)

    def _run_diag_search(self, query: str):
        if not self._dx_index:
            self.diag_adder.set_suggestions([])
            return
        future = self._search_executor.submit(self._dx_index.search, query, self._dx_search_limit)

        def _apply(fut):
            try:
                results = fut.result()
            except Exception:
                results = []
            if query != self._latest_diag_query:
                return
            self.diag_adder.set_suggestions(results)

        future.add_done_callback(lambda fut: QtCore.QTimer.singleShot(0, lambda: _apply(fut)))

    def _on_op_query_changed(self, text: str):
        self._latest_op_query = text or ""
        if self._op_index:
            self._op_search_timer.stop()
            self._op_search_timer.start()
        else:
            self._run_op_search(self._latest_op_query)

    def _on_op_search_timeout(self):
        self._run_op_search(self._latest_op_query)

    def _run_op_search(self, query: str):
        if not self._op_index:
            if not self._op_catalog_full:
                self.op_adder.set_suggestions([])
                return
            normalized_query = normalize_text(query)
            if not normalized_query:
                subset = self._op_catalog_full[: self._op_search_limit]
            else:
                subset = []
                for item in self._op_catalog_full:
                    if normalized_query in normalize_text(item):
                        subset.append(item)
                    if len(subset) >= self._op_search_limit:
                        break
            self.op_adder.set_suggestions(subset)
            return

        future = self._search_executor.submit(self._op_index.search, query, self._op_search_limit)

        def _apply(fut):
            try:
                results = fut.result()
            except Exception:
                results = []
            if query != self._latest_op_query:
                return
            self.op_adder.set_suggestions(results)

        future.add_done_callback(lambda fut: QtCore.QTimer.singleShot(0, lambda: _apply(fut)))

    def _set_doctor_visibility(self, visible: bool):
        self.row_doctor_label.setVisible(visible);
        self.cb_doctor.setVisible(visible)

    def _collect(self):
        qd = self.date.date()
        dt = datetime(qd.year(), qd.month(), qd.day(), self.time.time().hour(), self.time.time().minute())
        auto_period = _now_period(dt)
        ward_text = self.cb_ward.currentText().strip()
        if ward_text == WARD_PLACEHOLDER:
            ward_text = ""
        return ScheduleEntry(
            or_room=self.cb_or.currentText().strip(), dt=dt.date(), time_str=self.time.time().toString("HH:mm"),
            hn=self.ent_hn.text().strip(), name=self.ent_name.text().strip(), age=self.ent_age.text().strip() or "0",
            dept=(self.cb_dept.currentText().strip() if not self.cb_dept.currentText().startswith("â€”") else ""),
            doctor=self.cb_doctor.currentText().strip() if self.cb_doctor.isVisible() else "",
            diags=self.diag_adder.items(), ops=self.op_adder.items(),
            ward=ward_text,
            case_size=self.cb_case.currentText().strip(),
            queue=0,
            period=auto_period,
            urgency=self.cb_urgency.currentText().strip() or "Elective",
            assist1=self.cb_assist1.currentText().strip(),
            assist2=self.cb_assist2.currentText().strip(),
            scrub=self.cb_scrub.currentText().strip(),
            circulate=self.cb_circulate.currentText().strip(),
            time_start=(self.time_start.time().toString("HH:mm") if self.ck_time_start.isChecked() else ""),
            time_end=(self.time_end.time().toString("HH:mm") if self.ck_time_end.isChecked() else ""),
        )

    def _clear_form(self):
        self.cb_or.setCurrentIndex(0)
        self.ent_name.clear();
        self.ent_age.clear();
        self.ent_hn.clear()
        self.cb_dept.setCurrentIndex(0);
        self.cb_doctor.clear();
        self._set_doctor_visibility(False)
        self.diag_adder.clear();
        self.op_adder.clear()
        self.cb_ward.setCurrentIndex(0);
        self.cb_ward.setEditText(WARD_PLACEHOLDER)
        if hasattr(self, "cb_case"):
            self.cb_case.setCurrentIndex(0)
        if hasattr(self, "cb_urgency"):
            idx = self.cb_urgency.findText("Elective")
            self.cb_urgency.setCurrentIndex(idx if idx >= 0 else 0)
        for cb in (self.cb_assist1, self.cb_assist2, self.cb_scrub, self.cb_circulate):
            cb.setCurrentIndex(0)
            cb.setEditText("")
        self.ck_time_start.setChecked(False);
        self.time_start.setEnabled(False);
        self.time_start.setTime(QtCore.QTime.currentTime())
        self.ck_time_end.setChecked(False);
        self.time_end.setEnabled(False);
        self.time_end.setTime(QtCore.QTime.currentTime())
        self.date.setDate(QtCore.QDate.currentDate())
        self.time.setTime(QtCore.QTime.currentTime())
        self._update_period_info()
        self._on_dept_changed(self.cb_dept.currentText())
        self._set_add_mode()

    # ---------- ADD / UPDATE ----------
    def _set_edit_mode(self, idx: int):
        self._edit_idx = idx
        self.btn_add.setText("ðŸ’¾ à¸šà¸±à¸™à¸—à¸¶à¸à¸à¸²à¸£à¹à¸à¹‰à¹„à¸‚")
        self.btn_cancel_edit.show()
        self.toast.show_toast("à¹€à¸‚à¹‰à¸²à¸ªà¸¹à¹ˆà¹‚à¸«à¸¡à¸”à¹à¸à¹‰à¹„à¸‚")

    def _set_add_mode(self):
        self._edit_idx = None
        self.btn_add.setText("âž• à¹€à¸žà¸´à¹ˆà¸¡")
        self.btn_cancel_edit.hide()

    def _cancel_edit_mode(self):
        self._set_add_mode()
        self.toast.show_toast("à¸¢à¸à¹€à¸¥à¸´à¸à¹‚à¸«à¸¡à¸”à¹à¸à¹‰à¹„à¸‚")

    def _load_form_from_entry(self, e: ScheduleEntry):
        idx = self.cb_or.findText(e.or_room)
        if idx >= 0: self.cb_or.setCurrentIndex(idx)
        self.ent_name.setText(e.name or "")
        self.ent_age.setText(str(e.age or 0))
        self.ent_hn.setText(e.hn or "")
        if hasattr(self, "cb_urgency"):
            idx_u = self.cb_urgency.findText(e.urgency or "Elective")
            self.cb_urgency.setCurrentIndex(idx_u if idx_u >= 0 else 0)
        try:
            d = QtCore.QDate(e.date.year, e.date.month, e.date.day)
            self.date.setDate(d)
        except Exception:
            pass
        try:
            hh, mm = (e.time or "00:00").split(":")
            self.time.setTime(QtCore.QTime(int(hh), int(mm)))
        except Exception:
            pass
        self._update_period_info()
        if e.dept:
            for i in range(self.cb_dept.count()):
                if self.cb_dept.itemText(i).startswith(e.dept) or self.cb_dept.itemText(i) == e.dept:
                    self.cb_dept.setCurrentIndex(i);
                    break
        if e.doctor and self.cb_doctor.isVisible():
            j = self.cb_doctor.findText(e.doctor)
            if j >= 0: self.cb_doctor.setCurrentIndex(j)
        self.diag_adder.clear();
        [self.diag_adder.list.addItem(x) for x in (e.diags or [])]
        self.op_adder.clear();
        [self.op_adder.list.addItem(x) for x in (e.ops or [])]
        # Ward
        j = self.cb_ward.findText(e.ward) if e.ward else -1
        if j >= 0:
            self.cb_ward.setCurrentIndex(j)
        else:
            if e.ward:
                self.cb_ward.setEditText(e.ward)
            else:
                self.cb_ward.setCurrentIndex(0)
                self.cb_ward.setEditText(WARD_PLACEHOLDER)

        # Case size
        if hasattr(self, "cb_case"):
            k = self.cb_case.findText(e.case_size) if e.case_size else -1
            if k >= 0:
                self.cb_case.setCurrentIndex(k)
            else:
                self.cb_case.setCurrentIndex(0)

        # Nurse roles
        for combo, value in (
                (self.cb_assist1, e.assist1),
                (self.cb_assist2, e.assist2),
                (self.cb_scrub, e.scrub),
                (self.cb_circulate, e.circulate),
        ):
            val = value or ""
            idx_val = combo.findText(val) if val else 0
            if val and idx_val >= 0:
                combo.setCurrentIndex(idx_val)
            else:
                combo.setCurrentIndex(0)
            combo.setEditText(val)

        # Start/End time (optional)
        if e.time_start:
            self.ck_time_start.setChecked(True)
            try:
                hh, mm = e.time_start.split(":")
                self.time_start.setTime(QtCore.QTime(int(hh), int(mm)))
            except Exception:
                pass
        else:
            self.ck_time_start.setChecked(False)
            self.time_start.setEnabled(False)
            self.time_start.setTime(QtCore.QTime.currentTime())

        if e.time_end:
            self.ck_time_end.setChecked(True)
            try:
                hh, mm = e.time_end.split(":")
                self.time_end.setTime(QtCore.QTime(int(hh), int(mm)))
            except Exception:
                pass
        else:
            self.ck_time_end.setChecked(False)
            self.time_end.setEnabled(False)
            self.time_end.setTime(QtCore.QTime.currentTime())

    def _on_add_or_update(self):
        e = self._collect()
        errs = []
        if not e.or_room: errs.append("à¸à¸£à¸¸à¸“à¸²à¹€à¸¥à¸·à¸­à¸ OR")
        if not e.name: errs.append("à¸à¸£à¸¸à¸“à¸²à¸à¸£à¸­à¸à¸Šà¸·à¹ˆà¸­-à¸ªà¸à¸¸à¸¥")
        if not e.hn: errs.append("à¸à¸£à¸¸à¸“à¸²à¸à¸£à¸­à¸ HN")
        if e.hn and not e.hn.isdigit(): errs.append("HN à¸•à¹‰à¸­à¸‡à¹€à¸›à¹‡à¸™à¸•à¸±à¸§à¹€à¸¥à¸‚à¹€à¸—à¹ˆà¸²à¸™à¸±à¹‰à¸™")
        if errs:
            self.lbl_warn.setText(" â€¢ ".join(errs))
            try:
                QtWidgets.QApplication.beep()
            except Exception:
                pass
            return
        else:
            self.lbl_warn.setText("")

        # à¸ˆà¸³ uid à¹„à¸§à¹‰à¹€à¸žà¸·à¹ˆà¸­à¹‚à¸Ÿà¸à¸±à¸ªà¸«à¸¥à¸±à¸‡à¸šà¸±à¸™à¸—à¸¶à¸
        self._last_focus_uid = e.uid()

        loader = SweetAlert.loading(self, "à¸à¸³à¸¥à¸±à¸‡à¸šà¸±à¸™à¸—à¸¶à¸à¸‚à¹‰à¸­à¸¡à¸¹à¸¥...")
        QtWidgets.QApplication.processEvents()
        try:
            if self._edit_idx is None:
                self.sched.add(e)
                try:
                    self.db_logger.append_entry(e)
                except Exception:
                    pass
                self._notify("à¹€à¸žà¸´à¹ˆà¸¡à¸£à¸²à¸¢à¸à¸²à¸£à¹à¸¥à¹‰à¸§", f"OR {e.or_room} â€¢ {e.time} â€¢ HN {e.hn}")
                SweetAlert.success(self, "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", "à¸šà¸±à¸™à¸—à¸¶à¸à¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸«à¸¡à¹ˆà¹€à¸£à¸µà¸¢à¸šà¸£à¹‰à¸­à¸¢")
                # à¹„à¸¡à¹ˆà¹€à¸žà¸´à¹ˆà¸¡à¹€à¸‚à¹‰à¸² historical_monitor_seen à¸—à¸µà¹ˆà¸™à¸µà¹ˆ â€” à¸›à¸¥à¹ˆà¸­à¸¢à¹ƒà¸«à¹‰ monitor à¸£à¸²à¸¢à¸‡à¸²à¸™ HN à¸ˆà¸°à¹€à¸›à¹‡à¸™à¸„à¸™à¹€à¸žà¸´à¹ˆà¸¡
            else:
                if 0 <= self._edit_idx < len(self.sched.entries):
                    old_entry = self.sched.entries[self._edit_idx]
                    e.case_uid = old_entry.case_uid
                    e.version = int(old_entry.version or 1) + 1
                    e.state = old_entry.state
                    e.returning_started_at = old_entry.returning_started_at
                    e.returned_to_ward_at = old_entry.returned_to_ward_at
                    e.postop_completed = old_entry.postop_completed
                self.sched.update(self._edit_idx, e)
                self._notify("à¸šà¸±à¸™à¸—à¸¶à¸à¸à¸²à¸£à¹à¸à¹‰à¹„à¸‚à¹à¸¥à¹‰à¸§", f"OR {e.or_room} â€¢ {e.time} â€¢ HN {e.hn}")
                SweetAlert.success(self, "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", "à¸­à¸±à¸›à¹€à¸”à¸•à¸£à¸²à¸¢à¸à¸²à¸£à¹€à¸£à¸µà¸¢à¸šà¸£à¹‰à¸­à¸¢")
                self._set_add_mode()
        finally:
            loader.close()

        self._set_result_title()
        self._render_tree2()

        # à¹€à¸”à¹‰à¸‡à¹„à¸›à¹à¸—à¹‡à¸š Result à¹à¸¥à¸°à¹‚à¸Ÿà¸à¸±à¸ª/à¹„à¸®à¹„à¸¥à¸•à¹Œà¸Šà¸·à¹ˆà¸­à¸œà¸¹à¹‰à¸›à¹ˆà¸§à¸¢
        self.tabs.setCurrentIndex(1)
        QtCore.QTimer.singleShot(120, lambda: self._focus_uid(self._last_focus_uid))

        self._clear_form()

    # ---------- Helpers for monitor state sync ----------
    def _find_entry_by_hn_latest(self, hn: str):
        matches = [e for e in self.sched.entries if str(e.hn).strip() == str(hn).strip()]
        if not matches:
            return None

        def _key(entry: ScheduleEntry):
            try:
                hh, mm = (entry.time or "00:00").split(":")
                return (str(entry.date), int(hh) * 60 + int(mm))
            except Exception:
                return (str(entry.date), 0)

        matches.sort(key=_key, reverse=True)
        return matches[0]

    def _set_time_start_if_empty(self, entry: ScheduleEntry):
        if not entry.time_start:
            entry.time_start = datetime.now().strftime("%H:%M")
            entry.version = int(entry.version or 1) + 1

    def _set_time_end_if_empty(self, entry: ScheduleEntry):
        if not entry.time_end:
            entry.time_end = datetime.now().strftime("%H:%M")
            entry.version = int(entry.version or 1) + 1

    def _scan_monitor_status_transitions(self, rows: List[dict]):
        for row in rows:
            hn = str(row.get("hn_full") or row.get("id") or "").strip()
            if not hn:
                continue
            status = str(row.get("status") or "").strip()
            if not status:
                continue

            prev = self._last_status_by_hn.get(hn)
            if prev == status:
                continue
            self._last_status_by_hn[hn] = status

            entry = self._find_entry_by_hn_latest(hn)
            if not entry:
                continue

            changed = False
            if status == STATUS_OP_START:
                self._set_time_start_if_empty(entry)
                if entry.state in {"scheduled", "in_or", "operation_ended", "postop_pending", ""}:
                    entry.state = "operation_started"
                    changed = True
            elif status == STATUS_OP_END:
                self._set_time_end_if_empty(entry)
                if entry.state in {"operation_started", "in_or", "scheduled", ""}:
                    entry.state = "operation_ended"
                    changed = True
            elif status == STATUS_RETURNING:
                if not entry.time_end:
                    continue
                if entry.state != "returning_to_ward":
                    entry.state = "returning_to_ward"
                    entry.returning_started_at = _now_iso()
                    entry.version = int(entry.version or 1) + 1
                    changed = True

            if changed:
                self.sched._save()

    def _is_entry_completed(self, e: ScheduleEntry) -> bool:
        """à¸•à¸£à¸§à¸ˆà¸§à¹ˆà¸²à¸£à¸²à¸¢à¸à¸²à¸£à¸–à¸¹à¸à¹€à¸•à¸´à¸¡à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸«à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¸„à¸£à¸šà¸–à¹‰à¸§à¸™à¸žà¸­à¸ªà¸³à¸«à¸£à¸±à¸šà¸à¸²à¸£à¸›à¸´à¸”à¹€à¸„à¸ª"""
        return _is_postop_complete_entry(e)

    def _render_tree2(self):
        vbar = self.tree2.verticalScrollBar()
        hbar = self.tree2.horizontalScrollBar()
        vpos = vbar.value()
        hpos = hbar.value()

        expanded_state: Dict[str, bool] = {}
        root = self.tree2.invisibleRootItem()
        for i in range(root.childCount()):
            header_item = root.child(i)
            key = header_item.data(0, QtCore.Qt.UserRole)
            if not key:
                key = header_item.text(0).split('â€¢', 1)[0].strip()
            expanded_state[str(key)] = header_item.isExpanded()

        self.tree2.setUpdatesEnabled(False)
        self.tree2.blockSignals(True)
        try:
            self.tree2.clear()
            self._set_result_title()

            base_date = datetime.now().date()
            try:
                if hasattr(self, 'date'):
                    qdate = self.date.date()
                    if hasattr(qdate, 'toPython'):
                        base_date = qdate.toPython()
                    else:
                        base_date = date(qdate.year(), qdate.month(), qdate.day())
            except Exception:
                base_date = datetime.now().date()

            entries_snapshot: List[ScheduleEntry] = list(self.sched.entries)
            entries_snapshot = normalize_owner_for_wednesday(entries_snapshot, base_date)

            def _resolved_date(entry: ScheduleEntry) -> date:
                day_val = getattr(entry, "date", None)
                if isinstance(day_val, datetime):
                    return day_val.date()
                if isinstance(day_val, date):
                    return day_val
                if hasattr(day_val, "toPython"):
                    try:
                        return day_val.toPython()
                    except Exception:
                        return base_date
                if isinstance(day_val, str):
                    try:
                        return datetime.fromisoformat(day_val).date()
                    except Exception:
                        return base_date
                return base_date

            entries_for_day = [entry for entry in entries_snapshot if _resolved_date(entry) == base_date]

            valid_pickups: Set[str] = set()
            for entry in entries_for_day:
                pid = self._pickup_id_for_entry(entry)
                if pid:
                    valid_pickups.add(pid)
            if valid_pickups:
                self._runner_finished_sent.intersection_update(valid_pickups)
            else:
                self._runner_finished_sent.clear()

            runner_status_map: Dict[str, dict] = {}
            runner_ready = False
            if entries_for_day:
                base = self._runner_base()
                runner_ready = runner_health_ok(base)
                if runner_ready:
                    self._push_rows_to_runner(entries_for_day, runner_ready=True)
                    runner_status_map = _fetch_runner_status_map(str(base_date), base)
                    self._auto_finish_runner_cases(entries_for_day, runner_status_map)
            self._runner_status_cache = runner_status_map

            indexed_entries: List[Tuple[int, ScheduleEntry]] = list(enumerate(entries_snapshot))
            if not indexed_entries:
                empty = QtWidgets.QTreeWidgetItem(['â€” à¹„à¸¡à¹ˆà¸¡à¸µà¸£à¸²à¸¢à¸à¸²à¸£ â€”'])
                _span_first_column(empty)
                self.tree2.addTopLevelItem(empty)
            else:
                groups: Dict[str, List[Tuple[int, ScheduleEntry]]] = {}

                def _effective_or_room(entry: ScheduleEntry) -> str:
                    day_val = getattr(entry, 'date', None)
                    resolved_date = base_date
                    if isinstance(day_val, datetime):
                        resolved_date = day_val.date()
                    elif isinstance(day_val, date):
                        resolved_date = day_val
                    elif hasattr(day_val, 'toPython'):
                        try:
                            resolved_date = day_val.toPython()
                        except Exception:
                            resolved_date = base_date
                    if resolved_date and isinstance(resolved_date, date) and resolved_date.weekday() == 2:
                        who = _infer_doctor_from_entry(entry)
                        if who:
                            normalized_who = normalize_doctor_name(who)
                            for owner_name, target_room in OWNER_WED_DOCTOR2OR.items():
                                if normalize_doctor_name(owner_name) == normalized_who:
                                    return target_room
                    return entry.or_room or '-'

                for idx, entry in indexed_entries:
                    effective_room = _effective_or_room(entry) or '-'
                    if effective_room not in ('', '-') and effective_room != (entry.or_room or ''):
                        entry.or_room = effective_room
                    bucket_key = effective_room or '-'
                    groups.setdefault(bucket_key, []).append((idx, entry))

                order = list(getattr(self.sched, 'or_rooms', []))

                def _room_sort_key(or_name: str) -> Tuple[int, int]:
                    if or_name in order:
                        return (0, order.index(or_name))
                    if str(or_name).strip() == '-':
                        return (2, 999)
                    digits = ''.join(ch for ch in str(or_name) if ch.isdigit())
                    num = int(digits) if digits else 999
                    return (1, num)

                def _time_tuple(hhmm: str) -> Tuple[int, int, int]:
                    if not hhmm or hhmm == 'TF':
                        return (1, 99, 99)
                    try:
                        hh, mm = [int(x) for x in hhmm.split(':')]
                        return (0, hh, mm)
                    except Exception:
                        return (1, 99, 99)

                def _queue_value(entry: ScheduleEntry) -> int:
                    try:
                        return int(entry.queue or 0)
                    except Exception:
                        return 0

                state_colors = {
                    'returning_to_ward': '#ede9fe',
                    'postop_pending': '#fff7ed',
                    'returned_to_ward': '#ecfdf5',
                }
                state_icons = {
                    'returning_to_ward': 'â³',
                    'postop_pending': 'âš ï¸',
                    'returned_to_ward': 'âœ…',
                }

                headers: List[Tuple[str, QtWidgets.QTreeWidgetItem]] = []

                for or_room in sorted(groups.keys(), key=_room_sort_key):
                    bucket = groups[or_room]

                    def _row_sort(pair: Tuple[int, ScheduleEntry]) -> Tuple[int, int, int, int, int, str]:
                        entry = pair[1]
                        q = _queue_value(entry)
                        flag, hh, mm = _time_tuple(entry.time)
                        return (
                            0 if q > 0 else 1,
                            q if q > 0 else 0,
                            flag,
                            hh,
                            mm,
                            str(entry.hn or ''),
                        )

                    bucket_sorted = sorted(bucket, key=_row_sort)
                    entries_only = [entry for _, entry in bucket_sorted]

                    header_item = QtWidgets.QTreeWidgetItem(['' for _ in range(self.tree2.columnCount())])
                    or_label = or_room or '-'
                    first_entry = entries_only[0] if entries_only else None
                    the_date = getattr(first_entry, 'date', base_date)
                    owner = resolve_or_owner(or_label, the_date, getattr(first_entry, 'doctor', None)) or '-'
                    header_item.setText(0, f"{or_label} â€¢ {owner}")
                    font = header_item.font(0)
                    font.setBold(True)
                    header_item.setFont(0, font)
                    _span_first_column(header_item)
                    header_item.setChildIndicatorPolicy(QtWidgets.QTreeWidgetItem.ShowIndicator)
                    header_item.setData(0, QtCore.Qt.UserRole, or_label)
                    self.tree2.addTopLevelItem(header_item)
                    headers.append((or_label, header_item))

                    for idx, entry in bucket_sorted:
                        diag_txt = ' ; '.join(entry.diags) if entry.diags else '-'
                        op_txt = ' ; '.join(entry.ops) if entry.ops else '-'
                        or_time = f"{or_label} â€¢ {entry.time or 'TF'}"
                        status_text = getattr(entry, 'status', '') or (entry.state or '') or '-'
                        case_size_txt = getattr(entry, 'case_size', '') or '-'
                        dept_txt = getattr(entry, 'dept', '') or '-'
                        row = QtWidgets.QTreeWidgetItem([
                            or_time,
                            entry.hn or '-',
                            entry.name or '-',
                            str(entry.age or 0),
                            diag_txt,
                            op_txt,
                            entry.doctor or '-',
                            entry.ward or '-',
                            case_size_txt,
                            dept_txt,
                            entry.time_start or '-',
                            entry.time_end or '-',
                            '',
                            entry.assist1 or '-',
                            entry.assist2 or '-',
                            entry.scrub or '-',
                            getattr(entry, 'circulate', '') or '-',
                            status_text,
                        ])
                        row.setData(0, QtCore.Qt.UserRole, entry.uid())
                        row.setData(0, QtCore.Qt.UserRole + 1, idx)
                        pickup_id = self._pickup_id_for_entry(entry, or_label)
                        row.setData(0, QtCore.Qt.UserRole + 2, pickup_id)
                        header_item.addChild(row)

                        badge = _period_badge(entry.period or 'in')
                        self.tree2.setItemWidget(row, 12, badge)

                        runner_info = runner_status_map.get(pickup_id, {})
                        runner_status = (runner_info or {}).get('status', '')
                        runner_label = self._runner_status_label(runner_status)
                        if runner_label:
                            chip_color = RUNNER_STATUS_COLORS.get(runner_status, '#64748b')
                            runner_chip = StatusChipWidget(runner_label, chip_color)
                            self.tree2.setItemWidget(row, 17, runner_chip)
                            row.setText(17, '')
                            tooltip = self._runner_status_tooltip(runner_info)
                            if tooltip:
                                row.setToolTip(17, tooltip)
                            name_txt = entry.name or '-'
                            row.setText(2, f"[{runner_label}] {name_txt}")
                        else:
                            self.tree2.setItemWidget(row, 17, None)
                            row.setText(17, status_text)
                            row.setToolTip(17, '')
                            row.setText(2, entry.name or '-')

                        monitor_status = self._last_status_by_hn.get(str(entry.hn).strip(), '')
                        if monitor_status:
                            color = STATUS_COLORS.get(monitor_status, '#64748b')
                            chip = StatusChipWidget(monitor_status, color, pulse=(monitor_status in PULSE_STATUS))
                            cell = QtWidgets.QWidget()
                            lay = QtWidgets.QHBoxLayout(cell)
                            lay.setContentsMargins(0, 0, 0, 0)
                            lay.setSpacing(6)
                            lay.addWidget(chip, 0)
                            lbl = QtWidgets.QLabel(or_time)
                            lay.addWidget(lbl, 0)
                            lay.addStretch(1)
                            self.tree2.setItemWidget(row, 0, cell)
                        else:
                            row.setText(0, or_time)
                            self.tree2.setItemWidget(row, 0, None)

                        state = entry.state or ''
                        if state in state_colors:
                            brush = QtGui.QBrush(QtGui.QColor(state_colors[state]))
                            for col_idx in range(self.tree2.columnCount()):
                                row.setBackground(col_idx, brush)
                            icon = state_icons.get(state)
                            if icon:
                                row.setText(2, f"{icon} {row.text(2)}")
                        if state:
                            tip = [f"State: {state}"]
                            if entry.returning_started_at:
                                tip.append(f"à¹€à¸£à¸´à¹ˆà¸¡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸: {entry.returning_started_at}")
                            if entry.returned_to_ward_at:
                                tip.append(f"à¸à¸¥à¸±à¸šà¸•à¸¶à¸à¹€à¸¡à¸·à¹ˆà¸­: {entry.returned_to_ward_at}")
                            if entry.postop_completed:
                                tip.append('(à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸«à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¸„à¸£à¸šà¸–à¹‰à¸§à¸™ âœ“)')
                            row.setToolTip(2, '\n'.join(tip))

                for or_label, header_item in headers:
                    header_item.setExpanded(expanded_state.get(or_label, True))
        finally:
            self.tree2.blockSignals(False)
            self.tree2.setUpdatesEnabled(True)
            self.tree2.clearSelection()
            self.tree2.setCurrentItem(None)

            def _restore_scroll():
                vbar.setValue(min(vpos, vbar.maximum()))
                hbar.setValue(min(hpos, hbar.maximum()))

            QtCore.QTimer.singleShot(0, _restore_scroll)

    def _apply_queue_select(self, uid: str, new_q: int):
        target = None;
        target_idx = None
        for i, entry in enumerate(self.sched.entries):
            if entry.uid() == uid:
                target = entry;
                target_idx = i;
                break
        if not target: return
        new_q = max(0, min(9, int(new_q)))
        if new_q == target.queue: return
        target.queue = int(new_q)
        self.sched._save()
        try:
            QtWidgets.QApplication.beep()
        except Exception:
            pass
        self._notify("à¸­à¸±à¸›à¹€à¸”à¸•à¸„à¸´à¸§à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", f"OR {target.or_room} â€¢ HN {target.hn} â†’ à¸„à¸´à¸§ {new_q or 'à¸•à¸²à¸¡à¹€à¸§à¸¥à¸²'}")
        self._set_result_title()
        self._render_tree2()
        self._flash_row_by_uid(uid)

    def _find_item_by_uid(self, uid: str):
        root = self.tree2.invisibleRootItem()
        for i in range(root.childCount()):
            parent = root.child(i)
            for j in range(parent.childCount()):
                ch = parent.child(j)
                if ch.data(0, QtCore.Qt.UserRole) == uid:
                    return ch
        return None

    def _flash_row_by_uid(self, uid: str):
        it = self._find_item_by_uid(uid)
        if not it: return
        rect = self.tree2.visualItemRect(it)
        overlay = QtWidgets.QFrame(self.tree2.viewport())
        overlay.setGeometry(0, rect.y(), self.tree2.viewport().width(), rect.height())
        overlay.setStyleSheet("QFrame{background:rgba(250,204,21,0.35);border-radius:4px;}")
        overlay.raise_();
        overlay.show()
        anim = QtCore.QPropertyAnimation(overlay, b"windowOpacity", self)
        anim.setDuration(900);
        anim.setStartValue(1.0);
        anim.setKeyValueAt(0.5, 0.0);
        anim.setEndValue(1.0);
        anim.setLoopCount(2)
        anim.finished.connect(overlay.deleteLater)
        anim.start(QtCore.QAbstractAnimation.DeleteWhenStopped)

    def _focus_uid(self, uid: str):
        if not uid: return
        it = self._find_item_by_uid(uid)
        if not it: return
        hbar = self.tree2.horizontalScrollBar()
        old_hval = hbar.value()

        rect = self.tree2.visualItemRect(it)
        if not rect.isValid():
            index = self.tree2.indexFromItem(it)
            if index.isValid():
                self.tree2.scrollTo(index, QtWidgets.QAbstractItemView.PositionAtCenter)
                rect = self.tree2.visualItemRect(it)

        if rect.isValid():
            vbar = self.tree2.verticalScrollBar()
            target = rect.y() + vbar.value() - (self.tree2.viewport().height() // 2)
            vbar.setValue(max(0, min(target, vbar.maximum())))

        self.tree2.setCurrentItem(it)
        self._flash_row_by_uid(uid)
        QtCore.QTimer.singleShot(0, lambda: hbar.setValue(min(old_hval, hbar.maximum())))

    # ---------- Result context menu / Double-click ----------
    def _result_ctx_menu(self, pos: QtCore.QPoint):
        it = self.tree2.itemAt(pos)
        if not it: return
        idx = it.data(0, QtCore.Qt.UserRole + 1)
        if idx is None: return
        idx_int = int(idx)
        entry = self.sched.entries[idx_int] if 0 <= idx_int < len(self.sched.entries) else None
        menu = QtWidgets.QMenu(self)
        a_edit = menu.addAction("à¹à¸à¹‰à¹„à¸‚à¸£à¸²à¸¢à¸à¸²à¸£")
        a_del = menu.addAction("à¸¥à¸šà¸£à¸²à¸¢à¸à¸²à¸£")
        runner_ack_action = runner_arrive_action = None
        if entry:
            pickup_id = it.data(0, QtCore.Qt.UserRole + 2)
            if pickup_id:
                menu.addSeparator()
                runner_ack_action = menu.addAction("ðŸ“¥ Runner: à¸£à¸±à¸šà¹€à¸„à¸ª")
                runner_arrive_action = menu.addAction("âœ… Runner: à¸–à¸¶à¸‡ OR")
        act = menu.exec(self.tree2.viewport().mapToGlobal(pos))
        if act == a_edit:
            self._on_result_double_click(it, 0)
        elif act == a_del:
            self._delete_entry_idx(idx_int)
        elif act and entry and runner_ack_action and act == runner_ack_action:
            self._handle_runner_action(entry, "ack")
        elif act and entry and runner_arrive_action and act == runner_arrive_action:
            self._handle_runner_action(entry, "arrive")

    def _delete_entry_idx(self, idx: int):
        if 0 <= idx < len(self.sched.entries):
            entry = self.sched.entries[idx]
            if not SweetAlert.confirm(
                    self,
                    "à¸¢à¸·à¸™à¸¢à¸±à¸™à¸à¸²à¸£à¸¥à¸š",
                    f"à¸•à¹‰à¸­à¸‡à¸à¸²à¸£à¸¥à¸š HN {entry.hn} ({entry.name}) à¸«à¸£à¸·à¸­à¹„à¸¡à¹ˆ?",
            ):
                return
            self.sched.delete(idx)
            self._render_tree2()
            SweetAlert.success(self, "à¸ªà¸³à¹€à¸£à¹‡à¸ˆ", "à¸¥à¸šà¸£à¸²à¸¢à¸à¸²à¸£à¹€à¸£à¸µà¸¢à¸šà¸£à¹‰à¸­à¸¢")

    def _on_monitor_double_click(self, item: QtWidgets.QTableWidgetItem):
        row = item.row()
        hn = self.table.item(row, 0).text().strip() if self.table.item(row, 0) else ""
        if not hn:
            self.toast.show_toast("à¹„à¸¡à¹ˆà¸žà¸š HN à¸‚à¸­à¸‡à¹à¸–à¸§à¸™à¸µà¹‰");
            return
        self._route_to_identify(hn)

    def _on_result_double_click(self, item: QtWidgets.QTreeWidgetItem, col: int):
        idx = item.data(0, QtCore.Qt.UserRole + 1)
        if idx is None: return
        if 0 <= int(idx) < len(self.sched.entries):
            entry = self.sched.entries[int(idx)]
            self._load_form_from_entry(entry)
            self._set_edit_mode(int(idx))
            self.tabs.setCurrentIndex(0)  # à¹„à¸›à¸—à¸µà¹ˆà¸Ÿà¸­à¸£à¹Œà¸¡

    # ---------- Identify routing ----------
    def _find_entry_index_by_hn(self, hn: str) -> Optional[int]:
        for i, entry in enumerate(self.sched.entries):
            if str(entry.hn).strip() == str(hn).strip():
                return i
        return None

    def _route_to_identify(self, hn: str):
        QtWidgets.QApplication.clipboard().setText(hn)
        idx = self._find_entry_index_by_hn(hn)
        if idx is not None:
            entry = self.sched.entries[idx]
            self._load_form_from_entry(entry)
            self._set_edit_mode(idx)
            self.tabs.setCurrentIndex(0)
            self.toast.show_toast(f"HN {hn}: à¸žà¸šà¹ƒà¸™à¸£à¸²à¸¢à¸à¸²à¸£ â†’ à¹‚à¸«à¸¡à¸”à¹à¸à¹‰à¹„à¸‚")
        else:
            self._clear_form()
            self.ent_hn.setText(hn)
            self._set_add_mode()
            self.tabs.setCurrentIndex(0)
            self.toast.show_toast(f"HN {hn}: à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µ â†’ à¹€à¸žà¸´à¹ˆà¸¡à¹ƒà¸«à¸¡à¹ˆ")

    def apply_external_update(self, uid: str, patch: dict) -> bool:
        """à¸£à¸±à¸šà¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ˆà¸²à¸ client à¸ à¸²à¸¢à¸™à¸­à¸à¹€à¸žà¸·à¹ˆà¸­à¹€à¸•à¸´à¸¡à¸£à¸²à¸¢à¸¥à¸°à¹€à¸­à¸µà¸¢à¸”à¸«à¸¥à¸±à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”"""
        accepted_keys = {
            "assist1",
            "assist2",
            "scrub",
            "circulate",
            "time_start",
            "time_end",
            "ward",
            "doctor",
        }
        accepted_keys |= {
            "state",
            "returning_started_at",
            "returned_to_ward_at",
            "postop_completed",
            "version",
        }

        intent = str(patch.get("_intent") or "").strip().lower()

        for entry in self.sched.entries:
            if entry.uid() == uid:
                if intent == "mark_returning":
                    if not entry.time_end:
                        self.toast.show_toast("à¸¢à¸±à¸‡à¹„à¸¡à¹ˆà¸¡à¸µà¹€à¸§à¸¥à¸² 'à¸ˆà¸šà¸œà¹ˆà¸²à¸•à¸±à¸”' â€” à¸•à¸±à¹‰à¸‡à¸ªà¸–à¸²à¸™à¸°à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸à¹„à¸¡à¹ˆà¹„à¸”à¹‰")
                        return False
                    entry.state = "returning_to_ward"
                    entry.returning_started_at = _now_iso()
                    entry.postop_completed = False
                    entry.returned_to_ward_at = ""
                    entry.version = int(entry.version or 1) + 1
                    self.sched._save()
                    self._render_tree2()
                    self._flash_row_by_uid(uid)
                    self.toast.show_toast("à¸•à¸±à¹‰à¸‡à¸ªà¸–à¸²à¸™à¸° 'à¸à¸³à¸¥à¸±à¸‡à¸ªà¹ˆà¸‡à¸à¸¥à¸±à¸šà¸•à¸¶à¸' à¹à¸¥à¹‰à¸§ (à¹€à¸£à¸´à¹ˆà¸¡à¸™à¸±à¸š 3 à¸™à¸²à¸—à¸µ)")
                    return True

                string_fields = {
                    "assist1",
                    "assist2",
                    "scrub",
                    "circulate",
                    "time_start",
                    "time_end",
                    "ward",
                    "doctor",
                    "state",
                    "returning_started_at",
                    "returned_to_ward_at",
                }

                for key in accepted_keys:
                    if key not in patch:
                        continue
                    value = patch.get(key)
                    if key == "version":
                        # version à¸ˆà¸°à¸–à¸¹à¸à¸›à¸£à¸±à¸šà¹€à¸žà¸´à¹ˆà¸¡à¸—à¹‰à¸²à¸¢à¸Ÿà¸±à¸‡à¸à¹Œà¸Šà¸±à¸™
                        continue
                    if key == "postop_completed":
                        entry.postop_completed = bool(value)
                        continue
                    if key in string_fields:
                        setattr(entry, key, str(value or ""))
                        continue
                    setattr(entry, key, value)

                entry.version = int(entry.version or 1) + 1
                if entry.state == "returning_to_ward" and not entry.returning_started_at:
                    entry.returning_started_at = _now_iso()
                self.sched._save()
                self._render_tree2()
                self._flash_row_by_uid(uid)
                self.toast.show_toast("à¸­à¸±à¸›à¹€à¸”à¸•à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ˆà¸²à¸ Client à¸ªà¸³à¹€à¸£à¹‡à¸ˆ")
                return True
        return False

    # ---------- export ----------
    def _export_csv(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export CSV", "monitor.csv", "CSV (*.csv)")
        if not path: return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f);
                w.writerow(["ID", "Patient ID", "Status", "Timestamp", "ETA(min)"])
                for r in self.rows_cache: w.writerow(
                    [r.get("id", ""), r.get("patient_id", ""), r.get("status", ""), r.get("timestamp", ""),
                     r.get("eta_minutes", "")])
            QtWidgets.QMessageBox.information(self, "à¸ªà¹ˆà¸‡à¸­à¸­à¸à¹à¸¥à¹‰à¸§", path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "à¸œà¸´à¸”à¸žà¸¥à¸²à¸”", str(e))

    def _export_deid_csv(self):
        """
        à¸ªà¹ˆà¸‡à¸­à¸­à¸à¸‚à¹‰à¸­à¸¡à¸¹à¸¥à¸ªà¸³à¸«à¸£à¸±à¸šà¸§à¸´à¹€à¸„à¸£à¸²à¸°à¸«à¹Œà¹à¸šà¸šà¹„à¸¡à¹ˆà¸£à¸°à¸šà¸¸à¸•à¸±à¸§à¸•à¸™ (de-identified)
        à¹à¸«à¸¥à¹ˆà¸‡à¸‚à¹‰à¸­à¸¡à¸¹à¸¥: self.sched.entries (à¸•à¸²à¸£à¸²à¸‡ Result Schedule à¸ à¸²à¸¢à¹ƒà¸™à¹€à¸„à¸£à¸·à¹ˆà¸­à¸‡)
        à¸Ÿà¸´à¸¥à¸”à¹Œà¸ªà¸³à¸„à¸±à¸: hn_hash, dept, or, queue, period, scheduled date/time, time_start, time_end, diags, ops, ward
        """
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export De-Identified CSV", "cases_deid.csv",
                                                        "CSV (*.csv)")
        if not path: return
        try:
            rows = []
            for e in self.sched.entries:
                rows.append({
                    "hn_hash": hn_hash(e.hn or ""),
                    "dept": e.dept or "",
                    "or": e.or_room or "",
                    "queue": int(e.queue or 0),
                    "period": e.period or "",
                    "scheduled_date": str(e.date or ""),
                    "scheduled_time": e.time or "",
                    "time_start": e.time_start or "",
                    "time_end": e.time_end or "",
                    "diag": " | ".join(e.diags or []),
                    "op": " | ".join(e.ops or []),
                    "ward": e.ward or "",
                    "case_size": e.case_size or "",
                    "urgency": e.urgency or "",
                    "doctor": e.doctor or "",
                    # à¸«à¸¡à¸²à¸¢à¹€à¸«à¸•à¸¸: à¹„à¸¡à¹ˆà¸ªà¹ˆà¸‡à¸­à¸­à¸ HN/à¸Šà¸·à¹ˆà¸­
                })
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                cols = ["hn_hash", "dept", "or", "queue", "period", "scheduled_date", "scheduled_time", "time_start",
                        "time_end", "diag", "op", "ward", "case_size", "urgency", "doctor"]
                w = csv.DictWriter(f, fieldnames=cols)
                w.writeheader();
                w.writerows(rows)
            QtWidgets.QMessageBox.information(self, "à¸ªà¹ˆà¸‡à¸­à¸­à¸à¹à¸¥à¹‰à¸§", path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "à¸œà¸´à¸”à¸žà¸¥à¸²à¸”", str(e))

    # ---------- notify ----------
    def _notify(self, title: str, msg: str):
        try:
            self.tray.showMessage(title, msg, QtWidgets.QSystemTrayIcon.Information, 3000)
        except Exception:
            pass

    def _set_result_title(self):
        now = datetime.now()
        txt = f"à¸•à¸²à¸£à¸²à¸‡à¸à¸²à¸£à¸œà¹ˆà¸²à¸•à¸±à¸” à¸›à¸£à¸°à¸ˆà¸³à¸§à¸±à¸™ ({now:%d/%m/%Y}) à¹€à¸§à¸¥à¸² {now:%H:%M} à¸™. à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¹‚à¸£à¸‡à¸žà¸¢à¸²à¸šà¸²à¸¥à¸«à¸™à¸­à¸‡à¸šà¸±à¸§à¸¥à¸³à¸ à¸¹"
        self.result_banner.set_icon("ðŸ“")
        self.result_banner.set_title(txt)
        self.result_banner.set_subtitle("à¸«à¹‰à¸­à¸‡à¸œà¹ˆà¸²à¸•à¸±à¸”à¹‚à¸£à¸‡à¸žà¸¢à¸²à¸šà¸²à¸¥à¸«à¸™à¸­à¸‡à¸šà¸±à¸§à¸¥à¸³à¸ à¸¹")
        self.card_result.title_lbl.setText(txt)

    # ---------- seq watcher ----------
    def _check_seq(self):
        cur = self.sched.seq()
        if cur != self.seq_seen:
            self.seq_seen = cur
            self.sched.entries = self.sched._load()
            self.sched.or_rooms = self.sched._load_or()
            self._refresh_or_cb(self.cb_or)
            self._render_tree2()


class WrapItemDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter, option, index):
        text = index.data(QtCore.Qt.DisplayRole)
        opt = QtWidgets.QStyleOptionViewItem(option);
        self.initStyleOption(opt, index);
        opt.text = ""
        style = opt.widget.style() if opt.widget else QtWidgets.QApplication.style()
        style.drawControl(QtWidgets.QStyle.CE_ItemViewItem, opt, painter, opt.widget)
        rect = style.subElementRect(QtWidgets.QStyle.SE_ItemViewItemText, opt, opt.widget)
        doc = QtGui.QTextDocument();
        doc.setDefaultFont(opt.font)
        topt = QtGui.QTextOption();
        topt.setWrapMode(QtGui.QTextOption.WordWrap);
        doc.setDefaultTextOption(topt)
        doc.setTextWidth(rect.width());
        doc.setPlainText(str(text) if text is not None else "")
        painter.save();
        painter.translate(rect.topLeft());
        doc.drawContents(painter, QtCore.QRectF(0, 0, rect.width(), rect.height()));
        painter.restore()

    def sizeHint(self, option, index):
        text = index.data(QtCore.Qt.DisplayRole) or ""
        doc = QtGui.QTextDocument();
        doc.setDefaultFont(option.font)
        topt = QtGui.QTextOption();
        topt.setWrapMode(QtGui.QTextOption.WordWrap);
        doc.setDefaultTextOption(topt)
        # à¹ƒà¸Šà¹‰à¸„à¸§à¸²à¸¡à¸à¸§à¹‰à¸²à¸‡à¸„à¸­à¸¥à¸±à¸¡à¸™à¹Œà¸ˆà¸£à¸´à¸‡à¸‚à¸­à¸‡ tree à¹€à¸žà¸·à¹ˆà¸­à¸¥à¸”à¸›à¸±à¸à¸«à¸²à¸„à¸§à¸²à¸¡à¸ªà¸¹à¸‡à¸›à¸£à¸°à¹€à¸¡à¸´à¸™à¸•à¹ˆà¸³
        tree = option.widget if isinstance(option.widget, QtWidgets.QTreeWidget) else None
        col_w = tree.columnWidth(index.column()) if tree else option.rect.width()
        # à¹€à¸œà¸·à¹ˆà¸­à¸£à¸°à¸¢à¸°à¸‚à¸­à¸šà¸™à¸´à¸”à¸«à¸™à¹ˆà¸­à¸¢
        w = max(120, int(col_w) - 12)
        doc.setTextWidth(w)
        doc.setPlainText(str(text))
        s = doc.size()
        return QtCore.QSize(w, int(s.height()) + 12)


class SearchSelectAdder(QtWidgets.QWidget):
    """Searchable selector with a multi-select list.

    - Enter / à¸›à¸¸à¹ˆà¸¡ "âž• à¹€à¸žà¸´à¹ˆà¸¡"  : à¹€à¸žà¸´à¹ˆà¸¡à¸¥à¸‡à¸£à¸²à¸¢à¸à¸²à¸£à¸‚à¸­à¸‡à¹€à¸„à¸ª (à¹„à¸¡à¹ˆà¹à¸•à¸°à¸„à¸¥à¸±à¸‡à¸«à¸¥à¸±à¸)
    - à¸›à¸¸à¹ˆà¸¡ "ðŸ’¾ à¸šà¸±à¸™à¸—à¸¶à¸à¹€à¸›à¹‡à¸™à¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸«à¸¡à¹ˆ" : à¸ªà¹ˆà¸‡à¸ªà¸±à¸à¸à¸²à¸“à¹ƒà¸«à¹‰à¸ à¸²à¸¢à¸™à¸­à¸à¸šà¸±à¸™à¸—à¸¶à¸à¹€à¸‚à¹‰à¸²à¸„à¸¥à¸±à¸‡à¸«à¸¥à¸±à¸
    """

    itemsChanged = QtCore.Signal(list)
    requestPersist = QtCore.Signal(str)

    def __init__(self, placeholder="à¸„à¹‰à¸™à¸«à¸² ICD-10...", suggestions=None, parent=None):
        super().__init__(parent)
        v = QtWidgets.QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)

        row = QtWidgets.QHBoxLayout()
        row.setSpacing(6)
        self.combo = NoWheelComboBox()  # à¸à¸±à¸™ scroll à¹€à¸›à¸¥à¸µà¹ˆà¸¢à¸™à¸„à¹ˆà¸²à¹‚à¸”à¸¢à¹„à¸¡à¹ˆà¸•à¸±à¹‰à¸‡à¹ƒà¸ˆ
        self.combo.setEditable(True)
        self.combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo.setMinimumWidth(280)
        self.combo.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.search_line = self.combo.lineEdit()
        if self.search_line:
            self.search_line.setPlaceholderText(placeholder)

        self.btn_add = QtWidgets.QPushButton("âž• à¹€à¸žà¸´à¹ˆà¸¡")
        self.btn_add.setProperty("variant", "ghost")
        self.btn_persist = QtWidgets.QPushButton("ðŸ’¾ à¸šà¸±à¸™à¸—à¸¶à¸à¹€à¸›à¹‡à¸™à¸£à¸²à¸¢à¸à¸²à¸£à¹ƒà¸«à¸¡à¹ˆ")
        self.btn_persist.setProperty("variant", "ghost")

        row.addWidget(self.combo, 1)
        row.addWidget(self.btn_add)
        row.addWidget(self.btn_persist)
        v.addLayout(row)

        self.list = QtWidgets.QListWidget()
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.list.setStyleSheet("QListWidget{ border:1px dashed #e6eaf2; border-radius:12px; background:#fff; }")
        v.addWidget(self.list)

        self._completer: Optional[QtWidgets.QCompleter] = None

        self.set_suggestions(suggestions or [])

        # --- signals: à¸£à¸­à¸‡à¸£à¸±à¸šà¸—à¸±à¹‰à¸‡ Enter / à¸„à¸¥à¸´à¸à¸„à¸­à¸¡à¹‚à¸š / à¹€à¸¥à¸·à¸­à¸à¸ˆà¸²à¸à¸„à¸­à¸¡à¸žà¸¥à¸µà¸—à¹€à¸•à¸­à¸£à¹Œ ---
        if self.search_line:
            self.search_line.returnPressed.connect(self._add_current)

        self.btn_add.clicked.connect(self._add_current)
        self.btn_persist.clicked.connect(self._persist_current)

        self.list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.list.customContextMenuRequested.connect(self._ctx_menu)
        model = self.list.model()
        model.rowsInserted.connect(lambda *_: self._emit_items_changed())
        model.rowsRemoved.connect(lambda *_: self._emit_items_changed())

    def _ctx_menu(self, pos):
        menu = QtWidgets.QMenu(self)
        a1 = menu.addAction("à¸¥à¸šà¸£à¸²à¸¢à¸à¸²à¸£à¸—à¸µà¹ˆà¹€à¸¥à¸·à¸­à¸")
        a2 = menu.addAction("à¸¥à¸šà¸—à¸±à¹‰à¸‡à¸«à¸¡à¸”")
        act = menu.exec(self.list.mapToGlobal(pos))
        if act == a1:
            for it in self.list.selectedItems():
                self.list.takeItem(self.list.row(it))
        elif act == a2:
            self.list.clear()
        self._emit_items_changed()

    def _add_text(self, text: str):
        if not text:
            return
        current_lower = [self.list.item(i).text().lower().strip() for i in range(self.list.count())]
        if text.lower().strip() not in current_lower:
            self.list.addItem(text)
        self.combo.setCurrentIndex(0)
        self.combo.setEditText("")
        self._emit_items_changed()

    def _add_current(self):
        self._add_text(self.combo.currentText().strip())

    def _persist_current(self):
        text = self.combo.currentText().strip()
        if text:
            self.requestPersist.emit(text)

    def items(self) -> List[str]:
        return [self.list.item(i).text().strip() for i in range(self.list.count())]

    def clear(self):
        self.list.clear()
        self.combo.setCurrentIndex(0)
        self.combo.setEditText("")
        self._emit_items_changed()

    def set_suggestions(self, suggestions):
        seen = set()
        options: List[str] = []
        for value in suggestions or []:
            val = (value or "").strip()
            if not val or val in seen:
                continue
            seen.add(val)
            options.append(val)

        current_text = self.search_line.text() if self.search_line else ""

        self.combo.blockSignals(True)
        self.combo.clear()
        self.combo.addItem("")
        self.combo.addItems(options)
        self.combo.blockSignals(False)

        if self.search_line is not None:
            self.search_line.blockSignals(True)
            self.search_line.setText(current_text)
            self.search_line.setCursorPosition(len(current_text))
            self.search_line.blockSignals(False)

        self._completer = QtWidgets.QCompleter(options)
        self._completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self._completer.setFilterMode(QtCore.Qt.MatchContains)
        self.combo.setCompleter(self._completer)

        # à¸›à¸´à¸”à¸à¸²à¸£à¹€à¸¥à¸·à¹ˆà¸­à¸™à¸”à¹‰à¸§à¸¢à¸¥à¹‰à¸­à¹€à¸¡à¸²à¸ªà¹Œà¸šà¸™à¸„à¸­à¸¡à¹‚à¸š (à¸à¸±à¸™à¹€à¸›à¸¥à¸µà¹ˆà¸¢à¸™à¸„à¹ˆà¸²à¹€à¸§à¸¥à¸²à¹€à¸¥à¸·à¹ˆà¸­à¸™à¸«à¸™à¹‰à¸²)
        self.combo.setFocusPolicy(QtCore.Qt.StrongFocus)

    def _emit_items_changed(self):
        self.itemsChanged.emit(self.items())


def main():
    QLocale.setDefault(QLocale("en_US"))
    app = QtWidgets.QApplication(sys.argv);
    app.setApplicationName("RegistryPatientConnect");
    app.setOrganizationName(ORG_NAME);
    app.setWindowIcon(_load_app_icon())
    ap = argparse.ArgumentParser();
    ap.add_argument("--host", default="127.0.0.1");
    ap.add_argument("--port", type=int, default=DEFAULT_PORT);
    ap.add_argument("--token", default=DEFAULT_TOKEN)
    a = ap.parse_args()
    w = Main(a.host, a.port, a.token);
    w.show();
    sys.exit(app.exec())


if __name__ == "__main__":
    main()